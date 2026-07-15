#!/usr/bin/env python3
"""영업본부 중장기 매출 계획 엑셀 → Supabase longterm_revenue_plan 적재.

시트 '연도별 Booked 매출'의 요약표(B2:H11)만 읽는다. 보조 시트(중장기 DATA_*, 864행 원장)와
비율 행(13~14행: EDI/수주, 전망/수주)은 범위 밖.

레이아웃 (1-indexed):
  B2       환율 기준 문구 1줄 ('Booked 기준 (FX n,nnn원/USD, n,nnn원/EUR)')
  B3       '중장기 계획' / D3 '연도별 매출액 (백만원)'
  D4:H4    전망 연도 5개 — 문자열('2027년')
  B5       기준 1 라벨('26. 1Q', 병합 B5:B7) / 5~7행 계열 3종 / D:H 값
  B9       기준 2 라벨('26. 2Q', 병합 B9:B11) / 9~11행 계열 3종 / D:H 값

'N/A'·공란은 null(2026.1Q의 '고객 EDI 100%'가 전부 N/A).

**월별손익 엑셀과 다른 파일**이므로 sync_management_excel.py 오케스트레이터에 등록하지 않는다.

금액 비노출: 요약은 (기준·계열)별 행수·연도 커버리지·null 카운트만 출력. 금액·합계 출력 금지.
사용자가 직접 실행한다. WriteSession으로 자동 revalidate('longterm_revenue_plan').

사용법
-----
  python scripts/sync_longterm_revenue.py --dry-run
  python scripts/sync_longterm_revenue.py
  python scripts/sync_longterm_revenue.py --revalidate-prod

종료 코드
--------
0 정상
2 헤더/레이아웃 검증 실패
3 엑셀 파일 없음
"""
import argparse
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')
sys.path.insert(0, str(Path(__file__).parent))
from lib.db import WriteSession  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402

SHEET = '연도별 Booked 매출'
TABLE = 'longterm_revenue_plan'
CONFLICT = 'basis_year,basis_quarter,series,period_year'

FX_CELL = (2, 2)            # B2
TITLE_CELL = (3, 2)         # B3
UNIT_CELL = (3, 4)          # D3
YEAR_ROW = 4
COL_BASIS = 2               # B열 — 기준 라벨(병합 셀 좌상단)
COL_SERIES = 3              # C열 — 계열 라벨
COL_FIRST, COL_LAST = 4, 8  # D~H

EXPECTED_TITLE = '중장기 계획'
EXPECTED_UNIT = '연도별 매출액 (백만원)'
SERIES_ORDER = ('수주 Volume', '고객 EDI 100%', '한세 전망')
# 기준 블록: (라벨 셀 행, 계열 시작 행) — 병합 B5:B7 / B9:B11
BASIS_BLOCKS = ((5, 5), (9, 9))
BASIS_RE = re.compile(r'^(\d{2})\.\s*(\d)Q$')


def _txt(v: Any) -> str:
  return '' if v is None else str(v).strip()


def _num(v: Any) -> float | None:
  """숫자 셀만 값으로. 'N/A'·공란·문자열은 None."""
  if v is None or isinstance(v, bool):
    return None
  if isinstance(v, (int, float)):
    return float(v)
  return None


def resolve_excel() -> Path:
  """LONGTERM_EXCEL_PATH env 우선, 없으면 참고/영업계획/*.xlsx 최신."""
  env = os.environ.get('LONGTERM_EXCEL_PATH', '').strip()
  if env:
    p = Path(env)
    if not p.exists():
      raise FileNotFoundError(f'LONGTERM_EXCEL_PATH 파일 없음: {p}')
    return p
  base = Path(__file__).resolve().parent.parent / '참고' / '영업계획'
  cands = [p for p in base.glob('*.xlsx') if not p.name.startswith('~$')]  # 엑셀 잠금 파일 제외
  cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
  if not cands:
    raise FileNotFoundError(f'엑셀 없음: {base}/*.xlsx')
  return cands[0]


def parse_basis(label: str) -> tuple[int, int] | None:
  """'26. 1Q' → (2026, 1). 형식 불일치면 None."""
  m = BASIS_RE.match(label)
  if not m:
    return None
  yy, q = int(m.group(1)), int(m.group(2))
  if not (1 <= q <= 4):
    return None
  return 2000 + yy, q


def parse_year(v: Any) -> int | None:
  """'2027년' → 2027. 숫자 셀도 허용."""
  if isinstance(v, (int, float)) and not isinstance(v, bool):
    return int(v)
  m = re.search(r'(\d{4})', _txt(v))
  return int(m.group(1)) if m else None


def validate_layout(ws) -> list[str]:
  errs = []
  title = _txt(ws.cell(*TITLE_CELL).value)
  if title != EXPECTED_TITLE:
    errs.append(f'  B3: 기대 "{EXPECTED_TITLE}" 실제 "{title}"')
  unit = _txt(ws.cell(*UNIT_CELL).value)
  if unit != EXPECTED_UNIT:
    errs.append(f'  D3: 기대 "{EXPECTED_UNIT}" 실제 "{unit}"')
  for _, start in BASIS_BLOCKS:
    for i, exp in enumerate(SERIES_ORDER):
      actual = _txt(ws.cell(start + i, COL_SERIES).value)
      if actual != exp:
        errs.append(f'  C{start + i}: 기대 "{exp}" 실제 "{actual}"')
  for label_row, _ in BASIS_BLOCKS:
    label = _txt(ws.cell(label_row, COL_BASIS).value)
    if parse_basis(label) is None:
      errs.append(f'  B{label_row}: 기준 라벨 형식 불일치 (기대 "NN. NQ") 실제 "{label}"')
  return errs


def parse_years(ws) -> tuple[list[tuple[int, int]], list[str]]:
  """[(엑셀 열, 연도)], 오류 목록."""
  out, errs = [], []
  for c in range(COL_FIRST, COL_LAST + 1):
    y = parse_year(ws.cell(YEAR_ROW, c).value)
    if y is None:
      errs.append(f'  {YEAR_ROW}행 {c}열: 연도 파싱 실패 "{_txt(ws.cell(YEAR_ROW, c).value)}"')
    else:
      out.append((c, y))
  return out, errs


def build_entries(ws) -> list[dict[str, Any]]:
  fx = _txt(ws.cell(*FX_CELL).value) or None
  years, errs = parse_years(ws)
  if errs:
    raise ValueError('\n'.join(errs))
  entries: list[dict[str, Any]] = []
  for label_row, start in BASIS_BLOCKS:
    parsed = parse_basis(_txt(ws.cell(label_row, COL_BASIS).value))
    if parsed is None:
      continue
    by, bq = parsed
    for i, series in enumerate(SERIES_ORDER):
      for col, year in years:
        entries.append({
          'basis_year': by,
          'basis_quarter': bq,
          'series': series,
          'period_year': year,
          'value_mwon': _num(ws.cell(start + i, col).value),
          'fx_note': fx,
        })
  return entries


def summarize(entries: list[dict[str, Any]]) -> None:
  """(기준·계열)별 행수·연도 커버리지·null 카운트. 금액 비노출."""
  agg = defaultdict(lambda: {'rows': 0, 'years': set(), 'nulls': 0})
  for e in entries:
    k = (f"{e['basis_year']}.{e['basis_quarter']}Q", e['series'])
    agg[k]['rows'] += 1
    agg[k]['years'].add(e['period_year'])
    if e['value_mwon'] is None:
      agg[k]['nulls'] += 1
  logger.info('--- 중장기 매출 전망 요약 (기준·계열) — 금액 비노출 ---')
  for k in sorted(agg.keys()):
    v = agg[k]
    logger.info(f'  {k} | rows={v["rows"]} | years={sorted(v["years"])} | nulls={v["nulls"]}')


def main() -> int:
  ap = argparse.ArgumentParser(description='중장기 매출 계획 엑셀 → Supabase longterm_revenue_plan')
  ap.add_argument('--dry-run', action='store_true', help='실제 upsert 없이 파싱·검증만')
  ap.add_argument('--revalidate-prod', action='store_true',
                  help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                       '로컬 수동 실행 시 프로덕션 stale 방지용')
  args = ap.parse_args()

  try:
    path = resolve_excel()
  except FileNotFoundError as e:
    logger.error(str(e))
    return 3
  logger.info(f'엑셀 로드: {path.name}')

  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    if SHEET not in wb.sheetnames:
      logger.error(f'시트 없음: "{SHEET}" (보유: {wb.sheetnames})')
      return 2
    ws = wb[SHEET]
    errs = validate_layout(ws)
    if errs:
      logger.error(f'[{SHEET}] 레이아웃 불일치:\n' + '\n'.join(errs))
      return 2
    try:
      entries = build_entries(ws)
    except ValueError as e:
      logger.error(f'[{SHEET}] 연도 헤더 오류:\n{e}')
      return 2
  finally:
    wb.close()

  logger.info(f'[{SHEET}] {len(entries)}행 파싱 완료')
  summarize(entries)

  if args.dry_run:
    logger.success('dry-run 완료')
    return 0
  if not entries:
    logger.warning('적재할 행 없음')
    return 0

  with WriteSession() as w:
    w.table(TABLE).upsert(entries, on_conflict=CONFLICT).execute()
  logger.success(f'{TABLE} upsert 완료: {len(entries)}행')
  if args.revalidate_prod:
    revalidate_prod_for_tables([TABLE])
  return 0


if __name__ == '__main__':
  sys.exit(main())
