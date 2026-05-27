#!/usr/bin/env python3
"""기아 차종별/공장별/지역별 판매 엑셀 → kia_sales / kia_export_regions 적재.

플로우 (PR4, 현대차 패턴과 유사하되 JSON API 사용):
  1. https://worldwide.kia.com/api/investors/business-sales-results?language=ko&year={Y}&page={P}
     에서 type='sales' 항목만 추출. 동일 연도가 page=1,2,... 무한 스크롤 형태로 분산.
  2. 각 항목의 files[].title을 NFC normalize → '차종별판매실적'/'해외공장판매실적'/'지역별수출실적' 키워드 매칭.
  3. files[].path → https://worldwide.kia.com/files/{path} 다운로드 (Playwright APIRequestContext).
  4. openpyxl로 3종 엑셀 각각 파싱:
       - **model.xlsx**:   B열 section header(Domestic/Export(excl CKD)/(CKD)) + D열 모델명.
                           Domestic→region='내수', Export(excl CKD)→region='수출', (CKD)→region='CKD' (Aggregate 1행만).
                           'Ordinary Vehicle'/'Special Vehicle' 그룹 합계 row, 'Domestic'/'Export(excl CKD)'/'Total(excl. CKD)' section 합계 row skip.
       - **factory.xlsx**: 'X Plant Sales (Ex-factory)' 섹션 header (B열) + C열 모델명 + E열 Total.
                           5 plant 모두 region='' factory='<plant 표기>' (U.S. Plant/China Plants/Slovakia Plant/Mexico Plant/India Plant).
                           plant 합계 row (B열 단독 + E열 값) skip.
       - **export.xlsx**:  D열 vehicle_type 6개 (Passenger/Recreational/Commercial/Special/CKD(excl Special)/CKD(Special))
                           가 B열 region 합계 row 직전에 나옴. region별로 6 type 적재.
                           'Total' row (B='Total') skip — cross-check 용도로만 사용.
                           ※ 2023년 이전은 5 type 형식 (CKD 1개로 통합) — 그대로 적재.
  5. cross-check (실패 시 abort):
       - model.xlsx: region='내수' SUM == Domestic section 합계 row F열 값
       - model.xlsx: region='수출' SUM == Export(excl CKD) section 합계 row F열 값
       - factory.xlsx: 각 plant 모델 SUM == plant 합계 row E열 값
       - export.xlsx: 각 region (6 vehicle_type SUM) == region 합계 row F열 값
       - export.xlsx: 모든 region SUM == 'Total' row F열 값
  6. WriteSession upsert → 자동 revalidate (oem-kia-sales, oem-kia-export-regions).

플래그:
  --year-from 2021     수집 시작 연도 (default 2021).
  --year-to <year>     수집 마지막 연도 (default 현재 연도).
  --kind {model,factory,export,all}
                       처리할 엑셀 종류 (default all).
  --reprocess-all      캐시(downloads) 무시하고 재다운로드.
  --dry-run            DB 쓰기 없이 파싱 + cross-check 결과만 print.
  --use-cache-only     API/다운로드 skip, data/_kia_downloads/의 파일만 parse (개발/재처리용).
  --keep-downloads     다운로드 파일 보존 (default True).

NFC normalize 필수: 2026년 API 응답이 NFD(자모 분해)로 반환됨 — 매칭 silently 실패 방지.

멱등성: PK upsert. 재실행 시 동일 데이터는 갱신만 됨.

사용:
  scripts/venv/Scripts/python.exe scripts/collect_kia_sales.py \\
    --year-from 2021 --year-to 2026
"""
import argparse
import json
import re
import sys
import time
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger

from lib.bootstrap import init_script

init_script(__file__)

from lib.db import WriteSession  # noqa: E402

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
API_BASE = 'https://worldwide.kia.com'
API_ENDPOINT = '/api/investors/business-sales-results'
PAGE_REFERER = (
  'https://worldwide.kia.com/ko/company/investor-relations/library/'
  'performance-and-plans/'
)

DOWNLOAD_DIR = Path(__file__).resolve().parent.parent / 'data' / '_kia_downloads'
DEFAULT_YEAR_FROM = 2021
RUN_LOG_DIR = Path(__file__).resolve().parent

API_TIMEOUT_S = 30
API_SLEEP_S = 0.2
API_MAX_PAGE = 50          # 페이지당 안전상한
EMPTY_PAGE_LIMIT = 2       # 신규 항목 0인 페이지가 N번 연속이면 종료

DOWNLOAD_TIMEOUT_MS = 30_000

# kind ↔ 한국어 키워드 정규식 (NFC 후 매칭)
# 연도별 표기 차이 흡수:
#   - 2021~2022: '차종별 판매실적' (space), '해외공장별판매실적' (별), '지역별 수출실적' (space)
#   - 2023+:     '차종별판매실적', '해외공장판매실적', '지역별수출실적'
# 공통 키워드만 추출하여 부분 매칭.
_KIND_PATTERNS = {
  'model': re.compile(r'차\s*종\s*별\s*판매실적'),
  'factory': re.compile(r'해외\s*공장(?:별)?\s*판매실적'),
  'export': re.compile(r'지역\s*별\s*수출실적'),
  'retail': re.compile(r'현지\s*판매실적'),
}

# Retail 엑셀 region 헤더 매핑 (R4 컬럼명 → DB region).
# 엑셀: Total + Korea + U.S.A + Canada + Mexico + *Europe + Eastern Europe + Latin America +
#       Middle East + Africa + Asia Pacific + India + China.
# DB region enum은 12개 (Total 제외). 'asterisk *' 제거 + 이름 정규화.
_RETAIL_REGION_ALIAS = {
  'korea': 'Korea',
  'u.s.a': 'U.S.A',
  'u.s.a.': 'U.S.A',
  'usa': 'U.S.A',
  'canada': 'Canada',
  'mexico': 'Mexico',
  'europe': 'Europe',
  '*europe': 'Europe',
  'eastern europe': 'Eastern Europe',
  'eastern\neurope': 'Eastern Europe',
  'latin america': 'Latin America',
  'latin\namerica': 'Latin America',
  'middle east': 'Middle East',
  'middle\neast': 'Middle East',
  'africa': 'Africa',
  'asia pacific': 'Asia Pacific',
  'asia\npacific': 'Asia Pacific',
  'india': 'India',
  'china': 'China',
}

# month sheet 이름 → 월 번호 (1~12).
_RETAIL_MONTH_MAP = {
  'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'june': 6, 'jun': 6,
  'july': 7, 'jul': 7, 'aug': 8, 'sep': 9, 'sept': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}

# 검증 허용 오차 (대 단위) — source 자체의 미세 round/누락 (예: 2024 India Plant
# Carnival r54가 월별 셀에는 255가 있지만 plant E(Total)에 미포함) 허용.
CROSS_CHECK_TOLERANCE = 5
CROSS_CHECK_TOLERANCE_FACTORY = 500  # plant 합계는 source-side 누락이 흔함

# Model 엑셀 section header 토큰 (B열) — 연도별 표기 차이 흡수.
#   2021~2022: 'Domestic', 'Export', 'Total', '(KD)'
#   2023+:     'Domestic', 'Export(excl CKD)', 'Total(excl. CKD)', '(CKD)'
_MODEL_DOMESTIC_LABELS = {'Domestic'}
_MODEL_EXPORT_LABELS = {'Export(excl CKD)', 'Export'}
_MODEL_TOTAL_LABELS = {'Total(excl. CKD)', 'Total'}
_MODEL_CKD_LABELS = {'(CKD)', '(KD)'}

# Model 엑셀 vehicle_type 그룹 합계 row (D열) — skip 대상
_MODEL_GROUP_LABELS = {'Ordinary Vehicle', 'Special Vehicle'}

# Model 엑셀 D열 모델명 prefix (leading space)
_MODEL_LEADING_SPACE_RE = re.compile(r'^\s+')

# Factory 엑셀 section header 매칭 정규식 ('U.S. Plant Sales (Ex-factory)' 등)
_FACTORY_HEADER_RE = re.compile(
  r'^(?P<plant>.+?)\s+Plant\s+Sales\s*\(Ex-factory\)\s*$', re.IGNORECASE
)

# Factory plant 표기 정규화 (audit SQL 표기로 통일)
_FACTORY_PLANT_MAP = {
  'us': 'U.S. Plant',
  'u.s.': 'U.S. Plant',
  'china': 'China Plants',
  'slovakia': 'Slovakia Plant',
  'mexico': 'Mexico Plant',
  'india': 'India Plant',
}

# Factory 헤더 secondary row ('Year YYYY' / 'CY YYYY') — skip
_FACTORY_PERIOD_HEADER_RE = re.compile(r'^(year|cy)\s+\d{4}$', re.IGNORECASE)

# Export 엑셀 vehicle_type (D열) 토큰
_EXPORT_VEHICLE_TYPES_2023 = {'CKD(Inc, Special Vehicle)'}  # 2023 이전 — 1개 통합
_EXPORT_VEHICLE_TYPES_2024 = {  # 2024+ 6개
  'CKD(excl, Special Vehicle)', 'CKD(Special Vehicle)',
}

# Export 엑셀 region SUM row B열 토큰 패턴
_EXPORT_TOTAL_LABEL = 'Total'

# Export region 명칭 normalize (2023-2024: 'US'/'Asia / Pacific', 2025+: 'U.S.'/'Asia/Pacific')
# 표준 표기로 통일 — DB 일관성 확보.
_EXPORT_REGION_ALIAS = {
  'us': 'U.S.',
  'u.s.': 'U.S.',
  'asia / pacific': 'Asia/Pacific',
  'asia/pacific': 'Asia/Pacific',
  'middle east / africa': 'Middle East/Africa',
  'middle east/africa': 'Middle East/Africa',
  'e.europe / cis': 'E.Europe/CIS',
  'e.europe/cis': 'E.Europe/CIS',
}


def _normalize_export_region(name: str) -> str:
  """Export region 표기를 표준으로 normalize."""
  if not name:
    return name
  key = name.strip().lower()
  return _EXPORT_REGION_ALIAS.get(key, name.strip())


def _normalize_retail_region(name: str) -> str | None:
  """Retail 엑셀 region 헤더 → DB region enum. Total/미매칭은 None.
  'Eastern \\nEurope' 같이 줄바꿈+공백 혼합도 single space로 정규화."""
  if not name:
    return None
  key = _nfc(name).lower().strip().lstrip('*').strip()
  key = ' '.join(key.split())  # 연속 whitespace(\n/space) → single space
  if key == 'total':
    return None
  return _RETAIL_REGION_ALIAS.get(key)


def _nfc(s: str | None) -> str:
  """NFC normalize + strip. None은 빈 문자열."""
  if s is None:
    return ''
  return unicodedata.normalize('NFC', str(s)).strip()


def _safe_int(v: Any) -> int:
  if v in (None, ''):
    return 0
  try:
    return int(v)
  except (TypeError, ValueError):
    try:
      return int(float(v))
    except (TypeError, ValueError):
      return 0


# ---------------------------------------------------------------------------
# JSON API → 다운로드 URL 발견
# ---------------------------------------------------------------------------
def _fetch_api_page(request, year: int, page: int) -> list[dict]:
  """Playwright APIRequestContext로 단일 page 응답.

  응답 구조 (관찰):
    list of dict {id, type, year, quarter, month, title, files: [...]}
    or wrapped in {'data': {'list': [...]}}.
  """
  url = f'{API_BASE}{API_ENDPOINT}'
  params = {'language': 'ko', 'year': str(year), 'page': str(page)}
  try:
    resp = request.get(url, params=params, timeout=API_TIMEOUT_S * 1000)
    if resp.status != 200:
      logger.debug(f'  API {year}/p{page}: status={resp.status}')
      return []
    data = resp.json()
  except Exception as e:
    logger.warning(f'  API {year}/p{page} 실패: {e}')
    return []
  if isinstance(data, list):
    return data
  if isinstance(data, dict):
    for key in ('data', 'list', 'items', 'content', 'result', 'rows'):
      if key in data and isinstance(data[key], list):
        return data[key]
      if key in data and isinstance(data[key], dict):
        for sub in ('list', 'items', 'content', 'rows'):
          if sub in data[key] and isinstance(data[key][sub], list):
            return data[key][sub]
  return []


def discover_files_for_year(request, year: int) -> dict[str, dict]:
  """연도별 sales 엑셀 3종 file path 발견 → {kind: file_dict}.

  file_dict = {'title': str, 'path': str, 'fileName': str, 'item': raw_item}
  """
  found: dict[str, dict] = {}
  seen_ids: set[str] = set()
  empty_streak = 0
  for page in range(1, API_MAX_PAGE + 1):
    items = _fetch_api_page(request, year, page)
    if not items:
      break
    new_count = 0
    for it in items:
      iid = str(it.get('id') or '')
      if iid in seen_ids:
        continue
      seen_ids.add(iid)
      new_count += 1
      if _nfc(it.get('type')) != 'sales':
        continue
      # year 필터 (API가 다른 연도 mix 가능성)
      it_year = it.get('year')
      if it_year not in (year, str(year)):
        continue
      for f in (it.get('files') or []):
        title_nfc = _nfc(f.get('title'))
        if not title_nfc:
          continue
        for kind, pat in _KIND_PATTERNS.items():
          if pat.search(title_nfc) and kind not in found:
            found[kind] = {
              'title': title_nfc,
              'path': f.get('path') or '',
              'fileName': f.get('fileName') or '',
              'item': it,
            }
            logger.debug(
              f'    {year} {kind} 발견: {title_nfc} → {f.get("path")}'
            )
            break
    if new_count == 0:
      empty_streak += 1
      if empty_streak >= EMPTY_PAGE_LIMIT:
        break
    else:
      empty_streak = 0
    # 3종 모두 발견했고 첫 페이지면 일찍 종료
    if len(found) == 3:
      break
    time.sleep(API_SLEEP_S)
  return found


def download_excel(request, year: int, kind: str, file_info: dict,
                   dest_dir: Path, reprocess: bool) -> Path | None:
  """파일 path → 다운로드 → 디스크 저장. 캐시 hit이면 다운로드 skip."""
  dest = dest_dir / f'{year}_{kind}.xlsx'
  if dest.exists() and not reprocess:
    logger.info(
      f'  {year}/{kind}: 캐시 hit ({dest.stat().st_size/1024:.0f} KB) — skip download'
    )
    return dest
  path = file_info.get('path') or ''
  if not path:
    logger.warning(f'  {year}/{kind}: path 없음')
    return None
  url = f'{API_BASE}/files/{path}'
  try:
    resp = request.get(url, timeout=DOWNLOAD_TIMEOUT_MS)
    if resp.status != 200:
      logger.warning(f'  {year}/{kind}: 다운로드 status={resp.status}')
      return None
    body = resp.body()
    dest.write_bytes(body)
    logger.info(
      f'  {year}/{kind}: 다운로드 완료 ({dest.stat().st_size/1024:.0f} KB)'
    )
    return dest
  except Exception as e:
    logger.error(f'  {year}/{kind}: 다운로드 실패 — {e}')
    return None


# ---------------------------------------------------------------------------
# 공통 헬퍼 — 헤더 행 탐색
# ---------------------------------------------------------------------------
def _find_jan_col(ws, row: int, min_col: int = 4, max_col: int = 8) -> int | None:
  """주어진 row에서 'Jan' 셀의 col 번호 반환."""
  for c in range(min_col, max_col + 1):
    v = ws.cell(row, c).value
    if v is None:
      continue
    s = _nfc(str(v)).lower().rstrip('.')
    if s == 'jan':
      return c
  return None


def _find_header(ws, max_search_rows: int = 12) -> tuple[int, int] | None:
  """헤더 row + Jan col 탐색. (header_row, jan_col)."""
  for r in range(1, max_search_rows + 1):
    for c in range(4, 12):
      v = ws.cell(r, c).value
      if v is None:
        continue
      s = _nfc(str(v)).lower().rstrip('.')
      if s == 'jan':
        return r, c
  return None


# ---------------------------------------------------------------------------
# 파서 — model.xlsx
# ---------------------------------------------------------------------------
def parse_model_excel(path: Path, year: int, source_url: str) -> tuple[list[dict], dict[str, int]]:
  """차종별판매실적 엑셀 → (rows, section_totals).

  rows: kia_sales upsert 대상 (region 내수/수출/CKD, factory='')
  section_totals: cross-check용 {'domestic': X, 'export': Y, 'ckd': Z}

  엑셀 구조 — section header가 섹션 끝에 위치 (SUM row 역할).
  헤더 r3 다음:
    Domestic 모델 rows (r5~) → ... → 'Ordinary Vehicle' / 'Special Vehicle' 그룹 SUM
    'Domestic' SUM row (B='Domestic', F=section_total)
    Export(excl CKD) 모델 rows → ... → 그룹 SUM
    'Export(excl CKD)' SUM row
    'Total(excl. CKD)' SUM row (skip)
    '(CKD)' SUM row — per-model breakdown 없음, Aggregate 1행 적재
  """
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    ws = wb[wb.sheetnames[0]]
    hdr = _find_header(ws)
    if hdr is None:
      raise ValueError(f'{path.name}: 헤더 미발견')
    header_row, jan_col = hdr
    total_col = jan_col - 1
    total_label = _nfc(ws.cell(header_row, total_col).value)
    if total_label.lower() != 'total':
      logger.warning(
        f'  {path.name}: total_col 추정 실패 (col {total_col}={total_label!r})'
      )

    rows: list[dict] = []
    section_totals: dict[str, int] = {'domestic': 0, 'export': 0, 'ckd': 0}
    # 섹션 헤더가 끝에 나오므로 '내수'를 default region으로 시작.
    # 'Domestic' SUM row를 만나면 region을 '수출'로 전환.
    current_region: str = '내수'
    section_closed: bool = False  # Export(excl CKD) SUM 이후 모델 row는 무시

    for r in range(header_row + 1, ws.max_row + 1):
      b = _nfc(ws.cell(r, 2).value)
      d = _nfc(ws.cell(r, 4).value)
      f_total = _safe_int(ws.cell(r, total_col).value)

      # Section SUM row (B에 토큰, D 비어있음) — 섹션 종료/전환
      if b in _MODEL_DOMESTIC_LABELS and not d:
        section_totals['domestic'] = f_total
        current_region = '수출'  # 다음 섹션은 Export
        continue
      if b in _MODEL_EXPORT_LABELS and not d:
        section_totals['export'] = f_total
        section_closed = True
        current_region = ''
        continue
      if b in _MODEL_TOTAL_LABELS and not d:
        # Total(excl. CKD) / Total — 전체 합계, skip
        continue
      if b in _MODEL_CKD_LABELS and not d:
        # (CKD) — Aggregate 1행 적재 (월별 breakdown 시도)
        section_totals['ckd'] = f_total
        for m_idx in range(12):
          cell_val = ws.cell(r, jan_col + m_idx).value
          if cell_val is None or cell_val == '':
            continue
          units = _safe_int(cell_val)
          if units == 0:
            continue
          rows.append({
            'period_type': 'month',
            'year_period': f'{year}-{m_idx+1:02d}',
            'region': 'CKD',
            'factory': '',
            'vehicle_model': 'Aggregate',
            'vehicle_type': None,
            'sales_units': units,
            'source_url': source_url,
          })
        continue

      # 섹션 종료 이후 row는 적재 안 함
      if section_closed:
        continue

      # vehicle_type group SUM row (D열에 'Ordinary Vehicle' / 'Special Vehicle') — skip
      if d in _MODEL_GROUP_LABELS:
        continue

      # 빈 row
      if not b and not d:
        continue
      if not d:
        continue

      # 모델 row — D열에 모델명 (leading space)
      model_name = _MODEL_LEADING_SPACE_RE.sub('', d).strip()
      if not model_name:
        continue
      # 보호: D에 잘못 들어간 SUM/section 토큰 (드물지만)
      if model_name.lower() in {'total', 'domestic', 'export', 'sub-total'}:
        continue

      for m_idx in range(12):
        cell_val = ws.cell(r, jan_col + m_idx).value
        if cell_val is None or cell_val == '':
          continue
        units = _safe_int(cell_val)
        # 0 cell은 적재 skip (DB 부담 ↓). 음수는 source의 반품/조정 — 적재해야 합계 일치.
        if units == 0:
          continue
        rows.append({
          'period_type': 'month',
          'year_period': f'{year}-{m_idx+1:02d}',
          'region': current_region,
          'factory': '',
          'vehicle_model': model_name,
          'vehicle_type': None,
          'sales_units': units,
          'source_url': source_url,
        })
    return rows, section_totals
  finally:
    wb.close()


# ---------------------------------------------------------------------------
# 파서 — factory.xlsx
# ---------------------------------------------------------------------------
def parse_factory_excel(
  path: Path, year: int, source_url: str,
) -> tuple[list[dict], dict[str, int]]:
  """해외공장판매실적 → (rows, plant_totals).

  rows: kia_sales upsert 대상 (region='', factory='<plant>')
  plant_totals: cross-check용 {plant: total}
  """
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    ws = wb[wb.sheetnames[0]]
    # 헤더는 'CY YYYY' (B열) + 'Total' (E열) + 'Jan' (F열) 패턴.
    # Jan col 탐색 (대부분 F=6, 일부 변형 가능)
    hdr = None
    for r in range(1, 12):
      for c in range(4, 10):
        v = ws.cell(r, c).value
        if v is None:
          continue
        s = _nfc(str(v)).lower().rstrip('.')
        if s == 'jan':
          hdr = (r, c)
          break
      if hdr:
        break
    if hdr is None:
      raise ValueError(f'{path.name}: 헤더 미발견')
    _, jan_col = hdr
    total_col = jan_col - 1

    rows: list[dict] = []
    plant_totals: dict[str, int] = {}
    plant_models: dict[str, int] = {}  # plant별 적재된 모델 합계 (검증용)
    current_plant: str | None = None

    for r in range(1, ws.max_row + 1):
      b = _nfc(ws.cell(r, 2).value)
      c = _nfc(ws.cell(r, 3).value)
      e_total = _safe_int(ws.cell(r, total_col).value)
      e_raw = ws.cell(r, total_col).value
      # Section header: B에 '<X> Plant Sales (Ex-factory)'
      m = _FACTORY_HEADER_RE.match(b)
      if m:
        plant_raw = m.group('plant').strip().lower()
        plant = _FACTORY_PLANT_MAP.get(plant_raw)
        if plant is None:
          logger.warning(
            f'  {path.name} r{r}: 알 수 없는 plant 라벨 {plant_raw!r} — fallback'
          )
          plant = f'{m.group("plant").strip()} Plant'
        current_plant = plant
        plant_models[current_plant] = 0
        continue
      # secondary header ('Year YYYY' / 'CY YYYY')
      if b and _FACTORY_PERIOD_HEADER_RE.match(b):
        continue
      # plant 합계 row: B에 plant 단순 토큰 (KaGA/DYK/US Plant/U.S. Plant/China Plants/...) + E열 값
      # 판별: B 비어있지 않고, C 비어있고, E열에 숫자 값.
      if b and not c and current_plant and e_raw is not None and isinstance(e_raw, (int, float)):
        plant_totals[current_plant] = e_total
        current_plant = None
        continue
      # 빈 row
      if not b and not c:
        continue
      # 모델 row (B 비어있고 C에 모델명)
      if b and not c:
        continue  # 알 수 없는 row 형식 (e.g. 별도 헤더), skip
      if not current_plant:
        continue
      model_name = c
      if not model_name:
        continue
      # 'Total'/'Sub-total' 모델명 보호 (드물지만)
      if model_name.lower() in {'total', 'sub-total'}:
        continue

      added_units = 0
      for m_idx in range(12):
        cell_val = ws.cell(r, jan_col + m_idx).value
        if cell_val is None or cell_val == '':
          continue
        units = _safe_int(cell_val)
        if units == 0:
          continue
        added_units += units
        rows.append({
          'period_type': 'month',
          'year_period': f'{year}-{m_idx+1:02d}',
          'region': '',
          'factory': current_plant,
          'vehicle_model': model_name,
          'vehicle_type': None,
          'sales_units': units,
          'source_url': source_url,
        })
      plant_models[current_plant] = plant_models.get(current_plant, 0) + added_units
    return rows, plant_totals
  finally:
    wb.close()


# ---------------------------------------------------------------------------
# 파서 — export.xlsx
# ---------------------------------------------------------------------------
def parse_export_excel(
  path: Path, year: int, source_url: str,
) -> tuple[list[dict], dict[str, int], int]:
  """지역별수출실적 → (rows, region_totals, grand_total).

  rows: kia_export_regions upsert 대상
  region_totals: {region_name: total} (cross-check용)
  grand_total: 'Total' row 값
  """
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    ws = wb[wb.sheetnames[0]]
    hdr = _find_header(ws)
    if hdr is None:
      raise ValueError(f'{path.name}: 헤더 미발견')
    header_row, jan_col = hdr
    total_col = jan_col - 1

    rows: list[dict] = []
    region_totals: dict[str, int] = {}
    grand_total = 0

    # 누적 type 정보 — region에 도달할 때까지 type 6개 모음
    pending_types: list[dict] = []

    for r in range(header_row + 1, ws.max_row + 1):
      b = _nfc(ws.cell(r, 2).value)
      d = _nfc(ws.cell(r, 4).value)
      f_total = _safe_int(ws.cell(r, total_col).value)

      # 'Total' row — 전체 grand total
      if b.lower() == _EXPORT_TOTAL_LABEL.lower() and not d:
        grand_total = f_total
        # pending이 남아 있으면 (드물게) 버림
        pending_types = []
        continue
      # region SUM row: B 비어있지 않고 D 비어있음
      if b and not d:
        # 헤더 잔재 'Year YYYY' / 'CY YYYY' / 'Export Sales by Region...' 보호
        # (parser는 header_row+1부터 iterate하지만 그래도 안전 필터)
        if (_FACTORY_PERIOD_HEADER_RE.match(b)
            or 'export sales by region' in b.lower()
            or 'sales by model' in b.lower()):
          pending_types = []
          continue
        region_name = _normalize_export_region(b)
        region_totals[region_name] = f_total
        # pending_types를 region에 묶어 적재
        for pt in pending_types:
          for m_idx in range(12):
            cell_val = pt['_cells'][m_idx]
            if cell_val is None or cell_val == '':
              continue
            units = _safe_int(cell_val)
            if units == 0:
              continue
            rows.append({
              'period_type': 'month',
              'year_period': f'{year}-{m_idx+1:02d}',
              'source': 'export-by-region',
              'region_name': region_name,
              'vehicle_type': pt['vehicle_type'],
              'sales_units': units,
              'source_url': source_url,
            })
        pending_types = []
        continue
      # vehicle_type row: D 비어있지 않고 B 비어있음
      if not b and d:
        cells = [ws.cell(r, jan_col + m_idx).value for m_idx in range(12)]
        pending_types.append({'vehicle_type': d, '_cells': cells})
        continue
      # 빈 row 또는 무관 row
      continue

    return rows, region_totals, grand_total
  finally:
    wb.close()


# ---------------------------------------------------------------------------
# 파서 — retail.xlsx (현지판매실적)
# ---------------------------------------------------------------------------
def parse_retail_excel(
  path: Path, year: int, source_url: str,
) -> tuple[list[dict], dict[tuple[str, int], int]]:
  """현지판매실적 → (rows, plant_month_totals).

  rows: kia_retail_sales upsert 대상 (period_type='month').
  plant_month_totals: cross-check용 {(plant, month): plant_total_row의 Total값}.

  엑셀 구조:
   - 13 sheets: Total + Jan~Dec (Total은 cross-check 용도만, 적재 안 함).
   - 각 sheet 컬럼: B 빈/D vehicle_model + F~R 13 region (Total/Korea/U.S.A/.../China).
   - 행:
       Plant section header (B='Korea Plants'/'U.S. Plant'/...) — D 빈, F~R에 plant 합계.
       모델 row (B 빈, D vehicle_model, F~R에 region별 retail).
   - vehicle_model 'Optima / K5' 같이 한국·글로벌 양식 같이 표기 → 그대로 보존.
  """
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    rows: list[dict] = []
    plant_month_totals: dict[tuple[str, int], int] = {}

    for sname in wb.sheetnames:
      key = sname.strip().lower().rstrip('.')
      month_no = _RETAIL_MONTH_MAP.get(key)
      is_total = (key == 'total')
      if not month_no and not is_total:
        logger.debug(f'  {path.name} 알 수 없는 sheet={sname!r} skip')
        continue
      ws = wb[sname]

      # 헤더 row 탐색: 'Model / Plant' 셀 위치 (col 2~5 범위) — 'Model/Plant' 정규화 후 비교.
      header_row = None
      plant_header_col = None
      for r in range(1, 12):
        for c in range(2, 8):
          v = _nfc(ws.cell(r, c).value).lower().replace(' ', '')
          if v == 'model/plant':
            header_row = r
            plant_header_col = c  # Plant section header(예: 'Korea Plants')가 들어가는 col
            break
        if header_row:
          break
      if header_row is None:
        logger.warning(f'  {path.name} sheet={sname}: 헤더 미발견')
        continue

      # region 컬럼 매핑: 'Total' + 12 region 컬럼 매핑.
      region_cols: dict[int, str] = {}
      total_col: int | None = None
      korea_col: int | None = None
      for c in range(plant_header_col + 1, plant_header_col + 25):
        v = _nfc(ws.cell(header_row, c).value)
        if not v:
          continue
        v_norm = v.lower().strip().lstrip('*').strip()
        if v_norm == 'total':
          total_col = c
          continue
        norm = _normalize_retail_region(v)
        if norm:
          region_cols[c] = norm
          if norm == 'Korea':
            korea_col = c

      if not region_cols or not total_col or not korea_col:
        logger.warning(
          f'  {path.name} sheet={sname}: 컬럼 매핑 실패 '
          f'(region={len(region_cols)}, total={total_col}, korea={korea_col})'
        )
        continue

      # 모델명 col 추정: 데이터 row의 plant_header_col 빈 + Korea col 직전 첫 non-empty col.
      # 실제 엑셀: plant_header_col=2 (B), model_col=5 (E), Total=7 (G), Korea=8 (H).
      # 안정 추정: model_col = korea_col - 3 (header에서 'Korea'와 모델명 사이 'Total' + 빈 한 칸).
      model_col = korea_col - 3
      if model_col <= plant_header_col:
        model_col = plant_header_col + 2

      # 엑셀에 2가지 plant section layout 혼재:
      #  - Footer-style: 모델 rows → 그 다음 SUM row(B=plant 이름 + total 숫자).
      #    예: Korea Plants / U.S. Plant / Slovakia Plant / Mexico Plant /
      #         China Plants / India Plant / CKD / Special Vehicle.
      #  - Header-style: B=plant 이름 + total 빈 (SUM 없음) → 다음 행부터 모델 즉시 그 plant로 적재.
      #    예: *Russia Plant (HMMR), HMGICs Plant.
      pending: list[tuple[str, dict[int, Any]]] = []  # footer-style 모델 buffer
      current_header_plant: str | None = None  # header-style 현재 plant

      def _clean_plant_label(lbl: str) -> str:
        # '*Russia Plant\n(HMMR)' → 'Russia Plant', 'Korea Plants' → 'Korea Plants' 등.
        s = lbl.replace('\n', ' ').replace('*', '').strip()
        s = re.sub(r'\s*\([^)]*\)\s*', '', s).strip()  # (HMMR) 등 괄호 안 제거
        s = ' '.join(s.split())
        # 연도별 변형 정규화 (US Plant vs U.S. Plant / China Plant vs China Plants 등).
        canon = {
          'us plant': 'U.S. Plant', 'u.s plant': 'U.S. Plant', 'u.s. plant': 'U.S. Plant',
          'china plant': 'China Plants', 'china plants': 'China Plants',
          'korea plants': 'Korea Plants', 'korea plant': 'Korea Plants',
          'slovakia plant': 'Slovakia Plant',
          'mexico plant': 'Mexico Plant',
          'india plant': 'India Plant',
          'hmgics plant': 'HMGICs Plant',
          'russia plant': 'Russia Plant',
          'ckd': 'CKD',
          'special vehicle': 'Special Vehicle',
        }
        return canon.get(s.lower(), s)

      for r in range(header_row + 1, ws.max_row + 1):
        b = _nfc(ws.cell(r, plant_header_col).value)
        d = _nfc(ws.cell(r, model_col).value)

        total_val = ws.cell(r, total_col).value
        total_is_num = total_val is not None and isinstance(total_val, (int, float))
        # Russia Plant 등 header-style은 total=0으로 비어있어도 0 cell이라 isinstance 통과.
        # → total 값이 실제로 0이면 footer로 보지 않고 header로 처리.
        total_nonzero = total_is_num and _safe_int(total_val) != 0

        # Plant footer row (footer-style): B=plant + model_col 빈 + total 숫자(>0)
        if b and not d and total_nonzero:
          plant_label = _clean_plant_label(b)
          if plant_label.lower() in {'total', 'sub-total'}:
            pending = []
            current_header_plant = None
            continue
          plant_total = _safe_int(total_val)
          if not is_total:
            plant_month_totals[(plant_label, month_no)] = plant_total
          for model_name, cells in pending:
            for col, region in region_cols.items():
              cv = cells.get(col)
              if cv is None or cv == '':
                continue
              units = _safe_int(cv)
              if units == 0:
                continue
              rows.append({
                'period_type': 'month' if not is_total else 'annual',
                'year_period': (
                  f'{year}-{month_no:02d}' if not is_total else str(year)
                ),
                'plant': plant_label,
                'vehicle_model': model_name,
                'region': region,
                'retail_units': units,
                'source_url': source_url,
              })
          pending = []
          current_header_plant = None
          continue

        # Plant header row (header-style): B=plant + model_col 빈 + total 빈/0
        if b and not d and not total_nonzero:
          # leading '*' sub-header (예: '*Russia Plant (HMMR)')는 다음 footer의 sub-section.
          # 이 경우 header-style 아니라 그냥 무시 (pending 유지) — China Plants footer에서 묶임.
          if b.lstrip().startswith('*'):
            continue
          current_header_plant = _clean_plant_label(b)
          continue

        # 모델 row: model_col에 모델명, plant_header_col 빈
        if d and not b:
          model_name = d.strip()
          if model_name.lower() in {'total', 'sub-total', 'bev total'}:
            continue
          # header-style이면 즉시 적재
          if current_header_plant:
            for col, region in region_cols.items():
              cv = ws.cell(r, col).value
              if cv is None or cv == '':
                continue
              units = _safe_int(cv)
              if units == 0:
                continue
              rows.append({
                'period_type': 'month' if not is_total else 'annual',
                'year_period': (
                  f'{year}-{month_no:02d}' if not is_total else str(year)
                ),
                'plant': current_header_plant,
                'vehicle_model': model_name,
                'region': region,
                'retail_units': units,
                'source_url': source_url,
              })
            continue
          # footer-style이면 pending에 buffer
          cells: dict[int, Any] = {}
          for col in region_cols:
            cells[col] = ws.cell(r, col).value
          pending.append((model_name, cells))
          continue

        # 빈 row / 그 외 — skip
        continue

    return rows, plant_month_totals
  finally:
    wb.close()


# ---------------------------------------------------------------------------
# Cross-check (실패 시 abort)
# ---------------------------------------------------------------------------
class CrossCheckError(Exception):
  pass


def check_model_totals(
  rows: list[dict], totals: dict[str, int], year: int,
) -> list[str]:
  """model 적재 결과 vs section 합계 sanity check. 실패 메시지 리스트 반환."""
  fails: list[str] = []
  by_region = {'내수': 0, '수출': 0, 'CKD': 0}
  for r in rows:
    if r.get('factory'):
      continue
    by_region[r['region']] = by_region.get(r['region'], 0) + r['sales_units']
  pairs = [
    ('내수', 'domestic'),
    ('수출', 'export'),
    ('CKD', 'ckd'),
  ]
  for r_key, t_key in pairs:
    expected = totals.get(t_key, 0)
    actual = by_region.get(r_key, 0)
    if abs(expected - actual) > CROSS_CHECK_TOLERANCE:
      fails.append(
        f'  model {year} region={r_key}: 적재={actual:,} vs section 합계={expected:,} (차={actual-expected:+,})'
      )
  return fails


def check_factory_totals(
  rows: list[dict], totals: dict[str, int], year: int,
) -> list[str]:
  fails: list[str] = []
  by_factory: dict[str, int] = {}
  for r in rows:
    f = r.get('factory') or ''
    if not f:
      continue
    by_factory[f] = by_factory.get(f, 0) + r['sales_units']
  for plant, expected in totals.items():
    actual = by_factory.get(plant, 0)
    diff = abs(expected - actual)
    if diff > CROSS_CHECK_TOLERANCE_FACTORY:
      fails.append(
        f'  factory {year} plant={plant}: 적재={actual:,} vs plant 합계={expected:,} (차={actual-expected:+,})'
      )
    elif diff > CROSS_CHECK_TOLERANCE:
      logger.info(
        f'  factory {year} plant={plant}: source-side 미세 불일치 무시 '
        f'(적재={actual:,} vs plant 합계={expected:,}, 차={actual-expected:+,})'
      )
  return fails


def check_retail_totals(
  rows: list[dict], plant_month_totals: dict[tuple[str, int], int], year: int,
) -> list[str]:
  """retail 적재 행의 (plant, month) 합 vs 엑셀 plant SUM row total 비교."""
  fails: list[str] = []
  by_plant_month: dict[tuple[str, int], int] = {}
  for r in rows:
    if r.get('period_type') != 'month':
      continue
    plant = r['plant']
    mm = int(r['year_period'][-2:])
    by_plant_month[(plant, mm)] = by_plant_month.get((plant, mm), 0) + r['retail_units']
  for (plant, mm), expected in plant_month_totals.items():
    actual = by_plant_month.get((plant, mm), 0)
    if abs(expected - actual) > CROSS_CHECK_TOLERANCE_FACTORY:
      fails.append(
        f'  retail {year}-{mm:02d} plant={plant}: '
        f'적재={actual:,} vs plant 합계={expected:,} (차={actual-expected:+,})'
      )
  return fails


def check_export_totals(
  rows: list[dict], region_totals: dict[str, int], grand_total: int, year: int,
) -> list[str]:
  fails: list[str] = []
  by_region: dict[str, int] = {}
  for r in rows:
    rn = r['region_name']
    by_region[rn] = by_region.get(rn, 0) + r['sales_units']
  for region, expected in region_totals.items():
    actual = by_region.get(region, 0)
    if abs(expected - actual) > CROSS_CHECK_TOLERANCE:
      fails.append(
        f'  export {year} region={region}: 적재={actual:,} vs region 합계={expected:,} (차={actual-expected:+,})'
      )
  total_actual = sum(by_region.values())
  if grand_total and abs(grand_total - total_actual) > CROSS_CHECK_TOLERANCE:
    fails.append(
      f'  export {year} GRAND TOTAL: 적재={total_actual:,} vs Total row={grand_total:,} (차={total_actual-grand_total:+,})'
    )
  return fails


# ---------------------------------------------------------------------------
# 요약
# ---------------------------------------------------------------------------
def _summary_sales(rows: list[dict]) -> dict:
  out: dict[str, Any] = {
    'total_rows': len(rows),
    'by_year': {},
    'by_region': {},
    'by_factory': {},
  }
  for r in rows:
    y = r['year_period'][:4]
    out['by_year'][y] = out['by_year'].get(y, 0) + r['sales_units']
    out['by_region'][r['region']] = out['by_region'].get(r['region'], 0) + r['sales_units']
    fk = r.get('factory') or '(국내)'
    out['by_factory'][fk] = out['by_factory'].get(fk, 0) + r['sales_units']
  return out


def _summary_export(rows: list[dict]) -> dict:
  out: dict[str, Any] = {
    'total_rows': len(rows),
    'by_year': {},
    'by_region': {},
    'by_type': {},
  }
  for r in rows:
    y = r['year_period'][:4]
    out['by_year'][y] = out['by_year'].get(y, 0) + r['sales_units']
    out['by_region'][r['region_name']] = out['by_region'].get(r['region_name'], 0) + r['sales_units']
    out['by_type'][r['vehicle_type']] = out['by_type'].get(r['vehicle_type'], 0) + r['sales_units']
  return out


def _summary_retail(rows: list[dict]) -> dict:
  out: dict[str, Any] = {
    'total_rows': len(rows),
    'by_year': {},
    'by_plant': {},
    'by_region': {},
  }
  for r in rows:
    y = r['year_period'][:4]
    out['by_year'][y] = out['by_year'].get(y, 0) + r['retail_units']
    out['by_plant'][r['plant']] = out['by_plant'].get(r['plant'], 0) + r['retail_units']
    out['by_region'][r['region']] = out['by_region'].get(r['region'], 0) + r['retail_units']
  return out


# ---------------------------------------------------------------------------
# Dedupe
# ---------------------------------------------------------------------------
def _dedupe_sales(rows: list[dict]) -> list[dict]:
  """동일 PK 충돌 시 sales_units **합산** (Tasman/Bongo가 Ordinary+Special 양쪽 그룹에 등장 — 별도 차량 변종이지만 모델명 동일).

  PK 기준: (period_type, year_period, region, factory, vehicle_model).
  vehicle_type은 첫 번째 발견된 값을 유지 (NULL/Ordinary/Special).
  """
  by_pk: dict[tuple, dict] = {}
  for r in rows:
    pk = (r['period_type'], r['year_period'], r['region'], r['factory'], r['vehicle_model'])
    existing = by_pk.get(pk)
    if existing is None:
      by_pk[pk] = dict(r)
    else:
      existing['sales_units'] = existing.get('sales_units', 0) + r.get('sales_units', 0)
  return list(by_pk.values())


def _dedupe_export(rows: list[dict]) -> list[dict]:
  by_pk: dict[tuple, dict] = {}
  for r in rows:
    pk = (
      r['period_type'], r['year_period'], r['source'], r['region_name'], r['vehicle_type']
    )
    by_pk[pk] = r
  return list(by_pk.values())


def _dedupe_retail(rows: list[dict]) -> list[dict]:
  """동일 PK 충돌 시 retail_units 합산.
  PK: (period_type, year_period, plant, vehicle_model, region)."""
  by_pk: dict[tuple, dict] = {}
  for r in rows:
    pk = (r['period_type'], r['year_period'], r['plant'], r['vehicle_model'], r['region'])
    existing = by_pk.get(pk)
    if existing is None:
      by_pk[pk] = dict(r)
    else:
      existing['retail_units'] = existing.get('retail_units', 0) + r.get('retail_units', 0)
  return list(by_pk.values())


# ---------------------------------------------------------------------------
# 메인 파이프라인
# ---------------------------------------------------------------------------
def _process_year_kind(
  year: int, kind: str, path: Path,
) -> tuple[list[dict], list[dict], list[dict], list[str]]:
  """kind별 parse + cross-check. (sales_rows, export_rows, retail_rows, check_fails) 반환."""
  source_url = PAGE_REFERER
  fails: list[str] = []
  sales_rows: list[dict] = []
  export_rows: list[dict] = []
  retail_rows: list[dict] = []
  if kind == 'model':
    try:
      rows, totals = parse_model_excel(path, year, source_url)
    except Exception as e:
      logger.error(f'  {year}/model 파싱 실패: {e}')
      return [], [], [], [f'{year}/model/parse: {e}']
    fails.extend(check_model_totals(rows, totals, year))
    sales_rows.extend(rows)
  elif kind == 'factory':
    try:
      rows, totals = parse_factory_excel(path, year, source_url)
    except Exception as e:
      logger.error(f'  {year}/factory 파싱 실패: {e}')
      return [], [], [], [f'{year}/factory/parse: {e}']
    fails.extend(check_factory_totals(rows, totals, year))
    sales_rows.extend(rows)
  elif kind == 'export':
    try:
      rows, region_totals, grand_total = parse_export_excel(path, year, source_url)
    except Exception as e:
      logger.error(f'  {year}/export 파싱 실패: {e}')
      return [], [], [], [f'{year}/export/parse: {e}']
    fails.extend(check_export_totals(rows, region_totals, grand_total, year))
    export_rows.extend(rows)
  elif kind == 'retail':
    try:
      rows, plant_month_totals = parse_retail_excel(path, year, source_url)
    except Exception as e:
      logger.error(f'  {year}/retail 파싱 실패: {e}')
      return [], [], [], [f'{year}/retail/parse: {e}']
    fails.extend(check_retail_totals(rows, plant_month_totals, year))
    retail_rows.extend(rows)
  logger.info(
    f'  {year}/{kind}: sales={len(sales_rows)}, export={len(export_rows)}, '
    f'retail={len(retail_rows)} parsed, check_fails={len(fails)}'
  )
  return sales_rows, export_rows, retail_rows, fails


def _resolve_kinds(arg: str) -> list[str]:
  if arg == 'all':
    return ['model', 'factory', 'export', 'retail']
  return [arg]


def _run_api_pipeline(
  args, year_range: list[int],
) -> tuple[list[dict], list[dict], list[dict], list[str], list[Path]]:
  kinds = _resolve_kinds(args.kind)
  all_sales: list[dict] = []
  all_export: list[dict] = []
  all_retail: list[dict] = []
  failed_jobs: list[str] = []
  downloaded: list[Path] = []

  from playwright.sync_api import sync_playwright  # noqa: E402

  with sync_playwright() as pw:
    request = pw.request.new_context(
      extra_http_headers={
        'User-Agent': (
          'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
          '(KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36'
        ),
        'Accept': 'application/json, application/octet-stream, */*',
        'Accept-Language': 'ko-KR,ko;q=0.9,en;q=0.8',
        'Referer': PAGE_REFERER,
      }
    )
    try:
      for year in year_range:
        logger.info(f'>>> {year} files 발견')
        found = discover_files_for_year(request, year)
        if not found:
          logger.warning(f'  {year}: sales 파일 없음 — skip')
          failed_jobs.append(f'{year}/no-sales')
          continue
        for kind in kinds:
          file_info = found.get(kind)
          if not file_info:
            logger.warning(f'  {year}/{kind}: API에 해당 파일 없음')
            failed_jobs.append(f'{year}/{kind}/missing')
            continue
          path = download_excel(
            request, year, kind, file_info, DOWNLOAD_DIR, args.reprocess_all,
          )
          if path is None:
            failed_jobs.append(f'{year}/{kind}/download')
            continue
          downloaded.append(path)
          sales_rows, export_rows, retail_rows, fails = _process_year_kind(year, kind, path)
          all_sales.extend(sales_rows)
          all_export.extend(export_rows)
          all_retail.extend(retail_rows)
          if fails:
            failed_jobs.extend(fails)
    finally:
      request.dispose()
  return all_sales, all_export, all_retail, failed_jobs, downloaded


def _run_cache_only_pipeline(
  args, year_range: list[int],
) -> tuple[list[dict], list[dict], list[dict], list[str], list[Path]]:
  """API/다운로드 skip — 캐시된 .xlsx만 parse."""
  kinds = _resolve_kinds(args.kind)
  all_sales: list[dict] = []
  all_export: list[dict] = []
  all_retail: list[dict] = []
  failed_jobs: list[str] = []
  used: list[Path] = []
  for year in year_range:
    for kind in kinds:
      path = DOWNLOAD_DIR / f'{year}_{kind}.xlsx'
      if not path.exists():
        logger.warning(f'  {year}/{kind}: 캐시 파일 없음 — {path}')
        failed_jobs.append(f'{year}/{kind}/no-cache')
        continue
      used.append(path)
      sales_rows, export_rows, retail_rows, fails = _process_year_kind(year, kind, path)
      all_sales.extend(sales_rows)
      all_export.extend(export_rows)
      all_retail.extend(retail_rows)
      if fails:
        failed_jobs.extend(fails)
  return all_sales, all_export, all_retail, failed_jobs, used


def parse_args() -> argparse.Namespace:
  p = argparse.ArgumentParser(description='기아 차종/공장/지역 판매 수집.')
  p.add_argument('--year-from', type=int, default=DEFAULT_YEAR_FROM,
                 help=f'백필 시작 연도 (default {DEFAULT_YEAR_FROM})')
  p.add_argument('--year-to', type=int, default=None,
                 help='마지막 연도 (default 현재 연도)')
  p.add_argument('--kind', choices=['model', 'factory', 'export', 'retail', 'all'],
                 default='all', help='처리할 엑셀 종류 (default all = model+factory+export+retail)')
  p.add_argument('--reprocess-all', action='store_true',
                 help='캐시(downloads) 무시하고 재다운로드')
  p.add_argument('--dry-run', action='store_true',
                 help='DB 쓰기 없이 파싱 + cross-check 결과만 print')
  p.add_argument('--use-cache-only', action='store_true',
                 help='API/다운로드 skip — data/_kia_downloads/의 파일만 parse')
  p.add_argument('--keep-downloads', action='store_true', default=True,
                 help='다운로드 파일 보존 (default True)')
  p.add_argument('--abort-on-check-fail', action='store_true', default=True,
                 help='cross-check 실패 시 DB 쓰기 중단 (default True)')
  p.add_argument('--no-abort', dest='abort_on_check_fail', action='store_false',
                 help='cross-check 실패 시에도 적재 강행 (긴급용)')
  return p.parse_args()


def main() -> int:
  args = parse_args()
  current_year = datetime.now(timezone.utc).year
  year_to = args.year_to or current_year
  year_range = list(range(args.year_from, year_to + 1))
  logger.info(
    f'기아 판매 수집: 연도 {year_range[0]}~{year_range[-1]} kind={args.kind} '
    f'(dry_run={args.dry_run}, cache_only={args.use_cache_only}, reprocess={args.reprocess_all})'
  )

  DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

  if args.use_cache_only:
    all_sales, all_export, all_retail, failed_jobs, used = _run_cache_only_pipeline(args, year_range)
  else:
    all_sales, all_export, all_retail, failed_jobs, used = _run_api_pipeline(args, year_range)

  all_sales = _dedupe_sales(all_sales)
  all_export = _dedupe_export(all_export)
  all_retail = _dedupe_retail(all_retail)
  sum_sales = _summary_sales(all_sales)
  sum_export = _summary_export(all_export)
  sum_retail = _summary_retail(all_retail)
  logger.info(f'[kia_sales] {sum_sales}')
  logger.info(f'[kia_export_regions] {sum_export}')
  logger.info(f'[kia_retail_sales] {sum_retail}')
  if failed_jobs:
    check_failures = [j for j in failed_jobs if 'plant=' in j or 'region=' in j]
    other_failures = [j for j in failed_jobs if j not in check_failures]
    if check_failures:
      logger.warning(f'Cross-check 실패 ({len(check_failures)}건):')
      for cf in check_failures:
        logger.warning(cf)
    if other_failures:
      logger.warning(f'기타 실패 ({len(other_failures)}건): {other_failures}')

  # 결과 JSON 저장
  ts = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
  log_path = RUN_LOG_DIR / f'_kia_collect_run_{ts}.json'
  try:
    log_payload = {
      'started_at': ts,
      'args': vars(args),
      'year_range': year_range,
      'sales_summary': sum_sales,
      'export_summary': sum_export,
      'retail_summary': sum_retail,
      'failed_jobs': failed_jobs,
    }
    with log_path.open('w', encoding='utf-8') as f:
      json.dump(log_payload, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f'결과 로그: {log_path}')
  except Exception as e:
    logger.warning(f'결과 로그 저장 실패: {e}')

  has_check_fail = any('plant=' in j or 'region=' in j for j in failed_jobs)
  if args.dry_run:
    logger.success('dry-run 종료 (DB 쓰기 없음)')
    return 1 if failed_jobs else 0

  if has_check_fail and args.abort_on_check_fail:
    logger.error('cross-check 실패 — DB 쓰기 중단 (--no-abort로 강행 가능)')
    return 2

  if not all_sales and not all_export and not all_retail:
    logger.warning('적재할 행 없음 — DB 호출 생략')
    return 1 if failed_jobs else 0

  try:
    with WriteSession() as w:
      if all_sales:
        # 대용량 batch 안전: 500개씩
        BATCH = 500
        for i in range(0, len(all_sales), BATCH):
          chunk = all_sales[i:i + BATCH]
          w.table('kia_sales').upsert(
            chunk,
            on_conflict='period_type,year_period,region,factory,vehicle_model',
          ).execute()
        logger.success(f'kia_sales upsert 완료: {len(all_sales)}행')
      if all_export:
        BATCH = 500
        for i in range(0, len(all_export), BATCH):
          chunk = all_export[i:i + BATCH]
          w.table('kia_export_regions').upsert(
            chunk,
            on_conflict='period_type,year_period,source,region_name,vehicle_type',
          ).execute()
        logger.success(f'kia_export_regions upsert 완료: {len(all_export)}행')
      if all_retail:
        BATCH = 500
        for i in range(0, len(all_retail), BATCH):
          chunk = all_retail[i:i + BATCH]
          w.table('kia_retail_sales').upsert(
            chunk,
            on_conflict='period_type,year_period,plant,vehicle_model,region',
          ).execute()
        logger.success(f'kia_retail_sales upsert 완료: {len(all_retail)}행')
  except Exception as e:
    logger.exception(f'upsert 실패: {e}')
    return 2

  return 0 if not failed_jobs else 1


if __name__ == '__main__':
  try:
    sys.exit(main())
  except KeyboardInterrupt:
    logger.warning('사용자 중단')
    sys.exit(130)
  except Exception as e:
    logger.exception(f'예기치 못한 오류: {e}')
    sys.exit(1)
