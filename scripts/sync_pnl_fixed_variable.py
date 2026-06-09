#!/usr/bin/env python3
"""손익 엑셀의 '고정비' 시트 → Supabase pnl_fixed_variable 적재.

고정비 시트 헤더(1행, 8열):
  연도 | 연간/월 | 고정/변동 | 계정분류1 | 계정분류2 | 계정분류3 | 계정명 | 밸류(백만원)

데이터 2행부터 시작.
- 연도 '기준'(변동비율 기준행)·계정분류1 '매출' 행은 cost_type 필터로 자연 제외.
- 밸류(백만원)는 현재 비어있을 수 있어, None이면 SKIP(값 입력 후 재실행 대비).
"""
import argparse
import sys
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')

sys.path.insert(0, str(Path(__file__).parent))
from lib.db import get_client, upsert_rows  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402

def _latest_excel() -> Path:
  base = Path(__file__).resolve().parents[1] / '참고' / '손익'
  cands = sorted(base.glob('자료정리_월별손익*.xlsx'))
  if not cands:
    raise FileNotFoundError(f'손익 엑셀 없음: {base}/자료정리_월별손익*.xlsx')
  return cands[-1]


EXCEL_PATH = _latest_excel()
TABLE = 'pnl_fixed_variable'
SHEET = '고정비'
CONFLICT_COLS = 'period_year,period_kind,period_month,cost_type,category2,category3,account'
DATA_START_ROW = 2

# 고정/변동 열(C) 화이트리스트. '변동비율'(기준행)·None은 통과 안 함 → 매출·기준행 자연 제외.
# DB CHECK 제약(cost_type IN ('고정비','변동비'))·UI 필터와 일치해야 하므로 한글 값을 그대로 저장.
COST_TYPE_MAP = {'고정비': '고정비', '변동비': '변동비'}

# 헤더 8열 기대값 (A~H = row[0]~row[7])
EXPECTED_HEADER = ['연도', '연간/월', '고정/변동', '계정분류1', '계정분류2', '계정분류3', '계정명', '밸류(백만원)']


def parse_period(raw_period: Any) -> tuple[str, int] | None:
  """'연간' → ('annual', 0). 1..12 → ('monthly', N)."""
  if raw_period is None:
    return None
  if isinstance(raw_period, (int, float)) and not isinstance(raw_period, bool):
    m = int(raw_period)
    if 1 <= m <= 12:
      return ('monthly', m)
    return None
  s = str(raw_period).strip()
  if s == '연간':
    return ('annual', 0)
  return None


def parse_year(raw_year: Any) -> int | None:
  """연도 열(A) 파싱. 정수형(2023 등)만 통과, '기준' 등 문자열은 SKIP."""
  if isinstance(raw_year, (int, float)) and not isinstance(raw_year, bool):
    return int(raw_year)
  return None


def parse_rows(ws: Any) -> tuple[list[dict[str, Any]], int]:
  """시트를 순회해 적재 대상 행 dict 목록과 SKIP 수를 반환한다.

  세 종류를 적재한다:
    1. 기준 변동비율 (C='변동비율'): period_year=0, cost_type='변동비율', value_mwon=비율(0~1).
       연도 무관 가정치라 '기준' 행에서만 나온다.
    2. 매출 (계정분류1='매출'): cost_type='매출', category2/3/account='매출' 센티넬.
    3. 비용 (C='고정비'|'변동비'): 계정명별 금액.
  값(H)이 비어 있으면 SKIP(값 입력 후 재실행 대비)."""
  rows: list[dict[str, Any]] = []
  skipped = 0
  for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
    raw_year, raw_period, raw_cost = row[0], row[1], row[2]
    cat1, category2, category3, account, value = row[3], row[4], row[5], row[6], row[7]
    cost_label = str(raw_cost).strip() if raw_cost is not None else ''

    # 1) 기준 변동비율 — 연도 무관(period_year=0). '기준' 매출 행은 값이 없어 자연 SKIP.
    if cost_label == '변동비율':
      if value is None or category2 is None or category3 is None or account is None:
        skipped += 1
        continue
      rows.append({
        'period_year': 0,
        'period_kind': 'annual',
        'period_month': 0,
        'cost_type': '변동비율',
        'category2': str(category2).strip(),
        'category3': str(category3).strip(),
        'account': str(account).strip(),
        'value_mwon': float(value),
      })
      continue

    # 2)·3)은 실연도 행만 ('기준' SKIP).
    year = parse_year(raw_year)
    if year is None:
      skipped += 1
      continue
    period = parse_period(raw_period)
    if period is None:
      skipped += 1
      continue

    if cat1 is not None and str(cat1).strip() == '매출':
      # 2) 매출 행 — 고정/변동 구분 없음. 센티넬로 저장.
      cost_type, c2, c3, acc = '매출', '매출', '매출', '매출'
    else:
      # 3) 비용 행.
      cost_type = COST_TYPE_MAP.get(cost_label)
      if cost_type is None:  # 그 외(None 등) → SKIP
        skipped += 1
        continue
      c2, c3, acc = str(category2).strip(), str(category3).strip(), str(account).strip()

    if value is None:  # H 미입력 → SKIP
      skipped += 1
      continue

    rows.append({
      'period_year': year,
      'period_kind': period[0],
      'period_month': period[1],
      'cost_type': cost_type,
      'category2': c2,
      'category3': c3,
      'account': acc,
      'value_mwon': float(value),
    })
  return rows, skipped


def check_consistency(rows: list[dict[str, Any]]) -> None:
  """고정비 시트 매출·파생 영업이익을 DB pnl_cost_structure 저장값과 대조.

  사외비 정책: 금액은 로그에 찍지 않고 mismatch 연도 수만 보고. 임계 0.5%.
  매출 행이 없으면(값 미입력) 검증 대상이 없어 조용히 종료."""
  # 시트 집계: 연도별 매출 / 비용합계(고정+변동). period_year=0(기준)·변동비율 제외.
  sheet_rev: dict[int, float] = {}
  sheet_cost: dict[int, float] = {}
  ytd_max = 0
  for r in rows:
    y, ck = r['period_year'], r['cost_type']
    if y == 0 or ck == '변동비율':
      continue
    if y == 2026 and r['period_kind'] == 'monthly':
      ytd_max = max(ytd_max, r['period_month'])
    if ck == '매출':
      sheet_rev[y] = sheet_rev.get(y, 0.0) + (r['value_mwon'] or 0.0)
    elif ck in ('고정비', '변동비'):
      sheet_cost[y] = sheet_cost.get(y, 0.0) + (r['value_mwon'] or 0.0)

  if not sheet_rev:
    logger.info('정합성: 매출 데이터 없음 — 검증 생략')
    return

  # DB pnl_cost_structure의 매출·영업이익(actual) 조회.
  try:
    resp = (get_client().table('pnl_cost_structure')
            .select('period_year,period_kind,period_month,account,value_mwon')
            .in_('account', ['매출', '영업이익']).eq('kind', 'actual').execute())
    db_rows = resp.data or []
  except Exception as e:
    logger.warning(f'정합성: DB(pnl_cost_structure) 조회 실패 — 검증 생략 ({e})')
    return

  def db_value(year: int, account: str) -> float | None:
    total, found = 0.0, False
    for d in db_rows:
      if d['account'] != account or d['period_year'] != year:
        continue
      if year == 2026:
        if d['period_kind'] != 'monthly' or not (1 <= d['period_month'] <= ytd_max):
          continue
      elif d['period_kind'] != 'annual':
        continue
      total += d['value_mwon'] or 0.0
      found = True
    return total if found else None

  def mismatch(a: float, b: float | None) -> bool:
    if b is None:
      return False  # 비교 대상 없음 → mismatch로 치지 않음
    if b == 0:
      return abs(a) > 1e-6
    return abs(a - b) / abs(b) > 0.005

  rev_bad, op_bad = 0, 0
  for y in sorted(sheet_rev):
    if mismatch(sheet_rev[y], db_value(y, '매출')):
      rev_bad += 1
    sheet_op = sheet_rev.get(y, 0.0) - sheet_cost.get(y, 0.0)
    if mismatch(sheet_op, db_value(y, '영업이익')):
      op_bad += 1

  if rev_bad or op_bad:
    logger.warning(f'정합성 불일치(임계 0.5%): 매출 {rev_bad}개 연도, 영업이익 {op_bad}개 연도')
  else:
    logger.success('정합성 OK: 매출·영업이익이 DB와 일치(임계 0.5%)')


def main() -> int:
  parser = argparse.ArgumentParser(description='고정비 시트 → Supabase pnl_fixed_variable 적재')
  parser.add_argument('--revalidate-prod', action='store_true',
                      help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                           '로컬 수동 실행 시 프로덕션 stale 방지용')
  parser.add_argument('--dry-run', action='store_true',
                      help='파싱·행수 집계까지만 수행하고 DB 적재는 건너뛴다 (검증용)')
  args = parser.parse_args()

  if not EXCEL_PATH.exists():
    logger.error(f'엑셀 파일 없음: {EXCEL_PATH}')
    return 1

  logger.info(f'엑셀 로드: {EXCEL_PATH}')
  # 메모리 메모: read_only=True 단독 인덱싱 오정렬 사례 → read_only=False로 안전하게.
  wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=False)
  try:
    ws = wb[SHEET]
    # 헤더 검증 (8열, A~H)
    actual = [str(ws.cell(1, c).value or '').strip() for c in range(1, 9)]
    if actual != EXPECTED_HEADER:
      logger.error(f'[{SHEET}] 헤더 불일치: 기대 {EXPECTED_HEADER}, 실제 {actual}')
      return 2

    rows, skipped = parse_rows(ws)
  finally:
    wb.close()

  ratio_n = sum(1 for r in rows if r['cost_type'] == '변동비율')
  rev_n = sum(1 for r in rows if r['cost_type'] == '매출')
  cost_n = sum(1 for r in rows if r['cost_type'] in ('고정비', '변동비'))
  years = sorted({r['period_year'] for r in rows if r['period_year'] != 0})
  logger.info(f'적재 대상 {len(rows)}행 (SKIP {skipped}) — '
              f'변동비율 {ratio_n} / 매출 {rev_n} / 비용 {cost_n}, 연도 {years}')

  # 매출·영업이익 DB 정합성 체크 (금액 비노출, mismatch 연도 수만).
  check_consistency(rows)

  if not rows:
    # 현재 비용 금액이 전부 비어있으면 비용/매출 0행이 정상(변동비율만 적재될 수 있음).
    logger.info('적재 대상 0행 — DB 변경 없이 종료')
    return 0

  if args.dry_run:
    logger.info('[dry-run] 파싱까지만 수행, DB 적재 건너뜀')
    return 0

  try:
    n = upsert_rows(TABLE, rows, conflict_cols=CONFLICT_COLS)
    logger.success(f'pnl_fixed_variable upsert 완료: {n}행')
    if args.revalidate_prod:
      revalidate_prod_for_tables([TABLE])
    return 0
  except Exception as e:
    logger.error(f'upsert 실패: {e}')
    return 3


if __name__ == '__main__':
  sys.exit(main())
