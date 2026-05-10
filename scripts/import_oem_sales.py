#!/usr/bin/env python3
"""MarkLines OEM 글로벌 판매량 엑셀 → Supabase 사전 집계 4종 적재.

처리 흐름:
  1. 참고/oem 판매량/MarkLines_sales_data_*.xlsx (5개) 모두 로드
  2. 행 = (Country, Group, Maker, Type, Segment, Model, PowerTrain) + 월별 sales 컬럼
  3. PowerTrain 정규화 (ICE/HV/PHEV/EV/FCV/Other)
  4. 4개 사전 집계 테이블에 upsert
     - oem_sales_group_month
     - oem_sales_group_pt_month
     - oem_sales_group_country_month
     - oem_sales_type_seg_month

원시 long 테이블(oem_sales_monthly)은 데이터량(>500K행) 부담으로 적재 생략.
필요 시 별도 옵션 추가.

재실행 안전: PRIMARY KEY 충돌 시 UPDATE (멱등).

사용법:
  python scripts/import_oem_sales.py
"""
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')

sys.path.insert(0, str(Path(__file__).parent))
from lib.db import upsert_rows  # noqa: E402

EXCEL_DIR = Path(__file__).resolve().parents[1] / '참고' / 'oem 판매량'

# PowerTrain 정규화 매핑 (전동화 우선순위: EV > FCV > PHEV > HV > ICE)
PT_PRIORITY = ['EV', 'FCV', 'PHEV', 'HV', 'ICE']
PT_TOKEN_MAP = {
  'EV': 'EV',
  'FCV': 'FCV',
  'PHV': 'PHEV',   # PHV = Plug-in Hybrid Vehicle
  'HV': 'HV',
  'MHV': 'HV',     # Mild Hybrid → HV로 묶음
  'ICE': 'ICE',
}


def normalize_powertrain(raw: str | None) -> str:
  """PowerTrain 원본 값을 6종 정규화 값으로 변환.

  - 단일 토큰: 직접 매핑
  - 복합 토큰(예: 'HV/EV/PHV'): 전동화 우선순위(EV>FCV>PHEV>HV>ICE)로 1개 선택
  - 'N/A' 또는 미매핑: 'Other'
  """
  if not raw or raw == 'N/A':
    return 'Other'
  tokens = [t.strip() for t in raw.split('/') if t.strip()]
  mapped = {PT_TOKEN_MAP[t] for t in tokens if t in PT_TOKEN_MAP}
  if not mapped:
    return 'Other'
  for pt in PT_PRIORITY:
    if pt in mapped:
      return pt
  return 'Other'


def iter_excel_rows(excel_paths: list[Path]):
  """엑셀 파일들에서 (필드 dict, year_month, sales) 튜플 yield.

  엑셀 구조: header row=2, 컬럼 0~6 = 메타, 컬럼 7~ = 월별(YYYYMM) sales.
  sales가 0/None/'-'/공백이면 스킵.
  """
  for path in excel_paths:
    logger.info(f'엑셀 로딩: {path.name}')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    header = [c.value for c in ws[2]]
    # 월 컬럼 인덱스 추출
    month_cols = [
      (i, int(v)) for i, v in enumerate(header)
      if isinstance(v, (int, float)) and v and v > 200000
    ]
    if not month_cols:
      logger.warning(f'{path.name}: 월 컬럼 없음, 스킵')
      wb.close()
      continue

    n_rows = 0
    n_cells = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
      country, group, maker, vtype, segment, model, pt = (row[i] or '' for i in range(7))
      if not group:  # Group 없는 행 스킵
        continue
      pt_norm = normalize_powertrain(pt)
      meta = {
        'country': country.strip() if isinstance(country, str) else '',
        'oem_group': group.strip() if isinstance(group, str) else '',
        'maker': maker.strip() if isinstance(maker, str) else '',
        'vehicle_type': vtype.strip() if isinstance(vtype, str) else '',
        'segment': segment.strip() if isinstance(segment, str) else '',
        'model': model.strip() if isinstance(model, str) else '',
        'powertrain': pt_norm,
      }
      n_rows += 1
      for col_idx, ym in month_cols:
        v = row[col_idx]
        if not isinstance(v, (int, float)):
          continue
        sales = int(v)
        if sales <= 0:
          continue
        n_cells += 1
        yield meta, ym, sales
    wb.close()
    logger.info(f'  → {n_rows}행, {n_cells}개 (그룹×월) 셀')


def aggregate(rows_iter):
  """엑셀 셀 단위 데이터를 4개 사전 집계 dict로 정리."""
  group_month = defaultdict(int)         # (group, ym) → sales
  group_pt_month = defaultdict(int)      # (group, pt, ym) → sales
  group_country_month = defaultdict(int) # (group, country, ym) → sales
  type_seg_month = defaultdict(int)      # (type, segment, ym) → sales

  for meta, ym, sales in rows_iter:
    g = meta['oem_group']
    group_month[(g, ym)] += sales
    group_pt_month[(g, meta['powertrain'], ym)] += sales
    if meta['country']:
      group_country_month[(g, meta['country'], ym)] += sales
    if meta['vehicle_type'] and meta['segment']:
      type_seg_month[(meta['vehicle_type'], meta['segment'], ym)] += sales

  return group_month, group_pt_month, group_country_month, type_seg_month


def main() -> int:
  excel_paths = sorted(EXCEL_DIR.glob('MarkLines_sales_data*.xlsx'))
  if not excel_paths:
    logger.error(f'엑셀 파일 없음: {EXCEL_DIR}')
    return 1
  logger.info(f'엑셀 파일 {len(excel_paths)}개 발견')

  group_month, group_pt_month, group_country_month, type_seg_month = aggregate(
    iter_excel_rows(excel_paths)
  )
  logger.info(
    f'집계 결과: group×month={len(group_month)}, '
    f'group×pt×month={len(group_pt_month)}, '
    f'group×country×month={len(group_country_month)}, '
    f'type×seg×month={len(type_seg_month)}'
  )

  # upsert
  upsert_rows(
    'oem_sales_group_month',
    [{'oem_group': g, 'year_month': ym, 'sales': s} for (g, ym), s in group_month.items()],
    conflict_cols='oem_group,year_month',
  )
  upsert_rows(
    'oem_sales_group_pt_month',
    [{'oem_group': g, 'powertrain': pt, 'year_month': ym, 'sales': s}
     for (g, pt, ym), s in group_pt_month.items()],
    conflict_cols='oem_group,powertrain,year_month',
  )
  upsert_rows(
    'oem_sales_group_country_month',
    [{'oem_group': g, 'country': c, 'year_month': ym, 'sales': s}
     for (g, c, ym), s in group_country_month.items()],
    conflict_cols='oem_group,country,year_month',
  )
  upsert_rows(
    'oem_sales_type_seg_month',
    [{'vehicle_type': t, 'segment': s, 'year_month': ym, 'sales': v}
     for (t, s, ym), v in type_seg_month.items()],
    conflict_cols='vehicle_type,segment,year_month',
  )
  logger.success('OEM 판매량 적재 완료')
  return 0


if __name__ == '__main__':
  sys.exit(main())
