#!/usr/bin/env python3
"""재무 시트(자료정리_월별손익*.xlsx '재무') → Supabase finance_entries 적재.

대차대조표 계정(자산·부채·자본·채권·채무·재고·유형자산·무형자산·현금성자산·차입·증자).
시점 정규화: '연간' 또는 월=12 → annual(period_month=12, 연말), 월=1~11 → monthly.
  (전체 과거=숫자 12, 미국 과거='연간' 텍스트가 섞여 있어 둘 다 annual/12로 통일.)

금액 비노출: 요약은 (자회사·기간종류)별 행수·연도 커버리지·null 카운트만 출력.
검증: 자산 == 부채 + 자본 항등식 mismatch 행수만 보고(금액 비노출, 임계 0.5%).
사용자가 직접 실행한다. WriteSession으로 자동 revalidate('finance_entries').

사용법
-----
  python scripts/sync_finance.py --dry-run
  python scripts/sync_finance.py
  python scripts/sync_finance.py --revalidate-prod

종료 코드
--------
0 정상
2 헤더 검증 실패
"""
import argparse
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')
sys.path.insert(0, str(Path(__file__).parent))
from lib.db import WriteSession  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402

SHEET = '재무'
TABLE = 'finance_entries'
CONFLICT = 'subsidiary,consolidation,period_year,period_kind,period_month,account'
HEADER_ROW = 1
DATA_START = 2

# 1-indexed 컬럼 매핑
COL = {'year': 1, 'pm': 2, 'cons': 3, 'sub': 4, 'account': 5, 'value': 6}
EXPECTED_HEADERS = {
  1: '연도', 2: '연간/월', 3: '연결/별도', 4: '자회사', 5: '계정명', 6: '밸류(백만원)',
}
BATCH_SIZE = 500
TOLERANCE_PCT = 0.5  # 자산 == 부채+자본 mismatch 임계


def _num(v: Any) -> float | None:
  if v is None or v == '' or isinstance(v, bool):
    return None
  if isinstance(v, (int, float)):
    return float(v)
  return None


def _txt(v: Any) -> str:
  return '' if v is None else str(v).strip()


def parse_period(v: Any) -> tuple[str, int] | None:
  """'연간'/월=12 → ('annual',12), 월=1~11 → ('monthly',m). 그 외 None."""
  if isinstance(v, (int, float)) and not isinstance(v, bool):
    m = int(v)
    if m == 12:
      return ('annual', 12)
    if 1 <= m <= 11:
      return ('monthly', m)
    return None
  s = _txt(v)
  if s == '연간':
    return ('annual', 12)
  if s.isdigit():
    m = int(s)
    if m == 12:
      return ('annual', 12)
    if 1 <= m <= 11:
      return ('monthly', m)
  return None


def validate_headers(ws) -> list[str]:
  errs = []
  for c, exp in EXPECTED_HEADERS.items():
    actual = _txt(ws.cell(HEADER_ROW, c).value)
    if actual != exp:
      errs.append(f'  컬럼 {c}: 기대 "{exp}" 실제 "{actual}"')
  return errs


def row_to_entry(ws, r: int) -> dict[str, Any] | None:
  year = ws.cell(r, COL['year']).value
  if not isinstance(year, (int, float)) or isinstance(year, bool):
    return None
  pk = parse_period(ws.cell(r, COL['pm']).value)
  if pk is None:
    return None
  kind, month = pk
  consolidation = _txt(ws.cell(r, COL['cons']).value)
  subsidiary = _txt(ws.cell(r, COL['sub']).value)
  account = _txt(ws.cell(r, COL['account']).value)
  if not consolidation or not subsidiary or not account:
    return None
  return {
    'subsidiary': subsidiary,
    'consolidation': consolidation,
    'period_year': int(year),
    'period_kind': kind,
    'period_month': month,
    'account': account,
    'value_mwon': _num(ws.cell(r, COL['value']).value),
  }


def dedupe(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
  """PK 중복(시트의 '자본' 중복행 등) 제거 — 마지막 행 우선.

  같은 upsert 배치에 동일 conflict key가 2번 들어가면 Postgres ON CONFLICT가 실패하므로 필수."""
  seen: dict[tuple, dict[str, Any]] = {}
  for e in entries:
    key = (
      e['subsidiary'], e['consolidation'], e['period_year'],
      e['period_kind'], e['period_month'], e['account'],
    )
    seen[key] = e
  return list(seen.values())


def summarize(entries: list[dict[str, Any]]) -> None:
  """(자회사·기간종류) 행수·연도 커버리지·null 카운트. 금액 비노출."""
  agg = defaultdict(lambda: {'rows': 0, 'years': set(), 'nulls': 0})
  for e in entries:
    k = (e['subsidiary'], e['period_kind'])
    agg[k]['rows'] += 1
    agg[k]['years'].add(e['period_year'])
    if e['value_mwon'] is None:
      agg[k]['nulls'] += 1
  logger.info('--- 재무 요약 (자회사·기간종류) — 금액 비노출 ---')
  for k in sorted(agg.keys()):
    v = agg[k]
    logger.info(f'  {k} | rows={v["rows"]} | years={sorted(v["years"])} | nulls={v["nulls"]}')


def validate_identity(entries: list[dict[str, Any]]) -> None:
  """자산 == 부채 + 자본 항등식 mismatch 행수만 보고 (금액 비노출, 경고만)."""
  data: dict[tuple, dict[str, float]] = defaultdict(dict)
  for e in entries:
    if e['value_mwon'] is None:
      continue
    key = (e['subsidiary'], e['consolidation'], e['period_year'],
           e['period_kind'], e['period_month'])
    if e['account'] in ('자산', '부채', '자본'):
      data[key][e['account']] = e['value_mwon']
  mismatches = 0
  checked = 0
  for d in data.values():
    if '자산' not in d or '부채' not in d or '자본' not in d:
      continue
    checked += 1
    total = d['자산']
    calc = d['부채'] + d['자본']
    if total == 0:
      continue
    if abs(total - calc) / abs(total) * 100 > TOLERANCE_PCT:
      mismatches += 1
  if checked == 0:
    logger.info('항등식 검증: 자산/부채/자본 동시 입력된 시점 없음 (건너뜀)')
  elif mismatches:
    logger.warning(f'자산 != 부채+자본 mismatch: {mismatches}/{checked}시점 (tol={TOLERANCE_PCT}%)')
  else:
    logger.info(f'검증 OK: 자산 == 부채+자본 ({checked}시점, tol={TOLERANCE_PCT}%)')


def _latest_excel() -> Path:
  base = Path(__file__).resolve().parents[1] / '참고' / '손익'
  cands = sorted(base.glob('자료정리_월별손익*.xlsx'))
  if not cands:
    raise FileNotFoundError(f'손익 엑셀 없음: {base}')
  return cands[-1]


def main() -> int:
  ap = argparse.ArgumentParser(description='재무 시트 → Supabase finance_entries 적재')
  ap.add_argument('--dry-run', action='store_true', help='실제 upsert 없이 파싱·검증만')
  ap.add_argument('--revalidate-prod', action='store_true',
                  help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                       '로컬 수동 실행 시 프로덕션 stale 방지용')
  args = ap.parse_args()

  path = _latest_excel()
  logger.info(f'엑셀 로드: {path}')
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    ws = wb[SHEET]
    errs = validate_headers(ws)
    if errs:
      logger.error(f'[{SHEET}] 헤더 불일치:\n' + '\n'.join(errs))
      return 2
    entries: list[dict[str, Any]] = []
    for r in range(DATA_START, ws.max_row + 1):
      e = row_to_entry(ws, r)
      if e is not None:
        entries.append(e)
    entries = dedupe(entries)
    logger.info(f'[{SHEET}] {len(entries)}행 파싱 완료 (중복 제거 후)')
  finally:
    wb.close()

  summarize(entries)
  validate_identity(entries)

  if args.dry_run:
    logger.success('dry-run 완료')
    return 0
  if not entries:
    logger.warning('적재할 행 없음')
    return 0

  with WriteSession() as w:
    for i in range(0, len(entries), BATCH_SIZE):
      chunk = entries[i:i + BATCH_SIZE]
      w.table(TABLE).upsert(chunk, on_conflict=CONFLICT).execute()
  logger.success(f'finance_entries upsert 완료: {len(entries)}행')
  if args.revalidate_prod:
    revalidate_prod_for_tables([TABLE])
  return 0


if __name__ == '__main__':
  sys.exit(main())
