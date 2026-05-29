#!/usr/bin/env python3
"""재고 시트(자료정리_월별손익*.xlsx '재고') → Supabase inventory_entries 적재.

금액 비노출: 요약은 (분류·항목·kind)별 행수·연도 커버리지·null 카운트만 출력.
검증: 4분류 합(운영+관리+보상+영업+미국환산+우즈벡환산) vs 전체재고 mismatch 행수만 보고.
사용자가 직접 실행한다. WriteSession으로 자동 revalidate('inventory_entries').

사용법
-----
  python scripts/sync_inventory.py --dry-run
  python scripts/sync_inventory.py

종료 코드
--------
0 정상
2 헤더 검증 실패
"""
import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')
sys.path.insert(0, str(Path(__file__).parent))
from lib.db import WriteSession  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402

SHEET = '재고'
TABLE = 'inventory_entries'
CONFLICT = 'category,item,kind,period_year,period_month'
HEADER_ROW = 1
DATA_START = 2

# 1-indexed 컬럼 매핑
COL = {'year': 1, 'pm': 2, 'kind': 3, 'fx': 4, 'category': 5, 'item': 6, 'unit': 7, 'value': 8}
EXPECTED_HEADERS = {
  1: '연도', 2: '월', 3: '계획/실적', 4: '적용환율',
  5: '분류', 6: '항목', 7: '단위', 8: '밸류',
}
KIND_MAP = {'계획': 'plan', '실적': 'actual'}
BATCH_SIZE = 500
TOLERANCE_PCT = 0.5  # 4분류합 vs 전체재고 mismatch 임계


def _num(v: Any) -> float | None:
  """숫자 셀을 float로 정규화."""
  if v is None or v == '' or isinstance(v, bool):
    return None
  if isinstance(v, (int, float)):
    return float(v)
  return None


def _txt(v: Any) -> str:
  return '' if v is None else str(v).strip()


def parse_fx(v: Any) -> float | None:
  """`"1,400원/$"` → 1400.0. 비숫자/null이면 None."""
  s = _txt(v)
  if not s:
    return None
  m = re.search(r'([\d,]+(?:\.\d+)?)', s)
  if not m:
    return None
  return float(m.group(1).replace(',', ''))


def validate_headers(ws) -> list[str]:
  errs = []
  for c, exp in EXPECTED_HEADERS.items():
    actual = _txt(ws.cell(HEADER_ROW, c).value)
    if actual != exp:
      errs.append(f'  컬럼 {c}: 기대 "{exp}" 실제 "{actual}"')
  return errs


def row_to_entry(ws, r: int) -> dict[str, Any] | None:
  year = ws.cell(r, COL['year']).value
  if not isinstance(year, (int, float)):
    return None
  month = ws.cell(r, COL['pm']).value
  if not isinstance(month, (int, float)) or not (1 <= int(month) <= 12):
    return None
  category = _txt(ws.cell(r, COL['category']).value)
  item = _txt(ws.cell(r, COL['item']).value)
  if not category or not item:
    return None
  kind = KIND_MAP.get(_txt(ws.cell(r, COL['kind']).value))
  if kind is None:
    return None
  unit_raw = _txt(ws.cell(r, COL['unit']).value)
  return {
    'category': category,
    'item': item,
    'kind': kind,
    'period_year': int(year),
    'period_month': int(month),
    'unit': unit_raw or None,
    'fx_rate': parse_fx(ws.cell(r, COL['fx']).value),
    'value': _num(ws.cell(r, COL['value']).value),
  }


def summarize(entries: list[dict[str, Any]]) -> None:
  """(분류·항목·kind) 행수·연도 커버리지·null 카운트. 금액 비노출."""
  agg = defaultdict(lambda: {'rows': 0, 'years': set(), 'nulls': 0})
  for e in entries:
    k = (e['category'], e['item'], e['kind'])
    agg[k]['rows'] += 1
    agg[k]['years'].add(e['period_year'])
    if e['value'] is None:
      agg[k]['nulls'] += 1
  logger.info('--- 재고 요약 (분류·항목·kind) — 금액 비노출 ---')
  for k in sorted(agg.keys()):
    v = agg[k]
    logger.info(f'  {k} | rows={v["rows"]} | years={sorted(v["years"])} | nulls={v["nulls"]}')


def validate_total(entries: list[dict[str, Any]]) -> None:
  """4분류 합 vs 전체재고 mismatch 행수만 보고 (금액 비노출, 경고만)."""
  data: dict[tuple, dict[str, float]] = defaultdict(dict)
  for e in entries:
    if e['value'] is None:
      continue
    key = (e['period_year'], e['period_month'], e['kind'])
    cat, item = e['category'], e['item']
    if cat == '전체' and item == '전체 재고':
      data[key]['전체'] = e['value']
    elif cat == '운영' and item == '운영 재고':
      data[key]['운영'] = e['value']
    elif cat == '관리' and item == '관리 재고':
      data[key]['관리'] = e['value']
    elif cat == '보상' and item == '보상 재고':
      data[key]['보상'] = e['value']
    elif cat == '운송' and item == '영업 재고':
      data[key]['영업'] = e['value']
    elif cat == '운송' and item == '미국 운송':
      data[key]['미국USD'] = e['value']
      data[key].setdefault('_fx', e['fx_rate'] or 1400.0)
    elif cat == '운송' and item == '우즈벡 운송':
      data[key]['우즈벡USD'] = e['value']
      data[key].setdefault('_fx', e['fx_rate'] or 1400.0)
  mismatches = 0
  for key, d in data.items():
    total = d.get('전체')
    if total is None:
      continue
    fx = d.get('_fx', 1400.0)
    calc = (
      d.get('운영', 0)
      + d.get('관리', 0)
      + d.get('보상', 0)
      + d.get('영업', 0)
      + d.get('미국USD', 0) * fx / 100
      + d.get('우즈벡USD', 0) * fx / 100
    )
    if total == 0:
      continue
    diff_pct = abs(total - calc) / abs(total) * 100
    if diff_pct > TOLERANCE_PCT:
      mismatches += 1
  if mismatches:
    logger.warning(f'4분류합 vs 전체재고 mismatch: {mismatches}행 (tol={TOLERANCE_PCT}%)')
  else:
    logger.info(f'검증 OK: 4분류합 == 전체재고 (tol={TOLERANCE_PCT}%)')


def _latest_excel() -> Path:
  base = Path(__file__).resolve().parents[1] / '참고' / '손익'
  cands = sorted(base.glob('자료정리_월별손익*.xlsx'))
  if not cands:
    raise FileNotFoundError(f'손익 엑셀 없음: {base}')
  return cands[-1]


def main() -> int:
  ap = argparse.ArgumentParser(description='재고 시트 → Supabase inventory_entries 적재')
  ap.add_argument('--dry-run', action='store_true', help='실제 upsert 없이 파싱·검증만')
  ap.add_argument('--revalidate-prod', action='store_true',
                  help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                       '로컬 수동 실행 시 프로덕션 stale 방지용')
  args = ap.parse_args()

  path = _latest_excel()
  logger.info(f'엑셀 로드: {path}')
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    ws = wb[SHEET]
    errs = validate_headers(ws)
    if errs:
      logger.error(f'[{SHEET}] 헤더 불일치:\n' + '\n'.join(errs))
      return 2
    entries: list[dict[str, Any]] = []
    for r in range(DATA_START, ws.max_row + 1):
      e = row_to_entry(ws, r)
      if e is not None:
        entries.append(e)
    logger.info(f'[{SHEET}] {len(entries)}행 파싱 완료')
  finally:
    wb.close()

  summarize(entries)
  validate_total(entries)

  if args.dry_run:
    logger.success('dry-run 완료')
    return 0
  if not entries:
    logger.warning('적재할 행 없음')
    return 0

  with WriteSession() as w:
    for i in range(0, len(entries), BATCH_SIZE):
      chunk = entries[i:i + BATCH_SIZE]
      w.table(TABLE).upsert(chunk, on_conflict=CONFLICT).execute()
  logger.success(f'inventory_entries upsert 완료: {len(entries)}행')
  if args.revalidate_prod:
    revalidate_prod_for_tables([TABLE])
  return 0


if __name__ == '__main__':
  sys.exit(main())
