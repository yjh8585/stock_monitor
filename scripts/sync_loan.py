#!/usr/bin/env python3
"""이인텔리전스 시트(자료정리_월별손익*.xlsx '이인텔리전스') → Supabase loan_entries 적재.

자회사(이인텔리전스) 대여금 계획·실적 (억원). kind는 '계획'/'실적' 한글 그대로
(DB CHECK ↔ sync ↔ TS union 일치). 공란은 null(미래월·결측월).

금액 비노출: 요약은 (연도·kind)별 행수·월 커버리지·null 카운트만 출력. 금액 합계 비노출.
사용자가 직접 실행한다. WriteSession으로 자동 revalidate('loan_entries').

사용법
-----
  python scripts/sync_loan.py --dry-run
  python scripts/sync_loan.py
  python scripts/sync_loan.py --revalidate-prod

종료 코드
--------
0 정상
2 헤더 검증 실패
"""
import argparse
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')
sys.path.insert(0, str(Path(__file__).parent))
from lib.db import WriteSession  # noqa: E402
from lib.management_excel import resolve_excel_path  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402

SHEET = '이인텔리전스'
TABLE = 'loan_entries'
CONFLICT = 'period_year,period_month,kind'
HEADER_ROW = 1
DATA_START = 2

# 1-indexed 컬럼 매핑
COL = {'year': 1, 'month': 2, 'kind': 3, 'loan': 4}
EXPECTED_HEADERS = {
  1: '연도', 2: '월', 3: '계획/실적', 4: '대여금(억원)',
}
VALID_KINDS = ('계획', '실적')
BATCH_SIZE = 500


def _num(v: Any) -> float | None:
  if v is None or v == '' or isinstance(v, bool):
    return None
  if isinstance(v, (int, float)):
    return float(v)
  return None


def _txt(v: Any) -> str:
  return '' if v is None else str(v).strip()


def validate_headers(ws) -> list[str]:
  errs = []
  for c, exp in EXPECTED_HEADERS.items():
    actual = _txt(ws.cell(HEADER_ROW, c).value)
    if actual != exp:
      errs.append(f'  컬럼 {c}: 기대 "{exp}" 실제 "{actual}"')
  return errs


def row_to_entry(ws, r: int) -> dict[str, Any] | None:
  year = ws.cell(r, COL['year']).value
  month = ws.cell(r, COL['month']).value
  if not isinstance(year, (int, float)) or isinstance(year, bool):
    return None
  if not isinstance(month, (int, float)) or isinstance(month, bool):
    return None
  m = int(month)
  if not (1 <= m <= 12):
    return None
  kind = _txt(ws.cell(r, COL['kind']).value)
  if kind not in VALID_KINDS:
    return None
  return {
    'period_year': int(year),
    'period_month': m,
    'kind': kind,
    'loan_eok': _num(ws.cell(r, COL['loan']).value),
  }


def dedupe(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
  """PK 중복 제거 — 마지막 행 우선 (upsert 배치 내 동일 conflict key 방지)."""
  seen: dict[tuple, dict[str, Any]] = {}
  for e in entries:
    key = (e['period_year'], e['period_month'], e['kind'])
    seen[key] = e
  return list(seen.values())


def summarize(entries: list[dict[str, Any]]) -> None:
  """(연도·kind) 행수·월 커버리지·null 카운트. 금액 비노출."""
  agg = defaultdict(lambda: {'rows': 0, 'months': set(), 'nulls': 0})
  for e in entries:
    k = (e['period_year'], e['kind'])
    agg[k]['rows'] += 1
    agg[k]['months'].add(e['period_month'])
    if e['loan_eok'] is None:
      agg[k]['nulls'] += 1
  logger.info('--- 대여금 요약 (연도·구분) — 금액 비노출 ---')
  for k in sorted(agg.keys()):
    v = agg[k]
    logger.info(f'  {k} | rows={v["rows"]} | months={sorted(v["months"])} | nulls={v["nulls"]}')


def _latest_excel() -> Path:
  return resolve_excel_path()


def main() -> int:
  ap = argparse.ArgumentParser(description='이인텔리전스 시트 → Supabase loan_entries 적재')
  ap.add_argument('--dry-run', action='store_true', help='실제 upsert 없이 파싱·검증만')
  ap.add_argument('--revalidate-prod', action='store_true',
                  help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                       '로컬 수동 실행 시 프로덕션 stale 방지용')
  args = ap.parse_args()

  path = _latest_excel()
  logger.info(f'엑셀 로드: {path}')
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    ws = wb[SHEET]
    errs = validate_headers(ws)
    if errs:
      logger.error(f'[{SHEET}] 헤더 불일치:\n' + '\n'.join(errs))
      return 2
    entries: list[dict[str, Any]] = []
    for r in range(DATA_START, ws.max_row + 1):
      e = row_to_entry(ws, r)
      if e is not None:
        entries.append(e)
    entries = dedupe(entries)
    logger.info(f'[{SHEET}] {len(entries)}행 파싱 완료 (중복 제거 후)')
  finally:
    wb.close()

  summarize(entries)

  if args.dry_run:
    logger.success('dry-run 완료')
    return 0
  if not entries:
    logger.warning('적재할 행 없음')
    return 0

  with WriteSession() as w:
    for i in range(0, len(entries), BATCH_SIZE):
      chunk = entries[i:i + BATCH_SIZE]
      w.table(TABLE).upsert(chunk, on_conflict=CONFLICT).execute()
  logger.success(f'loan_entries upsert 완료: {len(entries)}행')
  if args.revalidate_prod:
    revalidate_prod_for_tables([TABLE])
  return 0


if __name__ == '__main__':
  sys.exit(main())
