#!/usr/bin/env python3
"""손익 엑셀(자료정리_월별손익*.xlsx) → Supabase pnl_entries 적재.

처리 흐름
--------
1. openpyxl로 최신 엑셀 로드 (data_only=True, read_only=True)
2. 3개 시트(연간/연결_월/월) 파싱
3. 행 → dict 변환
   - year_label: '2025(E)' → period_year=2025/is_estimate=True
   - year_label: '2026(P)' → period_year=2026/is_plan=True
   - 매출이 None/0/음수면 SKIP
   - 빈 차원 값은 '' 로 정규화 (PK 일관성)
   - 거래처별 실(sil) 고정 매핑 정정 (SIL_BY_CUSTOMER — UZ Auto → 2실). 정정 시 시트당 경고 1줄
4. postgrest-py로 chunk(500) upsert
5. CLI: --dry-run 옵션

종료 코드
--------
0 정상
1 엑셀 파일 없음
2 시트 헤더 검증 실패
3 upsert 실패

사용법
-----
  python scripts/sync_pnl_excel.py
  python scripts/sync_pnl_excel.py --dry-run
"""
import argparse
import sys
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')

sys.path.insert(0, str(Path(__file__).parent))
from lib.db import upsert_rows  # noqa: E402
from lib.management_excel import resolve_excel_path  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402


def _latest_excel() -> Path:
  return resolve_excel_path()


EXCEL_PATH = _latest_excel()
TABLE = 'pnl_entries'
CONFLICT_COLS = 'basis,year_label,period_month,sil,division,factory,product,customer'
HEADER_ROW = 3  # 1-indexed; 헤더는 3행, 데이터는 4행부터
DATA_START_ROW = 4

# ===== 시트별 헤더 매핑 (1-indexed 컬럼) =====
# 연간 시트: 연도/기준/실/부문/공장/제품/거래처 + 지표 7종
HEADERS_ANNUAL = {
  'year_label':    (2, '연도'),
  'basis_label':   (3, '기준'),
  'sil':           (4, '실'),
  'division':      (5, '부문'),
  'factory':       (6, '공장'),
  'product':       (7, '제품'),
  'customer':      (8, '거래처'),
  'revenue':       (9, '매출'),
  'material_cost': (10, '재료비'),
  'labor_cost':    (12, '노무비'),
  'expense':       (14, '경비'),
  'sga':           (16, '판관비'),
  'rnd':           (18, '연구비'),
  'op_income':     (20, '영업이익'),
}

# 월/연결_월 시트: 기간/연도/월/기준/실/부문/공장/제품/거래처 + 지표 7종
# (매출총이익은 SKIP — DB에 저장 안 함)
HEADERS_MONTHLY = {
  'period_excel':  (2, '기간'),
  'year_label':    (3, '연도'),
  'period_month':  (4, '월'),
  'basis_label':   (5, '기준'),
  'sil':           (6, '실'),
  'division':      (7, '부문'),
  'factory':       (8, '공장'),
  'product':       (9, '제품'),
  'customer':      (10, '거래처'),
  'revenue':       (11, '매출'),
  'material_cost': (12, '재료비'),
  'labor_cost':    (14, '노무비'),
  'expense':       (16, '경비'),
  # 18: 매출총이익 — SKIP
  'sga':           (20, '판관비'),
  'rnd':           (22, '연구비'),
  'op_income':     (24, '영업이익'),
}

# basis 레이블 → DB enum 매핑
BASIS_MAP = {
  '연결': 'consolidated',
  '별도': 'standalone',
}

# 거래처별 실(sil) 고정 매핑 — 엑셀 표기가 어긋나도 화면 기준으로 정정한다.
# 키는 거래처명 소문자(공백 제거 전 strip). UZ Auto 실적은 2실로 표현(사용자 지시 2026-07-30).
SIL_BY_CUSTOMER = {
  'uz auto': '2실',
}

# 시트 → (기본 basis, 헤더 매핑, period_month 처리방식)
SHEETS = [
  ('연간',    'consolidated', HEADERS_ANNUAL,  'annual'),
  ('연결_월', 'consolidated', HEADERS_MONTHLY, 'monthly'),
  ('월',      'standalone',   HEADERS_MONTHLY, 'monthly'),
]


def validate_headers(ws, mapping: dict[str, tuple[int, str]]) -> list[str]:
  """엑셀 헤더가 매핑된 한글 라벨과 일치하는지 검증한다.

  Returns:
    불일치 메시지 리스트 (빈 리스트면 통과)
  """
  errors: list[str] = []
  for key, (col_idx, expected_label) in mapping.items():
    actual = ws.cell(HEADER_ROW, col_idx).value
    actual_str = '' if actual is None else str(actual).strip()
    if actual_str != expected_label:
      errors.append(f'  컬럼 {col_idx} ({key}): 기대 "{expected_label}", 실제 "{actual_str}"')
  return errors


def parse_year_label(raw: Any) -> tuple[str | None, int | None, bool, bool]:
  """연도 라벨을 (year_label, period_year, is_plan, is_estimate)로 정규화.

  - 정수: '2024' → ('2024', 2024, False, False)
  - '2025(E)' → ('2025(E)', 2025, False, True)
  - '2026(P)' → ('2026(P)', 2026, True, False)
  - 그 외/숫자아님: (None, None, ...) 반환 — 호출부에서 SKIP
  """
  if raw is None or raw == '':
    return (None, None, False, False)
  if isinstance(raw, (int, float)) and not isinstance(raw, bool):
    yr = int(raw)
    return (str(yr), yr, False, False)
  s = str(raw).strip()
  if not s:
    return (None, None, False, False)
  is_plan = s.endswith('(P)')
  is_estimate = s.endswith('(E)')
  digits = ''.join(ch for ch in s if ch.isdigit())
  if not digits:
    return (None, None, False, False)
  try:
    yr = int(digits[:4])
  except ValueError:
    return (None, None, False, False)
  return (s, yr, is_plan, is_estimate)


def norm_text(v: Any) -> str:
  """차원 텍스트를 안전하게 정규화 (None/공백 → '')."""
  if v is None:
    return ''
  s = str(v).strip()
  return s


def norm_num(v: Any) -> float | None:
  """숫자 셀을 float로 정규화. 비숫자/빈값은 None."""
  if v is None or v == '':
    return None
  if isinstance(v, bool):
    return None
  if isinstance(v, (int, float)):
    return float(v)
  return None


def normalize_sil(customer: str, sil: str) -> str:
  """거래처 기준으로 실(sil) 라벨을 정정한다. 매핑에 없는 거래처는 엑셀 값 그대로.

  sil이 upsert 충돌키에 포함되므로, 엑셀이 옛 실로 남아 있으면 정정된 행과 별개 행으로
  적재되어 합계가 이중 계산된다. 그래서 적재 전에 여기서 맞춘다.
  """
  expected = SIL_BY_CUSTOMER.get(customer.strip().lower())
  if expected is None:
    return sil
  return expected


def get_cell(row: tuple, col_idx_1based: int) -> Any:
  """1-indexed 컬럼 인덱스로 row 튜플에서 값 조회. 범위 밖이면 None."""
  i = col_idx_1based - 1
  return row[i] if 0 <= i < len(row) else None


def row_to_entry(
  row: tuple,
  default_basis: str,
  mapping: dict[str, tuple[int, str]],
  mode: str,
) -> dict[str, Any] | None:
  """엑셀 행 → pnl_entries dict 변환. SKIP할 경우 None 반환.

  SKIP 정책: revenue·비용·영업이익 7개 metric이 모두 None/0이면 의미 없는 행.
  (revenue=0이라도 비용성 행은 영업이익에 음수 기여하므로 보존해야 1번 차트와 4번 차트가 일치한다.)
  """
  metrics = {
    'revenue':       norm_num(get_cell(row, mapping['revenue'][0])),
    'material_cost': norm_num(get_cell(row, mapping['material_cost'][0])),
    'labor_cost':    norm_num(get_cell(row, mapping['labor_cost'][0])),
    'expense':       norm_num(get_cell(row, mapping['expense'][0])),
    'sga':           norm_num(get_cell(row, mapping['sga'][0])),
    'rnd':           norm_num(get_cell(row, mapping['rnd'][0])),
    'op_income':     norm_num(get_cell(row, mapping['op_income'][0])),
  }
  if all(v is None or v == 0 for v in metrics.values()):
    return None

  # 연도 라벨 파싱
  year_label, period_year, is_plan, is_estimate = parse_year_label(
    get_cell(row, mapping['year_label'][0])
  )
  if year_label is None or period_year is None:
    return None

  # basis 결정 (시트 기본값에 시트 내부 기준 컬럼이 있으면 그 값으로 덮어쓰기)
  basis_raw = norm_text(get_cell(row, mapping['basis_label'][0]))
  basis = BASIS_MAP.get(basis_raw, default_basis)

  # period_month (월별 시트만)
  if mode == 'monthly':
    raw_month = get_cell(row, mapping['period_month'][0])
    if isinstance(raw_month, (int, float)) and 1 <= int(raw_month) <= 12:
      period_month = int(raw_month)
    else:
      return None
  else:
    period_month = 0

  customer = norm_text(get_cell(row, mapping['customer'][0]))

  return {
    'basis': basis,
    'year_label': year_label,
    'period_year': period_year,
    'period_month': period_month,
    'is_plan': is_plan,
    'is_estimate': is_estimate,
    'sil':      normalize_sil(customer, norm_text(get_cell(row, mapping['sil'][0]))),
    'division': norm_text(get_cell(row, mapping['division'][0])),
    'factory':  norm_text(get_cell(row, mapping['factory'][0])),
    'product':  norm_text(get_cell(row, mapping['product'][0])),
    'customer': customer,
    **metrics,
  }


PK_KEYS = ('basis', 'year_label', 'period_month', 'sil', 'division',
           'factory', 'product', 'customer')
NUMERIC_KEYS = ('revenue', 'material_cost', 'labor_cost', 'expense',
                'sga', 'rnd', 'op_income')


def merge_by_pk(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
  """PK 동일 행을 숫자 컬럼 합산으로 병합.

  엑셀에는 동일 (실/부문/공장/제품/거래처) 조합이 sub-입력(예: 임대수익&Royalty)으로
  여러 행에 분할되어 있는 경우가 있다. PostgREST upsert는 마지막 행만 적용되어
  손실이 발생하므로, 적재 전에 합산 병합한다. 비숫자 메타(period_year/is_plan/...)는
  첫 행 값을 그대로 유지한다(동일 PK 내에서는 동일하다고 가정).
  """
  merged: dict[tuple, dict[str, Any]] = {}
  for e in entries:
    pk = tuple(e[k] for k in PK_KEYS)
    if pk not in merged:
      merged[pk] = {**e}
      continue
    target = merged[pk]
    for k in NUMERIC_KEYS:
      a, b = target.get(k), e.get(k)
      if a is None and b is None:
        target[k] = None
      elif a is None:
        target[k] = b
      elif b is None:
        target[k] = a
      else:
        target[k] = a + b
  return list(merged.values())


def parse_sheet(wb, sheet_name: str, default_basis: str,
                mapping: dict[str, tuple[int, str]], mode: str) -> list[dict[str, Any]]:
  """단일 시트를 파싱해 entry 리스트를 반환한다."""
  ws = wb[sheet_name]
  errors = validate_headers(ws, mapping)
  if errors:
    logger.error(f'[{sheet_name}] 헤더 불일치:\n' + '\n'.join(errors))
    raise ValueError(f'헤더 불일치: {sheet_name}')

  entries: list[dict[str, Any]] = []
  total_rows = 0
  skipped = 0
  # (거래처, 엑셀 실, 정정 실) → 행수. 행마다 경고를 찍으면 수백 줄이 되므로 집계 후 1줄.
  sil_fixes: dict[tuple[str, str, str], int] = {}
  for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
    total_rows += 1
    e = row_to_entry(row, default_basis, mapping, mode)
    if e is None:
      skipped += 1
      continue
    raw_sil = norm_text(get_cell(row, mapping['sil'][0]))
    if e['sil'] != raw_sil:
      key = (e['customer'], raw_sil, e['sil'])
      sil_fixes[key] = sil_fixes.get(key, 0) + 1
    entries.append(e)

  for (cust, before, after), cnt in sorted(sil_fixes.items()):
    logger.warning(
      f'[{sheet_name}] 실 정정: {cust} 엑셀 "{before or "(공백)"}" → "{after}" {cnt}행 '
      f'(엑셀 원본도 "{after}"로 맞추면 이 경고가 사라집니다)'
    )

  merged = merge_by_pk(entries)
  duplicates = len(entries) - len(merged)
  logger.info(
    f'[{sheet_name}] 총 {total_rows}행 → 유효 {len(entries)}행 '
    f'(SKIP {skipped}, PK 합산 병합 {duplicates}건) → 적재 {len(merged)}행'
  )
  return merged


def summarize(entries: Iterable[dict[str, Any]]) -> None:
  """dry-run 요약 출력 (기준×연도×월수). 금액 비노출 — 행수·월수만."""
  from collections import defaultdict
  agg: dict[tuple[str, int], dict[str, Any]] = defaultdict(
    lambda: {'rows': 0, 'months': set()}
  )
  for e in entries:
    key = (e['basis'], e['period_year'])
    agg[key]['rows'] += 1
    agg[key]['months'].add(e['period_month'])
  logger.info('--- dry-run 요약 (basis × period_year) — 금액 비노출 ---')
  for key in sorted(agg.keys()):
    v = agg[key]
    months = sorted(v['months'])
    logger.info(f'  {key[0]:<13} {key[1]} | rows={v["rows"]:>5} | months={months}')


def main() -> int:
  parser = argparse.ArgumentParser(description='손익 엑셀 → Supabase pnl_entries 적재')
  parser.add_argument('--dry-run', action='store_true',
                      help='실제 upsert 없이 파싱 결과만 요약 출력')
  parser.add_argument('--revalidate-prod', action='store_true',
                      help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                           '로컬 수동 실행 시 프로덕션 stale 방지용')
  args = parser.parse_args()

  if not EXCEL_PATH.exists():
    logger.error(f'엑셀 파일 없음: {EXCEL_PATH}')
    return 1

  logger.info(f'엑셀 로드: {EXCEL_PATH}')
  wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
  try:
    all_entries: list[dict[str, Any]] = []
    try:
      for sheet, default_basis, mapping, mode in SHEETS:
        entries = parse_sheet(wb, sheet, default_basis, mapping, mode)
        all_entries.extend(entries)
    except ValueError:
      return 2
  finally:
    wb.close()

  logger.info(f'전체 적재 대상 {len(all_entries)}행')
  summarize(all_entries)

  if args.dry_run:
    logger.success('dry-run 완료 — upsert 생략')
    return 0

  try:
    # upsert_rows는 내부적으로 revalidate_for_tables를 자동 호출한다 (기본 URL=localhost)
    n = upsert_rows(TABLE, all_entries, conflict_cols=CONFLICT_COLS)
    logger.success(f'pnl_entries upsert 완료: {n}행')
    if args.revalidate_prod:
      revalidate_prod_for_tables([TABLE])
    return 0
  except Exception as e:
    logger.error(f'upsert 실패: {e}')
    return 3


if __name__ == '__main__':
  sys.exit(main())
