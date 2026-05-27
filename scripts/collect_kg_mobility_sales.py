#!/usr/bin/env python3
"""KG모빌리티 차종별 판매실적 엑셀 → kg_mobility_sales 적재.

플로우 (2026-05 단순화 — 현재 연도 엑셀 1개로 통합):
  1. https://www.kg-mobility.com/cm/ir-data/sales-performance 진입.
  2. **현재 연도 엑셀 1개만 다운로드** — KG IR 엑셀은 2010~현재까지 모든 연도 시트
     단일 파일에 포함 (예: '2026년 판매실적.xlsx' 시트 = Cover/2026/2025/.../2010/Data).
     과거 연도 별도 파일 다운로드 비효율 + 데이터 미세 차이 존재(과거 파일이 누락 발견됨).
  3. `--year-from` ~ `--year-to` 시트만 파싱 (default 2021~현재).
  4. 헤더 4행 패턴 매칭 → 모델 × 12개월 × (Domestic/Export) → KgMobilitySaleRow normalize.
  5. 한글 → 영문 모델명 정규화 (예: '토레스 EVX' → 'Torres EVX' — 다국어 엑셀 통합).
  6. WriteSession에 upsert (PK: period_type,year_period,region,vehicle_model)
     → revalidate_for_tables(['kg_mobility_sales']) 자동.

플래그:
  --year-from 2021     수집 시작 연도 (default 2021).
  --year-to <year>     수집 마지막 연도 (default 현재 연도).
  --reprocess-all      캐시 무시 (현재는 캐시 미운영 — flag만 보존).
  --dry-run            DB 쓰기 없이 파싱 결과 print.
  --keep-downloads     다운로드 엑셀을 data/_kg_downloads/에 보존 (디버그용).

멱등성: PK upsert. 재실행 시 동일 데이터는 갱신만 됨.

사용:
  scripts/venv/Scripts/python.exe scripts/collect_kg_mobility_sales.py \
    --year-from 2024 --year-to 2026 --dry-run --keep-downloads
"""
import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from loguru import logger

from lib.bootstrap import init_script

init_script(__file__)

from lib.db import WriteSession  # noqa: E402

SOURCE_URL = 'https://www.kg-mobility.com/cm/ir-data/sales-performance'
DOWNLOAD_DIR = Path(__file__).resolve().parent.parent / 'data' / '_kg_downloads'
DEFAULT_YEAR_FROM = 2021
PLAYWRIGHT_TIMEOUT_MS = 60_000
DOWNLOAD_TIMEOUT_MS = 30_000

# 엑셀 구조 상수 (정찰 결과 기반)
_MONTH_HEADERS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
# region 헤더 행의 A열 텍스트
_REGION_LABELS = {
  'total': '전체',     # 미사용 — 'Total' 섹션은 Domestic+Export 합이라 적재 안 함
  'domestic': '내수',
  'export': '수출',
  'cbu': '내수',       # 일부 시트에서 'CBU' 단독으로 시작하기도 함 — 컨텍스트로 보면 내수.
}
# 모델 행 중 합계/소계로 건너뛸 토큰
_SKIP_MODEL_PATTERNS = [
  re.compile(r'^total\s*\(cbu', re.IGNORECASE),
  re.compile(r'^sub\s*total', re.IGNORECASE),
  re.compile(r'^total$', re.IGNORECASE),
]
# vehicle_type 정규화 — 엑셀 B열 토큰 그대로 두되, 세그먼트 표기(L1/L3)는 SUV로 매핑
_VTYPE_NORMALIZE = {
  'SUV': 'SUV', 'L1': 'SUV', 'L3': 'SUV',
  'SUT': '픽업',
  'MPV': 'MPV',
  'PC': '세단',
  'EV': 'EV',     # 엑셀에서 type='EV'로 표기되면 그대로 (powertrain은 별도 컬럼)
}

# 모델명 정규화 — 한글/영문 혼재 (과거 시트에 한글 잔존, 신규 시트는 영문) → 영문 통일
_MODEL_NAME_NORMALIZE = {
  '토레스 EVX': 'Torres EVX',
  # 신규 한글 표기 발견 시 추가
}


# ---------------------------------------------------------------------------
# Playwright 다운로드
# ---------------------------------------------------------------------------
def fetch_excel_for_year(page, year: int, dest_dir: Path) -> Path | None:
  """KG 페이지에서 해당 연도 엑셀을 클릭 → expect_download → 저장 후 경로 반환.

  실패 시 None. 본 함수는 page를 재사용한다 (호출자 책임으로 1회 navigate).
  """
  selector = f'div.sale-item:has-text("{year}년") button.btn.file-down'
  try:
    btn = page.locator(selector).first
    btn.wait_for(state='visible', timeout=10_000)
  except Exception as e:
    logger.warning(f'{year}년: 다운로드 버튼 미발견 — {e}')
    return None

  try:
    with page.expect_download(timeout=DOWNLOAD_TIMEOUT_MS) as dl_info:
      btn.click()
    dl = dl_info.value
    dest = dest_dir / f'{year}_{dl.suggested_filename}'
    dl.save_as(str(dest))
    logger.info(f'{year}년: 다운로드 완료 ({dest.stat().st_size/1024:.0f} KB) → {dest.name}')
    return dest
  except Exception as e:
    logger.error(f'{year}년: 다운로드 실패 — {e}')
    return None


# ---------------------------------------------------------------------------
# 엑셀 파싱
# ---------------------------------------------------------------------------
def _find_header_row(ws, year: int) -> int | None:
  """헤더 행(`Total | ... | Jan | Feb | ...` 패턴)의 row 번호. 미발견 시 None.

  엑셀 정찰 결과: 시트마다 위치가 미세하게 다름 (r4 또는 r5). 안전하게 1~10행 스캔.
  """
  for r in range(1, 12):
    e = ws.cell(r, 5).value
    f = ws.cell(r, 6).value
    if e == 'Jan' and f == 'Feb':
      return r
  return None


def _is_skip_model(model: str) -> bool:
  return any(p.match(model.strip()) for p in _SKIP_MODEL_PATTERNS)


def _normalize_vehicle_type(raw: str | None) -> str:
  """B열 vehicle_type 정규화. 매핑 없으면 원본 str 사용.

  엑셀에서 SUV 세그먼트는 '(L1)', '(L3)' 처럼 괄호 포함으로 표기될 수 있음.
  괄호와 공백 제거 후 매핑 시도.
  """
  if not raw:
    return ''
  s = str(raw).strip()
  # '(L1)' / '(L3)' → 'L1' / 'L3'
  s_unparen = s.lstrip('(').rstrip(')').strip()
  return _VTYPE_NORMALIZE.get(s_unparen, _VTYPE_NORMALIZE.get(s, s))


def _normalize_model_name(raw: str) -> str:
  """모델명 한글→영문 통일. _MODEL_NAME_NORMALIZE 매핑 없으면 원본 trim 후 반환."""
  s = (raw or '').strip()
  return _MODEL_NAME_NORMALIZE.get(s, s)


def _normalize_region_label(raw: str) -> str | None:
  """region 텍스트 → '내수'/'수출'/'CKD'/'기타' 등 표준화.

  엑셀 패턴:
    'Total'    → None (Domestic+Export 합산이므로 skip)
    'Domestic' → '내수'
    'Export'   → '수출'
    그 외       → '기타' (소량 출현 시 보존)
  """
  s = (raw or '').strip().lower()
  if s == 'total' or s == '':
    return None
  if s == 'domestic':
    return '내수'
  if s == 'export':
    return '수출'
  if s == 'ckd':
    return 'CKD'
  return '기타'


def _iter_region_blocks(ws, header_row: int):
  """헤더 행 다음부터 데이터를 region 블록 단위로 yield.

  반환: (region_label, model_rows) — model_rows = list[(vehicle_type, model_name, [12개 sales])]
  새 region은 A열에 'Total'/'Domestic'/'Export' 등이 박힌 행을 시작점으로 인식.
  """
  current_region: str | None = None
  current_rows: list[tuple[str, str, list[int]]] = []
  carry_vtype = ''  # B열이 빈 경우 이전 행의 vehicle_type 상속 (merged cell 흔적)

  for r in range(header_row + 1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    c = ws.cell(r, 3).value

    a_str = str(a).strip() if a is not None else ''
    if a_str in ('Total', 'Domestic', 'Export', 'CBU') and ws.cell(r, 5).value == 'Jan':
      # 이건 sub-section의 새 헤더 행 — region 전환.
      if current_region and current_rows:
        yield current_region, current_rows
      current_region = a_str
      current_rows = []
      carry_vtype = ''
      continue
    if a_str in ('Total', 'Domestic', 'Export', 'CBU') and current_region is None:
      # 데이터 행이면서 동시에 region 라벨인 첫 행 (현재 행에 모델 있을 수 있음)
      current_region = a_str

    # 모델 행 — c(C열)에 모델명, e~p(E~P)에 12개월
    model_name = str(c).strip() if c is not None else ''
    if not model_name:
      continue
    if _is_skip_model(model_name):
      continue
    if b is not None and str(b).strip():
      carry_vtype = str(b).strip()
    monthly = []
    for m_idx in range(12):
      v = ws.cell(r, 5 + m_idx).value
      try:
        monthly.append(int(v) if v not in (None, '') else 0)
      except (TypeError, ValueError):
        monthly.append(0)
    current_rows.append((carry_vtype, model_name, monthly))

  if current_region and current_rows:
    yield current_region, current_rows


def _resolve_sheet(wb, target_year: int) -> str | None:
  """target_year에 매칭되는 시트명 반환.

  KG 엑셀 시트 패턴 변천:
    2022~2025 파일: 시트명이 정확히 'YYYY' (예: '2025', '2024')
    2021 파일: 'Total Sales Vol. 2021' 같은 prefix 포함 형태 + '2020','2019',...
    그 외: 시트명에 4자리 연도가 포함된 첫 매칭을 사용.
  """
  exact = str(target_year)
  if exact in wb.sheetnames:
    return exact
  # 시트명에 'YYYY' 4자리가 포함된 것 우선 (단, Cover 제외)
  for name in wb.sheetnames:
    if name.lower().startswith('cover'):
      continue
    if str(target_year) in name:
      return name
  return None


def parse_excel(path: Path, target_year: int) -> list[dict]:
  """엑셀 1개 → target_year 연도 시트만 추출해 row dict list 반환.

  하나의 파일에 다년치 시트가 있을 수 있으나 target_year와 일치하는 시트만 처리한다.
  매칭 실패 시 _kg_parse_failed_<year>.json에 raw 첫 5행 덤프 후 빈 list 반환.
  """
  wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
  try:
    sheet_name = _resolve_sheet(wb, target_year)
    if sheet_name is None:
      logger.warning(
        f'{target_year}년 시트 없음 in {path.name} (sheets={wb.sheetnames})'
      )
      return []
    ws = wb[sheet_name]
    header_row = _find_header_row(ws, target_year)
    if header_row is None:
      logger.error(f'{path.name}::{sheet_name} 헤더 미발견 — 구조 변경 가능')
      _dump_failed_rows(path, target_year, ws)
      return []
    logger.debug(f'{path.name}::{sheet_name} header_row={header_row}')

    rows: list[dict] = []
    for region_label, model_rows in _iter_region_blocks(ws, header_row):
      norm_region = _normalize_region_label(region_label)
      if not norm_region:
        continue  # Total 섹션 skip
      for vtype, model, monthly in model_rows:
        vtype_norm = _normalize_vehicle_type(vtype)
        model_norm = _normalize_model_name(model)
        for m_idx, units in enumerate(monthly):
          if units <= 0:
            continue
          year_period = f'{target_year}-{m_idx+1:02d}'
          rows.append({
            'period_type': 'month',
            'year_period': year_period,
            'region': norm_region,
            'vehicle_model': model_norm,
            'vehicle_type': vtype_norm,
            'powertrain': None,  # 매핑 테이블에서 별도 join
            'sales_units': units,
            'source_url': SOURCE_URL,
          })
    return rows
  finally:
    wb.close()


def _dump_failed_rows(path: Path, year: int, ws) -> None:
  """파싱 실패 시 raw 첫 5행을 덤프해 디버깅 용이하게."""
  dump_path = path.parent.parent / f'_kg_parse_failed_{year}.json'
  raw = []
  for r in range(1, 6):
    raw.append([ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)])
  try:
    with dump_path.open('w', encoding='utf-8') as f:
      json.dump({'file': str(path), 'year': year, 'rows': raw}, f,
                ensure_ascii=False, indent=2, default=str)
    logger.warning(f'  raw 행 5개 덤프 → {dump_path}')
  except Exception as e:
    logger.warning(f'  dump 실패: {e}')


# ---------------------------------------------------------------------------
# 요약 / 멱등 dedup
# ---------------------------------------------------------------------------
def dedupe_rows(rows: list[dict]) -> list[dict]:
  """동일 PK 충돌 시 sales_units **합산**. KG 엑셀은 같은 모델을 segment별로 분리해 표기
  (예: 'Tivoli' L1 + 'Tivoli' L3 두 행 — 같은 (year, region, model)) → 합산이 정확한 총량.
  vehicle_type/source_url 등 비-PK 필드는 첫 행 값을 유지 (sum 영향 없음).
  """
  by_pk: dict[tuple, dict] = {}
  for r in rows:
    pk = (r['period_type'], r['year_period'], r['region'], r['vehicle_model'])
    if pk in by_pk:
      by_pk[pk]['sales_units'] += r['sales_units']
    else:
      by_pk[pk] = dict(r)  # copy to avoid mutating original
  return list(by_pk.values())


def print_summary(rows: list[dict]) -> None:
  if not rows:
    logger.info('적재 대상 행 0건')
    return
  by_year: dict[str, int] = {}
  by_region: dict[str, int] = {}
  models: set[str] = set()
  total_units = 0
  for r in rows:
    y = r['year_period'][:4]
    by_year[y] = by_year.get(y, 0) + 1
    by_region[r['region']] = by_region.get(r['region'], 0) + 1
    models.add(r['vehicle_model'])
    total_units += r['sales_units']
  logger.info(f'행 총 {len(rows)}개, 모델 {len(models)}개, 합계 {total_units:,}대')
  logger.info(f'  연도별: {dict(sorted(by_year.items()))}')
  logger.info(f'  region별: {dict(sorted(by_region.items()))}')
  logger.info(f'  모델 샘플(15): {sorted(models)[:15]}')


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
  p = argparse.ArgumentParser(description='KG모빌리티 차종별 판매 수집.')
  p.add_argument('--year-from', type=int, default=DEFAULT_YEAR_FROM,
                 help=f'백필 시작 연도 (default {DEFAULT_YEAR_FROM})')
  p.add_argument('--year-to', type=int, default=None,
                 help='마지막 연도 (default 현재 연도)')
  p.add_argument('--reprocess-all', action='store_true',
                 help='캐시 무시 — 현재는 캐시 미운영, flag 예약')
  p.add_argument('--dry-run', action='store_true',
                 help='DB 쓰기 없이 파싱 결과만 print')
  p.add_argument('--keep-downloads', action='store_true',
                 help='다운로드 엑셀을 data/_kg_downloads/에 보존 (디버그용)')
  return p.parse_args()


def main() -> int:
  args = parse_args()
  current_year = datetime.now(timezone.utc).year
  year_to = args.year_to or current_year
  year_range = list(range(args.year_from, year_to + 1))
  logger.info(f'KG 모빌리티 차종별 판매 수집: 시트 {year_range[0]}~{year_range[-1]} '
              f'(현재 연도 엑셀 1개에서 추출 · dry_run={args.dry_run}, '
              f'keep_downloads={args.keep_downloads})')

  DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
  # Playwright는 lazy import (env / 시스템 브라우저 의존)
  from playwright.sync_api import sync_playwright  # noqa: E402

  all_rows: list[dict] = []
  downloaded: list[Path] = []
  failed_years: list[int] = []

  with sync_playwright() as pw:
    browser = pw.chromium.launch()
    ctx = browser.new_context(accept_downloads=True)
    page = ctx.new_page()
    try:
      page.goto(SOURCE_URL, wait_until='networkidle', timeout=PLAYWRIGHT_TIMEOUT_MS)
    except Exception as e:
      logger.error(f'KG IR 페이지 로드 실패: {e}')
      ctx.close()
      browser.close()
      return 1

    # 현재 연도 엑셀 1개만 다운로드 (모든 과거 연도 시트 포함)
    excel_path = fetch_excel_for_year(page, current_year, DOWNLOAD_DIR)
    ctx.close()
    browser.close()

    if excel_path is None:
      logger.error(f'현재 연도({current_year}) 엑셀 다운로드 실패 — 종료')
      return 1
    downloaded.append(excel_path)

  # 다년치 시트 파싱
  for year in year_range:
    try:
      year_rows = parse_excel(excel_path, year)
    except Exception as e:
      logger.error(f'{year}년 시트 파싱 실패: {e}')
      failed_years.append(year)
      continue
    if not year_rows:
      logger.warning(f'{year}년 시트 0행 — 시트 미존재 또는 파싱 실패')
      failed_years.append(year)
      continue
    logger.info(f'{year}년: 파싱 {len(year_rows)}행')
    all_rows.extend(year_rows)

  all_rows = dedupe_rows(all_rows)
  print_summary(all_rows)
  if failed_years:
    logger.warning(f'실패/누락 연도: {failed_years}')

  if args.dry_run:
    logger.success('dry-run 종료 (DB 쓰기 없음)')
    _maybe_cleanup_downloads(downloaded, args.keep_downloads)
    return 0 if not failed_years else 1

  if not all_rows:
    logger.warning('적재할 행 없음 — DB 호출 생략')
    _maybe_cleanup_downloads(downloaded, args.keep_downloads)
    return 1 if failed_years else 0

  try:
    with WriteSession() as w:
      w.table('kg_mobility_sales').upsert(
        all_rows,
        on_conflict='period_type,year_period,region,vehicle_model',
      ).execute()
    logger.success(f'kg_mobility_sales upsert 완료: {len(all_rows)}행')
  except Exception as e:
    logger.error(f'upsert 실패: {e}')
    _maybe_cleanup_downloads(downloaded, args.keep_downloads)
    return 2

  _maybe_cleanup_downloads(downloaded, args.keep_downloads)
  return 0 if not failed_years else 1


def _maybe_cleanup_downloads(files: list[Path], keep: bool) -> None:
  if keep:
    logger.info(f'다운로드 파일 보존: {len(files)}개 in {DOWNLOAD_DIR}')
    return
  for f in files:
    try:
      f.unlink(missing_ok=True)
    except Exception as e:
      logger.debug(f'  파일 삭제 실패 {f.name}: {e}')


if __name__ == '__main__':
  try:
    sys.exit(main())
  except KeyboardInterrupt:
    logger.warning('사용자 중단')
    sys.exit(130)
  except Exception as e:
    logger.exception(f'예기치 못한 오류: {e}')
    sys.exit(1)
