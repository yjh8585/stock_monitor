#!/usr/bin/env python3
"""현대차 미국/유럽 현지(retail) 판매 → hyundai_retail_sales 적재 (Phase 2C).

플로우:
  1. https://www.hyundai.com/worldwide/ko/company/ir/ir-resources/sales-results 진입.
  2. `#field-sales-type` dropdown으로 연도 순회 (--year-from~--year-to).
  3. 연도별 2개 다운로드 버튼 중 retail 2종 추출:
       - "YYYY년 미국 현지 판매파일 다운로드" → region='US'
       - "YYYY년 유럽 현지 판매파일 다운로드" → region='EU'
  4. `page.expect_download()`로 엑셀 캡처 → `data/_hyundai_downloads/{year}_retail_{region}.xlsx`.
  5. openpyxl 파싱:
       - 공통 헤더: r3 = B='Models', D~O=Jan~Dec, P=Total
       - 모델 row: B=vehicle_type(섹션 시작 row만, 이후 carry), C=모델명, D~O=월별, P=연간합
       - 'Sub-total'/'Total'/'Light CV' section row 처리
       - US 전용: 'US Total Industry' / 'HMC Market Share' row
  6. WriteSession에 upsert hyundai_retail_sales.
       - 모델 row × 12개월 = month rows + 연간합 = annual row
       - US Industry/MarketShare는 별도 vehicle_model='Industry'/'MarketShare' row

플래그:
  --year-from 2021      수집 시작 연도 (default 2021).
  --year-to <year>      수집 마지막 연도 (default 현재 연도).
  --region {US,EU,all}  대상 region (default all).
  --dry-run             DB 쓰기 없이 파싱 결과만 print.
  --keep-downloads      다운로드 엑셀 보존 (default 보존).

멱등성: PK upsert.

사용:
  scripts/venv/Scripts/python.exe scripts/collect_hyundai_retail.py \\
    --year-from 2024 --year-to 2024 --dry-run
"""
import argparse
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from loguru import logger

from lib.bootstrap import init_script

init_script(__file__)

from lib.db import WriteSession  # noqa: E402

SOURCE_URL = 'https://www.hyundai.com/worldwide/ko/company/ir/ir-resources/sales-results'
DOWNLOAD_DIR = Path(__file__).resolve().parent.parent / 'data' / '_hyundai_downloads'
DEFAULT_YEAR_FROM = 2021
PLAYWRIGHT_TIMEOUT_MS = 60_000
DOWNLOAD_TIMEOUT_MS = 30_000
DROPDOWN_WAIT_MS = 2_000

# 버튼 텍스트 매칭 → region
_REGION_LABELS = {
  'US': '미국 현지 판매',
  'EU': '유럽 현지 판매',
}

# Sub-total / Total / Grand Total 토큰
_SKIP_ROW_PATTERNS = [
  re.compile(r'^sub[\s-]*total$', re.IGNORECASE),
  re.compile(r'^grand\s*total$', re.IGNORECASE),
]

# vehicle_type 섹션 토큰
_VEHICLE_TYPES = {'PC', 'RV', 'Light CV', 'CV', 'LCV'}


def _is_skip_row(s: str) -> bool:
  return any(p.match((s or '').strip()) for p in _SKIP_ROW_PATTERNS)


def _safe_int(v) -> int | None:
  if v in (None, ''):
    return None
  if isinstance(v, str) and v.strip().startswith('#REF'):
    return None
  try:
    return int(v)
  except (TypeError, ValueError):
    try:
      return int(float(v))
    except (TypeError, ValueError):
      return None


def _safe_float(v) -> float | None:
  if v in (None, ''):
    return None
  if isinstance(v, str) and v.strip().startswith('#REF'):
    return None
  try:
    return float(v)
  except (TypeError, ValueError):
    return None


# ---------------------------------------------------------------------------
# Playwright 다운로드
# ---------------------------------------------------------------------------
def _select_year(page, year: int) -> bool:
  try:
    page.locator('#field-sales-type .btn-dropdown').click()
    page.wait_for_timeout(300)
    page.locator(f'#field-sales-type .btn-option:has-text("{year}")').first.click()
    page.wait_for_timeout(DROPDOWN_WAIT_MS)
    return True
  except Exception as e:
    logger.warning(f'{year}년 dropdown 선택 실패: {e}')
    return False


def fetch_excel(page, year: int, region: str, dest_dir: Path) -> Path | None:
  """연도/region 조합으로 retail 엑셀 다운로드. 페이지는 호출자 책임으로 navigate+select."""
  label = _REGION_LABELS[region]
  selector = f'button.btn-download:has-text("{label}")'
  try:
    btn = page.locator(selector).first
    btn.wait_for(state='visible', timeout=10_000)
  except Exception as e:
    logger.warning(f'{year}년 {region} retail: 버튼 미발견 — {e}')
    return None

  try:
    with page.expect_download(timeout=DOWNLOAD_TIMEOUT_MS) as dl_info:
      btn.click()
    dl = dl_info.value
    dest = dest_dir / f'{year}_retail_{region}.xlsx'
    dl.save_as(str(dest))
    logger.info(f'{year}년 {region} retail: 다운로드 완료 ({dest.stat().st_size/1024:.0f} KB)')
    return dest
  except Exception as e:
    logger.error(f'{year}년 {region} retail: 다운로드 실패 — {e}')
    return None


# ---------------------------------------------------------------------------
# 엑셀 파서
# ---------------------------------------------------------------------------
def _find_header_row(ws) -> tuple[int, int] | None:
  """'Jan'/'Jan.' 셀이 있는 row, col 반환."""
  for r in range(1, 12):
    for c in range(3, 6):
      v = ws.cell(r, c).value
      if v is None:
        continue
      s = str(v).strip().lower().rstrip('.')
      if s == 'jan':
        return r, c
  return None


def parse_retail_excel(path: Path, year: int, region: str) -> list[dict]:
  """미국/유럽 retail 엑셀 → rows.

  구조 공통:
    r1: 제목 ('Y2024 US Retail Sales' / 'Y2024 Europe Subsidiary Sales')
    r3: 헤더 (B='Models', D=Jan ... O=Dec, P=Total)
    이후: vehicle_type 섹션(B='PC'/'RV'/'Light CV') + 모델 row(B=type 또는 빈, C=모델명)
    'Sub-total' (C='Sub-total') → skip (model 차원 합계)
    'Total' (B='Total') → vehicle_model='Total' 적재
    US 전용: 'US Total Industry' (B), 'HMC Market Share' (B)

  반환 dict 형식 (월별 + annual 합계):
    {period_type, year_period, region, vehicle_type, vehicle_model,
     retail_units, market_share, industry_total, source_type, source_url}
  """
  wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
  ws = wb[wb.sheetnames[0]]
  hdr = _find_header_row(ws)
  if hdr is None:
    logger.error(f'{path.name}: 헤더 미발견')
    return []
  header_row, jan_col = hdr

  rows: list[dict] = []
  carry_vtype = ''
  for r in range(header_row + 1, ws.max_row + 1):
    b = ws.cell(r, 2).value
    c = ws.cell(r, 3).value
    s_b = str(b).strip() if b is not None else ''
    s_c = str(c).strip() if c is not None else ''

    # 빈 row
    if not s_b and not s_c:
      continue

    # Sub-total (C='Sub-total') → skip (vehicle_type 합계, Total로 충분)
    if _is_skip_row(s_c):
      continue

    # B에 vehicle_type 시작 + C 비어있음 → carry 설정 후 다음 row(같은 row C='' 무시)
    if s_b in _VEHICLE_TYPES and not s_c:
      carry_vtype = s_b
      continue

    # B에 vehicle_type + C 비어있지 않음 → 같은 row가 첫 모델 (US 패턴)
    if s_b in _VEHICLE_TYPES and s_c:
      carry_vtype = s_b
      # 이 row의 모델은 아래 일반 모델 처리 흐름으로 fallthrough
      model_name = s_c
      monthly = [_safe_int(ws.cell(r, jan_col + i).value) for i in range(12)]
      total = _safe_int(ws.cell(r, jan_col + 12).value)
      rows.extend(_emit_model_rows(
        year, region, carry_vtype, model_name, monthly, total, path,
      ))
      continue

    # B='Total' (전체 합계, EU 패턴 r50 / US 패턴 r33)
    if s_b.lower() == 'total' and not s_c:
      monthly = [_safe_int(ws.cell(r, jan_col + i).value) for i in range(12)]
      total = _safe_int(ws.cell(r, jan_col + 12).value)
      rows.extend(_emit_model_rows(
        year, region, '', 'Total', monthly, total, path,
      ))
      continue

    # B='Light CV' + C 비어있음 (EU r48: Light CV 자체가 한 row, 모델 breakdown 없음)
    if s_b == 'Light CV' and not s_c:
      monthly = [_safe_int(ws.cell(r, jan_col + i).value) for i in range(12)]
      total = _safe_int(ws.cell(r, jan_col + 12).value)
      rows.extend(_emit_model_rows(
        year, region, 'Light CV', 'Light CV', monthly, total, path,
      ))
      continue

    # B='US Total Industry' (US r35) → industry_total
    if region == 'US' and 'industry' in s_b.lower():
      monthly = [_safe_int(ws.cell(r, jan_col + i).value) for i in range(12)]
      total = _safe_int(ws.cell(r, jan_col + 12).value)
      rows.extend(_emit_industry_rows(year, region, monthly, total, path))
      continue

    # B='HMC Market Share' (US r37) → market_share
    if region == 'US' and 'market share' in s_b.lower():
      monthly = [_safe_float(ws.cell(r, jan_col + i).value) for i in range(12)]
      total = _safe_float(ws.cell(r, jan_col + 12).value)
      rows.extend(_emit_share_rows(year, region, monthly, total, path))
      continue

    # 일반 모델 row: B 비어있고 C에 모델명
    if not s_b and s_c and carry_vtype:
      model_name = s_c
      monthly = [_safe_int(ws.cell(r, jan_col + i).value) for i in range(12)]
      total = _safe_int(ws.cell(r, jan_col + 12).value)
      rows.extend(_emit_model_rows(
        year, region, carry_vtype, model_name, monthly, total, path,
      ))
      continue

    # Note: ... 등 무시
    if s_b.lower().startswith('note') or s_b.startswith('*'):
      continue

  return rows


def _emit_model_rows(
  year: int, region: str, vtype: str, model: str,
  monthly: list[int | None], total: int | None, path: Path,
) -> list[dict]:
  """모델 1개에 대해 월별 12 + 연간 1 dict 생성."""
  out = []
  for m_idx, val in enumerate(monthly):
    if val is None:
      continue
    out.append({
      'period_type': 'month',
      'year_period': f'{year}-{m_idx + 1:02d}',
      'region': region,
      'vehicle_type': vtype,
      'vehicle_model': model,
      'retail_units': val,
      'market_share': None,
      'industry_total': None,
      'source_type': 'hmc-ir',
      'source_url': SOURCE_URL,
    })
  if total is not None:
    out.append({
      'period_type': 'annual',
      'year_period': str(year),
      'region': region,
      'vehicle_type': vtype,
      'vehicle_model': model,
      'retail_units': total,
      'market_share': None,
      'industry_total': None,
      'source_type': 'hmc-ir',
      'source_url': SOURCE_URL,
    })
  return out


def _emit_industry_rows(
  year: int, region: str, monthly: list[int | None], total: int | None, path: Path,
) -> list[dict]:
  out = []
  for m_idx, val in enumerate(monthly):
    if val is None:
      continue
    out.append({
      'period_type': 'month',
      'year_period': f'{year}-{m_idx + 1:02d}',
      'region': region,
      'vehicle_type': '',
      'vehicle_model': 'Industry',
      'retail_units': None,
      'market_share': None,
      'industry_total': val,
      'source_type': 'hmc-ir',
      'source_url': SOURCE_URL,
    })
  if total is not None:
    out.append({
      'period_type': 'annual',
      'year_period': str(year),
      'region': region,
      'vehicle_type': '',
      'vehicle_model': 'Industry',
      'retail_units': None,
      'market_share': None,
      'industry_total': total,
      'source_type': 'hmc-ir',
      'source_url': SOURCE_URL,
    })
  return out


def _emit_share_rows(
  year: int, region: str, monthly: list[float | None], total: float | None, path: Path,
) -> list[dict]:
  out = []
  for m_idx, val in enumerate(monthly):
    if val is None:
      continue
    out.append({
      'period_type': 'month',
      'year_period': f'{year}-{m_idx + 1:02d}',
      'region': region,
      'vehicle_type': '',
      'vehicle_model': 'MarketShare',
      'retail_units': None,
      'market_share': round(val, 5),
      'industry_total': None,
      'source_type': 'hmc-ir',
      'source_url': SOURCE_URL,
    })
  if total is not None:
    out.append({
      'period_type': 'annual',
      'year_period': str(year),
      'region': region,
      'vehicle_type': '',
      'vehicle_model': 'MarketShare',
      'retail_units': None,
      'market_share': round(total, 5),
      'industry_total': None,
      'source_type': 'hmc-ir',
      'source_url': SOURCE_URL,
    })
  return out


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main() -> None:
  parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
  parser.add_argument('--year-from', type=int, default=DEFAULT_YEAR_FROM)
  parser.add_argument('--year-to', type=int, default=datetime.now(timezone.utc).year)
  parser.add_argument('--region', choices=['US', 'EU', 'all'], default='all')
  parser.add_argument('--dry-run', action='store_true')
  parser.add_argument('--keep-downloads', action='store_true', default=True)
  args = parser.parse_args()

  regions = ['US', 'EU'] if args.region == 'all' else [args.region]
  DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

  # 1) 다운로드 (Playwright)
  from playwright.sync_api import sync_playwright

  downloaded: dict[tuple[int, str], Path] = {}
  with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    ctx = browser.new_context(accept_downloads=True)
    page = ctx.new_page()
    page.set_default_timeout(PLAYWRIGHT_TIMEOUT_MS)
    page.goto(SOURCE_URL, wait_until='domcontentloaded')
    page.wait_for_timeout(3000)

    for year in range(args.year_from, args.year_to + 1):
      logger.info(f'=== {year}년 ===')
      if not _select_year(page, year):
        continue
      for region in regions:
        path = fetch_excel(page, year, region, DOWNLOAD_DIR)
        if path:
          downloaded[(year, region)] = path
    browser.close()

  # 2) 파싱
  all_rows: list[dict] = []
  for (year, region), path in downloaded.items():
    rows = parse_retail_excel(path, year, region)
    logger.info(f'{year} {region}: {len(rows)} rows 파싱')
    all_rows.extend(rows)

  if not all_rows:
    logger.warning('파싱 결과 0건 — 종료')
    return

  # 샘플 출력
  logger.info(f'총 {len(all_rows)} rows. 샘플 5개:')
  for row in all_rows[:5]:
    logger.info(f'  {row}')

  if args.dry_run:
    logger.info('--dry-run 모드 — DB 적재 skip')
    return

  # 3) upsert (WriteSession이 revalidate 자동 hook)
  with WriteSession() as w:
    chunk = 500
    for i in range(0, len(all_rows), chunk):
      batch = all_rows[i:i + chunk]
      w.table('hyundai_retail_sales').upsert(
        batch,
        on_conflict='period_type,year_period,region,vehicle_type,vehicle_model',
      ).execute()
  logger.success(f'적재 완료: {len(all_rows)} rows')


if __name__ == '__main__':
  main()
