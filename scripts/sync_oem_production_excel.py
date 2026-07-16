#!/usr/bin/env python3
"""MarkLines OEM 생산량 Excel 자동 다운로드 + 변경 감지 + DB 동기화.

판매 쪽 `sync_oem_excel.py`의 쌍둥이다. 소스 페이지만 vehicle_production으로 다르고
인증(MARKLINES_COOKIE)·발견·검증·md5 스킵 흐름은 동일하다.

처리 흐름
--------
1. MARKLINES_COOKIE 환경변수 + requests.Session 으로 인증
2. https://www.marklines.com/en/vehicle_production/search?rf=left_menu GET
3. HTML 파싱 → "…Latest Month…" 링크의 href = Excel 다운로드 URL
   (실측 링크 텍스트: 'Jan 2024 - Latest Month of 2026' → MarkLines_product_data_en.xlsx)
4. Excel 임시 파일로 다운로드 (스트리밍)
5. 구조 검증 — 메타 6열이 Country..Model 이고 7번째가 YYYYMM 인지
   **판매 export(7열, PowerTrain 포함)를 잘못 받는 것을 여기서 막는다.**
6. MD5 비교 → 기존 참고/oem 생산량/MarkLines_product_data_en.xlsx 와 동일 시 EXIT 0
7. 다르면 파일 교체 → import_oem_production.main() 호출 → DB upsert (멱등)
   upsert_rows()가 캐시 무효화까지 자동 수행한다(lib/revalidate.py COLUMN_TO_TAGS).

과거 연도(2020~2023) 파일은 같은 페이지에서 연도별로 따로 받으며, 한 번 적재하면 바뀌지 않아
자동 동기화 대상이 아니다(판매와 동일 관행). GHA 러너에는 최신 파일만 존재한다.

종료 코드
--------
0: 정상 (변경 없음 또는 업데이트 성공)
1: MARKLINES_COOKIE 미설정 / 쿠키 만료
2: Excel 다운로드 링크 파싱 실패
3: Excel 다운로드 HTTP 오류
4: 다운로드 파일 구조 불일치 (잘못된 export — 로컬 파일 보존)
5: import_oem_production.main() 실패

사용법
-----
  python scripts/sync_oem_production_excel.py
"""
import hashlib
import os
import shutil
import sys
import tempfile
from pathlib import Path
from urllib.parse import urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup
from loguru import logger

sys.path.insert(0, str(Path(__file__).parent))
from lib.bootstrap import init_script  # noqa: E402

init_script(__file__)

import import_oem_production  # noqa: E402

EXCEL_DIR = Path(__file__).resolve().parents[1] / '참고' / 'oem 생산량'
TARGET_FILE = EXCEL_DIR / 'MarkLines_product_data_en.xlsx'
SEARCH_URL = 'https://www.marklines.com/en/vehicle_production/search?rf=left_menu'
USER_AGENT = (
  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
  '(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'
)
REQUEST_TIMEOUT = 180

#: 생산 export의 메타 헤더. 판매는 여기에 'PowerTrain'이 붙은 7열이라 이 튜플로 구분된다.
EXPECTED_HEADER_PREFIX = ('Country', 'Group', 'Maker/Brand', 'Type', 'Segment', 'Model')

#: 다운로드 URL이 반드시 포함해야 할 파일명 조각. 판매(sales_data) 링크를 집는 사고 방지.
EXPECTED_FILE_TOKEN = 'product_data'


def build_session() -> requests.Session:
  """MARKLINES_COOKIE 기반 인증 세션 생성."""
  cookie = os.environ.get('MARKLINES_COOKIE', '').strip()
  if not cookie:
    logger.error('MARKLINES_COOKIE 미설정 — .env.local 에 등록 필요')
    sys.exit(1)
  s = requests.Session()
  s.headers.update({
    'User-Agent': USER_AGENT,
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9,ko;q=0.8',
    'Cookie': cookie,
    'Referer': 'https://www.marklines.com/en/',
  })
  return s


def fetch_search_page(session: requests.Session) -> str:
  """생산 검색 페이지 HTML 반환. 쿠키 만료 시 EXIT 1."""
  r = session.get(SEARCH_URL, timeout=REQUEST_TIMEOUT, allow_redirects=True)
  if r.status_code != 200:
    logger.error(f'검색 페이지 HTTP {r.status_code} — 쿠키 만료 가능')
    sys.exit(1)
  if '/login' in r.url or 'login_form' in r.text.lower():
    logger.error('로그인 페이지로 리다이렉트 — MARKLINES_COOKIE 만료')
    sys.exit(1)
  return r.text


def find_latest_excel_url(html: str) -> str:
  """HTML에서 "Latest Month" 링크의 absolute URL 반환.

  실측 링크 텍스트가 'Jan 2024 - Latest Month of 2026'이라 판매 페이지("Latest month")와
  같은 부분일치 규칙이 그대로 통한다. 다만 같은 페이지에 연도별 링크(2020~2023)도 있으므로
  **href에 product_data가 들어간 것만** 후보로 삼아 오선택을 막는다.
  """
  soup = BeautifulSoup(html, 'html.parser')
  candidates: list[tuple[str, str]] = []
  for a in soup.find_all('a', href=True):
    href = a['href']
    text = a.get_text(strip=True)
    if 'latest month' in text.lower() and EXPECTED_FILE_TOKEN in href:
      candidates.append((href, text))
  if not candidates:
    logger.error('"Latest Month" 생산 링크 미발견 — MarkLines 페이지 구조 변경 가능')
    sys.exit(2)
  href, label = candidates[0]
  full_url = urljoin(SEARCH_URL, href)
  logger.info(f'Excel 링크 발견: "{label}" → {full_url}')
  return full_url


def download_excel(session: requests.Session, url: str) -> Path:
  """Excel을 임시 파일로 다운로드. 경로 반환."""
  r = session.get(url, timeout=REQUEST_TIMEOUT, stream=True, allow_redirects=True)
  if r.status_code != 200:
    logger.error(f'Excel 다운로드 HTTP {r.status_code} — URL: {url}')
    sys.exit(3)
  ctype = r.headers.get('Content-Type', '')
  if 'html' in ctype.lower():
    logger.error(f'다운로드 응답이 HTML — 인증 실패 가능 (Content-Type={ctype})')
    sys.exit(3)
  tmp = tempfile.NamedTemporaryFile(prefix='marklines_prod_', suffix='.xlsx', delete=False)
  size = 0
  with tmp as f:
    for chunk in r.iter_content(chunk_size=64 * 1024):
      if chunk:
        f.write(chunk)
        size += len(chunk)
  logger.info(f'다운로드 완료: {size/1024/1024:.2f} MB → {tmp.name}')
  return Path(tmp.name)


def validate_excel_structure(path: Path) -> bool:
  """다운로드한 Excel이 생산 export 형식인지 검증.

  - Sheet1 존재
  - row 2 의 1~6번 컬럼이 ('Country','Group','Maker/Brand','Type','Segment','Model')
  - 7번째 컬럼이 YYYYMM 정수
  판매 export를 받았다면 7번째가 'PowerTrain' 문자열이라 여기서 걸린다.
  실패 시 로컬 파일 보존.
  """
  try:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
  except Exception as e:
    logger.error(f'Excel 열기 실패: {e}')
    return False
  try:
    if 'Sheet1' not in wb.sheetnames:
      logger.error(f'Sheet1 누락. 시트 목록: {wb.sheetnames}')
      return False
    ws = wb['Sheet1']
    header = tuple(ws.cell(2, c).value for c in range(1, 7))
    if header != EXPECTED_HEADER_PREFIX:
      logger.error(f'헤더 불일치 — 기대: {EXPECTED_HEADER_PREFIX}, 실제: {header}')
      return False
    first_month = ws.cell(2, 7).value
    if not (isinstance(first_month, (int, float)) and first_month > 200000):
      logger.error(
        f'7번째 컬럼이 YYYYMM이 아님: {first_month!r} — 판매 export(PowerTrain)를 받았을 수 있음'
      )
      return False
    logger.info(f'구조 검증 통과 (시작 월={int(first_month)})')
    return True
  finally:
    wb.close()


def md5_of(path: Path) -> str:
  h = hashlib.md5()
  with path.open('rb') as f:
    for chunk in iter(lambda: f.read(64 * 1024), b''):
      h.update(chunk)
  return h.hexdigest()


def main() -> int:
  EXCEL_DIR.mkdir(parents=True, exist_ok=True)
  session = build_session()
  html = fetch_search_page(session)
  excel_url = find_latest_excel_url(html)
  tmp_path = download_excel(session, excel_url)

  try:
    if not validate_excel_structure(tmp_path):
      logger.error('다운로드 파일 구조 불일치 — 로컬 파일 보존, 종료')
      return 4

    new_md5 = md5_of(tmp_path)
    if TARGET_FILE.exists():
      old_md5 = md5_of(TARGET_FILE)
      if new_md5 == old_md5:
        logger.success(f'변경 없음 (md5={new_md5[:12]}…) — DB upsert 스킵')
        return 0
      logger.info(f'변경 감지 (old={old_md5[:12]}… new={new_md5[:12]}…)')
    else:
      logger.info(f'기존 파일 없음 — 신규 저장 (md5={new_md5[:12]}…)')

    shutil.move(str(tmp_path), str(TARGET_FILE))
    logger.info(f'파일 교체 완료: {TARGET_FILE}')

    rc = import_oem_production.main()
    if rc != 0:
      logger.error(f'import_oem_production.main() 실패 (rc={rc})')
      return 5
    logger.success('OEM 생산량 Excel 동기화 완료')
    return 0
  finally:
    if tmp_path.exists():
      tmp_path.unlink(missing_ok=True)


if __name__ == '__main__':
  sys.exit(main())
