#!/usr/bin/env python3
"""MarkLines 판매 엑셀 → oem_model_segment 적재 (멱등).

`참고/oem 판매량/MarkLines_sales_data*.xlsx` 전부를 읽어 (model, country) 유니크로 병합한다.
연도별 파일에 같은 모델이 반복 등장하므로 최신 파일이 나중에 오도록 파일명 정렬 순서를 지킨다.
"""
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, str(Path(__file__).parent))

from lib.bootstrap import init_script  # noqa: E402

init_script(__file__)

from openpyxl import load_workbook  # noqa: E402
from loguru import logger  # noqa: E402

from lib.db import WriteSession, upsert_rows  # noqa: E402
from lib.model_segment import parse_segment_rows  # noqa: E402

EXCEL_GLOB = '참고/oem 판매량/MarkLines_sales_data*.xlsx'


def main() -> int:
  root = Path(__file__).parent.parent
  paths = sorted(root.glob(EXCEL_GLOB))
  if not paths:
    logger.error(f'엑셀 없음 — {EXCEL_GLOB}')
    return 1

  merged: dict[tuple[str, str], dict] = {}
  for path in paths:
    # read_only=True 는 이 MarkLines 엑셀에서 쓸 수 없다 — 시트 XML의 <dimension ref="A1"/>
    # 태그가 실제 데이터 범위를 반영하지 않아(생성 툴 버그로 추정) read_only 최적화 경로가
    # iter_rows(min_row=3, ...)에서 행을 0개로 잘라낸다(2026-08-13 실측, 5개 파일 전부 재현).
    # read_only=False 로 전체 로드하면 정상 32,181행(최대 파일 기준) 반환 확인.
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb['Sheet1']
    raw = list(ws.iter_rows(min_row=3, max_col=7, values_only=True))
    wb.close()
    for row in parse_segment_rows(raw):
      merged[(row['model'], row['country'])] = row
    logger.info(f'{path.name}: 누적 {len(merged):,}건')

  rows = list(merged.values())
  with WriteSession():
    upsert_rows('oem_model_segment', rows, conflict_cols='model,country')
  logger.success(f'oem_model_segment {len(rows):,}건 적재 완료')
  return 0


if __name__ == '__main__':
  sys.exit(main())
