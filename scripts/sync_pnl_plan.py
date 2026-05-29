#!/usr/bin/env python3
"""손익 계획 시트(자료정리_월별손익*.xlsx '계획') → Supabase pnl_plan 적재.

금액 비노출: 요약은 (분류·항목·kind)별 행수·연도 커버리지·null 카운트만 출력.
사용자가 직접 실행한다. WriteSession으로 자동 revalidate('pnl_plan').

사용법
-----
  python scripts/sync_pnl_plan.py --dry-run
  python scripts/sync_pnl_plan.py

종료 코드
--------
0 정상
2 헤더 검증 실패
"""
import argparse
import sys
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')
sys.path.insert(0, str(Path(__file__).parent))
from lib.db import WriteSession  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402

SHEET_PLAN = '계획'
SHEET_ORDER = '수주'
TABLE = 'pnl_plan'
CONFLICT = 'category,item,basis,kind,period_year,period_type,period_month'
HEADER_ROW = 1
DATA_START = 2
# 계획 시트 — 1-indexed 컬럼 (연결/별도 포함)
COL_PLAN = {'year': 1, 'pm': 2, 'kind': 3, 'basis': 4, 'category': 5, 'item': 6, 'unit': 7, 'value': 8}
EXPECTED_HEADERS_PLAN = {
  1: '연도', 2: '연간/월', 3: '계획/실적',
  4: '연결/별도', 5: '분류', 6: '항목', 7: '단위', 8: '밸류',
}
# 수주 시트 — 연결/별도 없음, basis는 'consolidated' 고정
COL_ORDER = {'year': 1, 'pm': 2, 'kind': 3, 'category': 4, 'item': 5, 'unit': 6, 'value': 7}
EXPECTED_HEADERS_ORDER = {
  1: '연도', 2: '연간/월', 3: '계획/실적',
  4: '분류', 5: '항목', 6: '단위', 7: '밸류',
}
BASIS_MAP = {'연결': 'consolidated', '별도': 'standalone'}
KIND_MAP = {'계획': 'plan', '실적': 'actual'}
BATCH_SIZE = 500


def _latest_excel() -> Path:
  """참고/손익/ 디렉터리에서 자료정리_월별손익*.xlsx 중 가장 최신(사전순 마지막)을 반환."""
  base = Path(__file__).resolve().parents[1] / '참고' / '손익'
  cands = sorted(base.glob('자료정리_월별손익*.xlsx'))
  if not cands:
    raise FileNotFoundError(f'손익 엑셀 없음: {base}')
  return cands[-1]


def _num(v: Any) -> float | None:
  """숫자 셀을 float로 정규화. 비숫자/빈값은 None."""
  if v is None or v == '' or isinstance(v, bool):
    return None
  if isinstance(v, (int, float)):
    return float(v)
  return None


def _txt(v: Any) -> str:
  """텍스트 셀을 안전하게 정규화 (None → '')."""
  return '' if v is None else str(v).strip()


def validate_headers(ws, expected: dict[int, str]) -> list[str]:
  """헤더 행이 expected 매핑과 일치하는지 검증한다."""
  errs = []
  for c, exp in expected.items():
    actual = _txt(ws.cell(HEADER_ROW, c).value)
    if actual != exp:
      errs.append(f'  컬럼 {c}: 기대 "{exp}" 실제 "{actual}"')
  return errs


def _parse_period(pm_raw: Any) -> tuple[str, int] | None:
  if isinstance(pm_raw, (int, float)) and 1 <= int(pm_raw) <= 12:
    return ('month', int(pm_raw))
  if _txt(pm_raw) == '연간':
    return ('annual', 0)
  return None


def row_to_entry_plan(ws, r: int) -> dict[str, Any] | None:
  """계획 시트 행 → pnl_plan dict (연결/별도 컬럼 있음)."""
  year = ws.cell(r, COL_PLAN['year']).value
  if not isinstance(year, (int, float)):
    return None
  category = _txt(ws.cell(r, COL_PLAN['category']).value)
  item = _txt(ws.cell(r, COL_PLAN['item']).value)
  if not category or not item:
    return None
  basis = BASIS_MAP.get(_txt(ws.cell(r, COL_PLAN['basis']).value))
  kind = KIND_MAP.get(_txt(ws.cell(r, COL_PLAN['kind']).value))
  if basis is None or kind is None:
    return None
  period = _parse_period(ws.cell(r, COL_PLAN['pm']).value)
  if period is None:
    return None
  return {
    'category': category, 'item': item, 'basis': basis, 'kind': kind,
    'period_year': int(year),
    'period_type': period[0], 'period_month': period[1],
    'unit': _txt(ws.cell(r, COL_PLAN['unit']).value),
    'value': _num(ws.cell(r, COL_PLAN['value']).value),
  }


def row_to_entry_order(ws, r: int) -> dict[str, Any] | None:
  """수주 시트 행 → pnl_plan dict. 연결/별도 컬럼 없음 → basis='consolidated' 고정.
  분류='수주', 항목∈{입찰총액|수주성공|수주실패|연기∙중단∙취소}."""
  year = ws.cell(r, COL_ORDER['year']).value
  if not isinstance(year, (int, float)):
    return None
  category = _txt(ws.cell(r, COL_ORDER['category']).value)
  item = _txt(ws.cell(r, COL_ORDER['item']).value)
  if not category or not item:
    return None
  kind = KIND_MAP.get(_txt(ws.cell(r, COL_ORDER['kind']).value))
  if kind is None:
    return None
  period = _parse_period(ws.cell(r, COL_ORDER['pm']).value)
  if period is None:
    return None
  return {
    'category': category, 'item': item, 'basis': 'consolidated', 'kind': kind,
    'period_year': int(year),
    'period_type': period[0], 'period_month': period[1],
    'unit': _txt(ws.cell(r, COL_ORDER['unit']).value),
    'value': _num(ws.cell(r, COL_ORDER['value']).value),
  }


def summarize(entries: list[dict[str, Any]]) -> None:
  """dry-run 요약 출력 (분류·항목·kind별 행수·연도 커버리지·null 카운트). 금액 비노출."""
  from collections import defaultdict
  agg = defaultdict(lambda: {'rows': 0, 'years': set(), 'nulls': 0})
  for e in entries:
    k = (e['category'], e['item'], e['kind'])
    agg[k]['rows'] += 1
    agg[k]['years'].add(e['period_year'])
    if e['value'] is None:
      agg[k]['nulls'] += 1
  logger.info('--- 계획 요약 (분류·항목·kind) — 금액 비노출 ---')
  for k in sorted(agg.keys()):
    v = agg[k]
    logger.info(f'  {k} | rows={v["rows"]} | years={sorted(v["years"])} | nulls={v["nulls"]}')


def main() -> int:
  ap = argparse.ArgumentParser(description='손익 계획 시트 → Supabase pnl_plan 적재')
  ap.add_argument('--dry-run', action='store_true',
                  help='실제 upsert 없이 파싱 결과만 요약 출력')
  ap.add_argument('--revalidate-prod', action='store_true',
                  help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                       '로컬 수동 실행 시 프로덕션 stale 방지용')
  args = ap.parse_args()

  path = _latest_excel()
  logger.info(f'엑셀 로드: {path}')
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    entries: list[dict[str, Any]] = []
    # 1) 계획 시트
    ws_plan = wb[SHEET_PLAN]
    errs = validate_headers(ws_plan, EXPECTED_HEADERS_PLAN)
    if errs:
      logger.error(f'[{SHEET_PLAN}] 헤더 불일치:\n' + '\n'.join(errs))
      return 2
    plan_n = 0
    for r in range(DATA_START, ws_plan.max_row + 1):
      e = row_to_entry_plan(ws_plan, r)
      if e is not None:
        entries.append(e)
        plan_n += 1
    logger.info(f'[{SHEET_PLAN}] {plan_n}행')
    # 2) 수주 시트 (분류=수주, 항목=입찰총액/수주성공/수주실패/연기∙중단∙취소)
    ws_order = wb[SHEET_ORDER]
    errs = validate_headers(ws_order, EXPECTED_HEADERS_ORDER)
    if errs:
      logger.error(f'[{SHEET_ORDER}] 헤더 불일치:\n' + '\n'.join(errs))
      return 2
    order_n = 0
    for r in range(DATA_START, ws_order.max_row + 1):
      e = row_to_entry_order(ws_order, r)
      if e is not None:
        entries.append(e)
        order_n += 1
    logger.info(f'[{SHEET_ORDER}] {order_n}행')
  finally:
    wb.close()

  logger.info(f'적재 대상 {len(entries)}행')
  summarize(entries)

  if args.dry_run:
    logger.success('dry-run 완료')
    return 0

  if not entries:
    logger.warning('적재할 행 없음')
    return 0

  # WriteSession: __exit__에서 revalidate_for_tables(['pnl_plan']) 자동 호출
  with WriteSession() as w:
    for i in range(0, len(entries), BATCH_SIZE):
      chunk = entries[i:i + BATCH_SIZE]
      w.table(TABLE).upsert(chunk, on_conflict=CONFLICT).execute()

  logger.success(f'pnl_plan upsert 완료: {len(entries)}행')
  if args.revalidate_prod:
    revalidate_prod_for_tables([TABLE])
  return 0


if __name__ == '__main__':
  sys.exit(main())
