#!/usr/bin/env python3
"""현대차 차종별/공장별/지역별 판매 엑셀 → hyundai_sales 적재.

플로우 (PR3, KG 패턴과 동일):
  1. https://www.hyundai.com/worldwide/ko/company/ir/ir-resources/sales-results 진입.
  2. `#field-sales-type` 연도 dropdown(2016~현재년)을 `--year-from`~`--year-to` 범위로 순회.
  3. 연도별 5개 다운로드 버튼 중 3종을 추출:
       - "YYYY년 차종별 매출실적파일 다운로드"       → kind='model'
       - "YYYY년 해외 공장별 판매파일 다운로드"       → kind='factory'
       - "YYYY년 지역별 수출실적파일 다운로드"         → kind='export'
     ("미국/유럽 현지 판매" 버튼은 본 스크립트 범위 외 — 추후 PR에서 확장 가능)
  4. `page.expect_download()`로 엑셀 캡처 → `data/_hyundai_downloads/{year}_{kind}.xlsx`.
  5. openpyxl로 3종 엑셀 각각 파싱:
       - **model**:   B=vehicle_type(PC/RV/CV), C=모델명, D~O=Jan~Dec.
                      A열 없음. 'Domestic'/'Export' section은 B열에 위치.
                      region='내수'(Domestic) / '수출'(Export). factory=''.
       - **factory**: B=공장코드(HMI/HAOS/BHMC/HMMA/...) 또는 'Domestic'/'Export'/'Total',
                      C=모델명, D~O=Jan~Dec. region: 'Domestic'→'내수', 'Export'→'수출' (공장 기준 판매처).
                      factory=공장코드.
       - **export**:  B=대권역(North America/Europe/Others), C=세부 지역(U.S.A./Canada/...),
                      D~O=Jan~Dec. region_name=세부 지역명. → 별도 테이블 `hyundai_export_regions`
                      (source='export-by-region', period_type='month').
  6. WriteSession에 upsert:
       - model/factory rows → hyundai_sales (PK: period_type,year_period,region,factory,vehicle_model)
       - export rows → hyundai_export_regions (PK: period_type,year_period,source,region_name)
     → revalidate_for_tables(['hyundai_sales','hyundai_export_regions']) 자동.

플래그:
  --year-from 2021     수집 시작 연도 (default 2021).
  --year-to <year>     수집 마지막 연도 (default 현재 연도).
  --kind {model,factory,export,all}
                       처리할 엑셀 종류 (default all).
  --reprocess-all      캐시 무시 — 현재는 캐시 미운영, flag만 보존.
  --dry-run            DB 쓰기 없이 파싱 결과만 print.
  --keep-downloads     다운로드 엑셀을 data/_hyundai_downloads/에 보존 (default 보존).

멱등성: PK upsert. 재실행 시 동일 데이터는 갱신만 됨.

사용:
  scripts/venv/Scripts/python.exe scripts/collect_hyundai_sales.py \\
    --year-from 2024 --year-to 2025 --dry-run --keep-downloads
"""
import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from loguru import logger

from lib.bootstrap import init_script

init_script(__file__)

from lib.db import WriteSession  # noqa: E402

SOURCE_URL = 'https://www.hyundai.com/worldwide/ko/company/ir/ir-resources/sales-results'
DOWNLOAD_DIR = Path(__file__).resolve().parent.parent / 'data' / '_hyundai_downloads'
DEFAULT_YEAR_FROM = 2021
PLAYWRIGHT_TIMEOUT_MS = 60_000
DOWNLOAD_TIMEOUT_MS = 30_000
DROPDOWN_WAIT_MS = 2_000

# 다운로드 버튼 텍스트 매칭 → kind
_KIND_LABELS = {
  'model': '차종별 매출실적',
  'factory': '해외 공장별 판매',
  'export': '지역별 수출실적',
}

# 합계/소계 row 토큰 (모든 kind 공통, 다양한 변형 포함)
_SKIP_ROW_PATTERNS = [
  re.compile(r'^sub[\s-]*total$', re.IGNORECASE),
  re.compile(r'^total$', re.IGNORECASE),
  re.compile(r'^grand\s*total$', re.IGNORECASE),
]

# Model 파일의 section header 토큰 (B열 단독)
_MODEL_SECTION_LABELS = {
  'Domestic': '내수',
  'Export': '수출',
}

# Factory 파일의 region 토큰 (B열, 공장 sub-section)
_FACTORY_REGION_LABELS = {
  'Domestic': '내수',
  'Export': '수출',
}

# Plant header 시점에 region이 사전 SET되어야 하는 plant들 (region sub-section 없이 바로 모델 row 진입).
#   - CKD: B='CKD' 다음 B='' C='PV'/'CV' (region 토큰 없음)
#   - Others (2021/2022): B='Others' 다음 B='' C='Others' (region 토큰 없음, 모델명도 'Others')
# region 표기는 plant 이름과 동일하게 ('CKD'/'Others') → UI에서 명확.
_REGIONLESS_PLANTS = {'CKD', 'Others'}

# HTBC/BHMC Export 등 region sub-section만 있고 모델 breakdown 없는 plant 케이스.
# 사용 위치: 'Aggregate' vehicle_model로 region 합계만 적재(Grand Total 누락 방지).
_PLANT_AGGREGATE_MODEL = 'Aggregate'


def _is_skip_row(s: str) -> bool:
  """Sub-total/Total/Grand Total 등 합계 row 판별."""
  return any(p.match((s or '').strip()) for p in _SKIP_ROW_PATTERNS)


# ---------------------------------------------------------------------------
# Playwright 다운로드
# ---------------------------------------------------------------------------
def _select_year(page, year: int) -> bool:
  """현대 dropdown(field-sales-type)에서 연도 선택. 성공 시 True."""
  try:
    page.locator('#field-sales-type .btn-dropdown').click()
    page.wait_for_timeout(300)
    page.locator(f'#field-sales-type .btn-option:has-text("{year}")').first.click()
    page.wait_for_timeout(DROPDOWN_WAIT_MS)
    return True
  except Exception as e:
    logger.warning(f'{year}년 dropdown 선택 실패: {e}')
    return False


def fetch_excel_for_year_kind(
  page, year: int, kind: str, dest_dir: Path,
) -> Path | None:
  """연도/kind 조합으로 엑셀 1개 다운로드 → 저장 후 경로 반환.

  본 함수는 page를 재사용한다 (호출자 책임으로 1회 navigate + _select_year).
  """
  label_substr = _KIND_LABELS[kind]
  selector = f'button.btn-download:has-text("{label_substr}")'
  try:
    btn = page.locator(selector).first
    btn.wait_for(state='visible', timeout=10_000)
  except Exception as e:
    logger.warning(f'{year}년 {kind}: 버튼 미발견 — {e}')
    return None

  try:
    with page.expect_download(timeout=DOWNLOAD_TIMEOUT_MS) as dl_info:
      btn.click()
    dl = dl_info.value
    dest = dest_dir / f'{year}_{kind}.xlsx'
    dl.save_as(str(dest))
    logger.info(
      f'{year}년 {kind}: 다운로드 완료 ({dest.stat().st_size/1024:.0f} KB)'
    )
    return dest
  except Exception as e:
    logger.error(f'{year}년 {kind}: 다운로드 실패 — {e}')
    return None


# ---------------------------------------------------------------------------
# 공통 엑셀 헬퍼
# ---------------------------------------------------------------------------
def _find_header_row(ws, jan_col_min: int = 3, jan_col_max: int = 6) -> tuple[int, int] | None:
  """'Jan' 셀이 있는 row, col 반환. row가 헤더 행, col이 1월 컬럼 시작 위치.

  엑셀 일관: 헤더는 r3, 'Jan' 위치는 파일별로 D(=4) 고정. fallback으로 1~6열 스캔.
  """
  for r in range(1, 12):
    for c in range(jan_col_min, jan_col_max + 1):
      v = ws.cell(r, c).value
      if v is None:
        continue
      s = str(v).strip().lower().rstrip('.')
      if s == 'jan':
        return r, c
  return None


def _safe_int(v) -> int:
  if v in (None, ''):
    return 0
  try:
    return int(v)
  except (TypeError, ValueError):
    try:
      return int(float(v))
    except (TypeError, ValueError):
      return 0


def _has_region_aggregate(ws, row: int, jan_col: int) -> bool:
  """factory.xlsx에서 region header row(B='Domestic'/'Export')가 모델 breakdown 없이
  자체적으로 region 합계만 표시하는 plant인지 판별.

  조건:
    1. 같은 row의 P(=jan_col+12) 또는 D~O 중 하나라도 값 > 0.
    2. 다음 non-empty row의 B가 다음 region/plant/'Total' 같은 토큰 (= 모델 row 아님)
       또는 C='Sub-total' (= 직전 region이 한 줄 합계로 종료).

  예: HTBC plant의 r179 'Domestic' (391), r180 'Export' (1252), r181 'Sub-total'.
  """
  # 같은 row에 값이 있는가
  has_value = any(
    _safe_int(ws.cell(row, jan_col + m).value) > 0 for m in range(12)
  )
  if not has_value:
    return False
  # look-ahead: 다음 non-empty row 확인
  for nr in range(row + 1, min(ws.max_row + 1, row + 6)):
    nb = str(ws.cell(nr, 2).value or '').strip()
    nc = str(ws.cell(nr, 3).value or '').strip()
    if not nb and not nc:
      continue
    # 다음 row가 모델 row면 (B 비어있고 C에 모델명) → 이건 region header — Aggregate 아님
    if not nb and nc and not _is_skip_row(nc):
      return False
    # 다음 row가 다른 region/plant/Total/Sub-total → 이 row가 plant 단위 합계 (Aggregate)
    return True
  return False


def _dump_failed_rows(path: Path, kind: str, year: int, ws) -> None:
  """파싱 실패 시 raw 첫 8행을 덤프 (디버깅용)."""
  dump_path = path.parent.parent / f'_hyundai_parse_failed_{year}_{kind}.json'
  raw = []
  for r in range(1, 9):
    raw.append([ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)])
  try:
    with dump_path.open('w', encoding='utf-8') as f:
      json.dump({'file': str(path), 'year': year, 'kind': kind, 'rows': raw}, f,
                ensure_ascii=False, indent=2, default=str)
    logger.warning(f'  raw 행 8개 덤프 → {dump_path}')
  except Exception as e:
    logger.warning(f'  dump 실패: {e}')


# ---------------------------------------------------------------------------
# 엑셀 파서 — model (차종별 매출실적)
# ---------------------------------------------------------------------------
def parse_model_excel(path: Path, year: int) -> list[dict]:
  """차종별 매출실적 엑셀 → rows.

  구조:
    r1: 'YYYY Unit Sales by Model' (제목)
    r3: 헤더 ('Model', '', 'Jan.', ..., 'Total')
    r6: 'Domestic' section header (B열)
    r8~: PC/RV/CV vehicle_type rows. B=vehicle_type(첫 row만), C=model명.
    중간: 'Sub-total' row (skip), 'Total' row (skip), 'Export' section (B열).

  region: 'Domestic'→'내수', 'Export'→'수출'. factory=''.
  """
  wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
  try:
    ws = wb[wb.sheetnames[0]]
    hdr = _find_header_row(ws)
    if hdr is None:
      logger.error(f'{path.name} 헤더 미발견')
      _dump_failed_rows(path, 'model', year, ws)
      return []
    header_row, jan_col = hdr

    rows: list[dict] = []
    current_region: str | None = None
    carry_vtype = ''
    for r in range(header_row + 1, ws.max_row + 1):
      b = ws.cell(r, 2).value
      c = ws.cell(r, 3).value
      s_b = str(b).strip() if b is not None else ''
      s_c = str(c).strip() if c is not None else ''

      # Section header — B에 'Domestic'/'Export', C 비어있음
      if s_b in _MODEL_SECTION_LABELS and not s_c:
        current_region = _MODEL_SECTION_LABELS[s_b]
        carry_vtype = ''
        continue
      # 'Total' / 'Grand Total' section header — skip
      if _is_skip_row(s_b) and not s_c:
        carry_vtype = ''
        continue
      # Sub-total row (C='Sub-total' 또는 'Total') — skip
      if _is_skip_row(s_c):
        continue
      # 빈 row
      if not s_b and not s_c:
        continue
      # 모델 row — region 미설정이면 skip (제목 row 등)
      if current_region is None:
        continue
      # vehicle_type carry (B 비면 직전 값 유지)
      if s_b:
        carry_vtype = s_b
      model_name = s_c
      if not model_name:
        continue

      # 12개월 → row append
      for m_idx in range(12):
        units = _safe_int(ws.cell(r, jan_col + m_idx).value)
        if units <= 0:
          continue
        year_period = f'{year}-{m_idx+1:02d}'
        rows.append({
          'period_type': 'month',
          'year_period': year_period,
          'region': current_region,
          'factory': '',
          'vehicle_model': model_name,
          'vehicle_type': carry_vtype,
          'powertrain': None,
          'sales_units': units,
          'source_url': SOURCE_URL,
        })
    return rows
  finally:
    wb.close()


# ---------------------------------------------------------------------------
# 엑셀 파서 — factory (해외 공장별 판매)
# ---------------------------------------------------------------------------
def parse_factory_excel(path: Path, year: int) -> list[dict]:
  """해외 공장별 판매 엑셀 → rows.

  구조:
    r6: 'HMI' section header (B열 공장코드, C 비어있음)
    r8: B='Domestic' (region), C=모델명 (sub-section 시작)
    r9~: B 비어있음, C=모델명 (carry: 'Domestic' 유지)
    'Sub-total' row → skip
    'Export' sub-section → region='수출'
    'Total' (B='Total', C 비어있음) → skip
    다음 공장 (B='HAOS', ...) → 새 plant.

  region: 'Domestic'→'내수' / 'Export'→'수출' (공장 입장에서의 판매처).
  factory: 공장코드 (HMI/HAOS/BHMC/HMMA/HMGMA/HMMC/HMMR/HMB/HMMI/HTBC/KMX/Others/Russia/Vietnam/Singapore/CKD).
  """
  wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
  try:
    ws = wb[wb.sheetnames[0]]
    hdr = _find_header_row(ws)
    if hdr is None:
      logger.error(f'{path.name} 헤더 미발견')
      _dump_failed_rows(path, 'factory', year, ws)
      return []
    header_row, jan_col = hdr

    rows: list[dict] = []
    current_factory: str | None = None
    current_region: str | None = None
    for r in range(header_row + 1, ws.max_row + 1):
      b = ws.cell(r, 2).value
      c = ws.cell(r, 3).value
      s_b = str(b).strip() if b is not None else ''
      s_c = str(c).strip() if c is not None else ''

      # 'Grand Total' — 종료
      if s_b.lower().startswith('grand total'):
        break
      # Factory section header — B에 공장코드, C 비어있음, 'Domestic'/'Export'/'Total'/Sub-total 아님.
      if s_b and not s_c and s_b not in _FACTORY_REGION_LABELS and not _is_skip_row(s_b):
        current_factory = s_b
        # region sub-section 없이 바로 모델 row가 오는 plant는 region을 plant 이름으로 사전 설정.
        current_region = s_b if s_b in _REGIONLESS_PLANTS else None
        continue
      # Plant 내 region sub-section: B='Domestic'/'Export'
      if s_b in _FACTORY_REGION_LABELS:
        current_region = _FACTORY_REGION_LABELS[s_b]
        # HTBC/BHMC Export/Others 등 — 모델 breakdown 없이 region 합계만 표시되는 plant.
        # 같은 row의 P(Total) 값이 있고, 다음 row가 모델 row 아닌 경우 'Aggregate'로 적재.
        if (not s_c or _is_skip_row(s_c)) and current_factory:
          if _has_region_aggregate(ws, r, jan_col):
            for m_idx in range(12):
              units = _safe_int(ws.cell(r, jan_col + m_idx).value)
              if units <= 0:
                continue
              rows.append({
                'period_type': 'month',
                'year_period': f'{year}-{m_idx+1:02d}',
                'region': current_region,
                'factory': current_factory,
                'vehicle_model': _PLANT_AGGREGATE_MODEL,
                'vehicle_type': '',
                'powertrain': None,
                'sales_units': units,
                'source_url': SOURCE_URL,
              })
          continue
        # fall-through to row processing
      # 'Total' (B='Total', C 비어있음) — plant 종료
      if _is_skip_row(s_b) and not s_c:
        current_region = None
        continue
      # Sub-total row (C='Sub-total') — skip
      if _is_skip_row(s_c):
        continue
      # 빈 row
      if not s_c and not (s_b in _FACTORY_REGION_LABELS):
        continue
      # 정상 모델 row
      if current_factory is None or current_region is None:
        continue
      model_name = s_c
      if not model_name:
        continue

      for m_idx in range(12):
        units = _safe_int(ws.cell(r, jan_col + m_idx).value)
        if units <= 0:
          continue
        year_period = f'{year}-{m_idx+1:02d}'
        rows.append({
          'period_type': 'month',
          'year_period': year_period,
          'region': current_region,
          'factory': current_factory,
          'vehicle_model': model_name,
          'vehicle_type': '',  # 공장 엑셀에는 type 미제공
          'powertrain': None,
          'sales_units': units,
          'source_url': SOURCE_URL,
        })
    return rows
  finally:
    wb.close()


# ---------------------------------------------------------------------------
# 엑셀 파서 — export (지역별 수출실적)
# ---------------------------------------------------------------------------
def parse_export_excel(path: Path, year: int) -> list[dict]:
  """지역별 수출실적 엑셀 → hyundai_export_regions 형식 rows.

  구조:
    r3: 'Region' 헤더
    r6~: B=대권역(North America/Europe/Others, merged), C=세부지역(U.S.A./Canada/Mexico/Sub-total/...),
         D~O=Jan~Dec.
    'Sub-total' row → skip (대권역 합계, 세부 region 합으로 재계산 가능).
    r22: 'Total' row → skip.

  반환 row 스키마 (hyundai_export_regions PK: period_type,year_period,source,region_name):
    {period_type='month', year_period='YYYY-MM', source='export-by-region',
     region_name=세부지역, sales_units, source_url}
  대권역(B열)은 사용하지 않음 — 세부 region_name이 PK 일부로 충분.
  """
  wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
  try:
    ws = wb[wb.sheetnames[0]]
    hdr = _find_header_row(ws)
    if hdr is None:
      logger.error(f'{path.name} 헤더 미발견')
      _dump_failed_rows(path, 'export', year, ws)
      return []
    header_row, jan_col = hdr

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
      b = ws.cell(r, 2).value
      c = ws.cell(r, 3).value
      s_b = str(b).strip() if b is not None else ''
      s_c = str(c).strip() if c is not None else ''

      # 'Total' row (B='Total') — 종료
      if s_b.lower() == 'total':
        break
      # Sub-total row (C='Sub-total') — skip
      if _is_skip_row(s_c):
        continue
      # 세부 region row (C 비어있지 않음)
      if not s_c:
        continue
      region_name = s_c
      # 대권역 row(C에 region 없고 B만) — 본 엑셀에선 발생 안 하지만 안전.
      if region_name.lower() == 'total':
        continue

      for m_idx in range(12):
        units = _safe_int(ws.cell(r, jan_col + m_idx).value)
        if units <= 0:
          continue
        year_period = f'{year}-{m_idx+1:02d}'
        rows.append({
          'period_type': 'month',
          'year_period': year_period,
          'source': 'export-by-region',
          'region_name': region_name,
          'sales_units': units,
          'source_url': SOURCE_URL,
        })
    return rows
  finally:
    wb.close()


# ---------------------------------------------------------------------------
# 요약 / dedup
# ---------------------------------------------------------------------------
def dedupe_rows(rows: list[dict]) -> list[dict]:
  """hyundai_sales 동일 PK 충돌 제거 (PK: period_type, year_period, region, factory, vehicle_model).

  같은 파일에서 동일 PK가 두 번 등장하는 경우 (예: factory 파일에서 같은 plant의 Domestic 섹션에 동일 모델이 중복 표기) 뒷 행이 우선.
  """
  by_pk: dict[tuple, dict] = {}
  for r in rows:
    pk = (r['period_type'], r['year_period'], r['region'],
          r['factory'], r['vehicle_model'])
    by_pk[pk] = r
  return list(by_pk.values())


def dedupe_export_rows(rows: list[dict]) -> list[dict]:
  """hyundai_export_regions 동일 PK 충돌 제거 (PK: period_type,year_period,source,region_name)."""
  by_pk: dict[tuple, dict] = {}
  for r in rows:
    pk = (r['period_type'], r['year_period'], r['source'], r['region_name'])
    by_pk[pk] = r
  return list(by_pk.values())


def _split_export_rows(all_rows: list[dict]) -> tuple[list[dict], list[dict]]:
  """parser가 섞어 반환하는 rows를 (sales_rows, export_rows)로 분리.

  hyundai_export_regions row는 source 키가 존재 (export-by-region/ir-summary).
  hyundai_sales row는 region/factory/vehicle_model 키를 가진다.
  """
  sales: list[dict] = []
  exports: list[dict] = []
  for r in all_rows:
    if 'source' in r and r.get('source') == 'export-by-region':
      exports.append(r)
    else:
      sales.append(r)
  return sales, exports


def print_summary(rows: list[dict]) -> None:
  if not rows:
    logger.info('hyundai_sales 적재 대상 행 0건')
    return
  by_year: dict[str, int] = {}
  by_region: dict[str, int] = {}
  by_factory: dict[str, int] = {}
  models: set[str] = set()
  total_units = 0
  for r in rows:
    y = r['year_period'][:4]
    by_year[y] = by_year.get(y, 0) + 1
    by_region[r['region']] = by_region.get(r['region'], 0) + 1
    fkey = r['factory'] or '(국내/지역수출)'
    by_factory[fkey] = by_factory.get(fkey, 0) + 1
    models.add(r['vehicle_model'])
    total_units += r['sales_units']
  logger.info(f'[hyundai_sales] 행 총 {len(rows)}개, 모델 {len(models)}개, 합계 {total_units:,}대')
  logger.info(f'  연도별: {dict(sorted(by_year.items()))}')
  logger.info(f'  region별 (상위 10): {dict(sorted(by_region.items(), key=lambda x: -x[1])[:10])}')
  logger.info(f'  factory별 (상위 10): {dict(sorted(by_factory.items(), key=lambda x: -x[1])[:10])}')


def print_export_summary(rows: list[dict]) -> None:
  if not rows:
    logger.info('hyundai_export_regions 적재 대상 행 0건')
    return
  by_year: dict[str, int] = {}
  by_region: dict[str, int] = {}
  total_units = 0
  for r in rows:
    y = r['year_period'][:4]
    by_year[y] = by_year.get(y, 0) + r['sales_units']
    by_region[r['region_name']] = by_region.get(r['region_name'], 0) + r['sales_units']
    total_units += r['sales_units']
  logger.info(f'[hyundai_export_regions] 행 총 {len(rows)}개, 합계 {total_units:,}대')
  logger.info(f'  연도별 합계: {dict(sorted(by_year.items()))}')
  logger.info(f'  region별 합계 (상위 10): {dict(sorted(by_region.items(), key=lambda x: -x[1])[:10])}')


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
  p = argparse.ArgumentParser(description='현대차 차종/공장/지역 판매 수집.')
  p.add_argument('--year-from', type=int, default=DEFAULT_YEAR_FROM,
                 help=f'백필 시작 연도 (default {DEFAULT_YEAR_FROM})')
  p.add_argument('--year-to', type=int, default=None,
                 help='마지막 연도 (default 현재 연도)')
  p.add_argument('--kind', choices=['model', 'factory', 'export', 'all'],
                 default='all', help='처리할 엑셀 종류 (default all)')
  p.add_argument('--reprocess-all', action='store_true',
                 help='캐시 무시 — 현재는 캐시 미운영, flag 예약')
  p.add_argument('--dry-run', action='store_true',
                 help='DB 쓰기 없이 파싱 결과만 print')
  p.add_argument('--keep-downloads', action='store_true', default=True,
                 help='다운로드 엑셀을 data/_hyundai_downloads/에 보존 (default True)')
  p.add_argument('--use-cache-only', action='store_true',
                 help='Playwright 없이 data/_hyundai_downloads/의 캐시된 .xlsx만 parse (개발/재처리용)')
  return p.parse_args()


def _run_cache_only_pipeline(
  args, year_range: list[int],
) -> tuple[list[dict], list[Path], list[str]]:
  """캐시된 .xlsx만 parse (Playwright 미사용)."""
  kinds = ['model', 'factory', 'export'] if args.kind == 'all' else [args.kind]
  all_rows: list[dict] = []
  used: list[Path] = []
  failed_jobs: list[str] = []
  parsers = {
    'model': parse_model_excel,
    'factory': parse_factory_excel,
    'export': parse_export_excel,
  }
  for year in year_range:
    for kind in kinds:
      path = DOWNLOAD_DIR / f'{year}_{kind}.xlsx'
      if not path.exists():
        logger.warning(f'{year}/{kind}: 캐시 파일 없음 — {path}')
        failed_jobs.append(f'{year}/{kind}/no-cache')
        continue
      used.append(path)
      try:
        year_rows = parsers[kind](path, year)
      except Exception as e:
        logger.exception(f'{year}년 {kind} 파싱 실패: {e}')
        failed_jobs.append(f'{year}/{kind}/parse')
        continue
      logger.info(f'{year}년 {kind}: 파싱 {len(year_rows)}행 (cache)')
      all_rows.extend(year_rows)
  return all_rows, used, failed_jobs


def _run_pipeline(args, year_range: list[int]) -> tuple[list[dict], list[Path], list[str]]:
  """Playwright + 파싱 파이프라인. (all_rows, downloaded_paths, failed_jobs) 반환."""
  kinds = ['model', 'factory', 'export'] if args.kind == 'all' else [args.kind]
  all_rows: list[dict] = []
  downloaded: list[Path] = []
  failed_jobs: list[str] = []

  from playwright.sync_api import sync_playwright  # noqa: E402

  with sync_playwright() as pw:
    browser = pw.chromium.launch()
    ctx = browser.new_context(accept_downloads=True)
    page = ctx.new_page()
    try:
      page.goto(SOURCE_URL, wait_until='domcontentloaded',
                timeout=PLAYWRIGHT_TIMEOUT_MS)
      page.wait_for_load_state('load', timeout=30_000)
      page.wait_for_timeout(3_000)
    except Exception as e:
      logger.error(f'현대 IR 페이지 로드 실패: {e}')
      ctx.close()
      browser.close()
      return [], [], [f'page-load: {e}']

    for year in year_range:
      if not _select_year(page, year):
        failed_jobs.append(f'{year}/year-select')
        continue
      for kind in kinds:
        path = fetch_excel_for_year_kind(page, year, kind, DOWNLOAD_DIR)
        if path is None:
          failed_jobs.append(f'{year}/{kind}/download')
          continue
        downloaded.append(path)
        parser = {
          'model': parse_model_excel,
          'factory': parse_factory_excel,
          'export': parse_export_excel,
        }[kind]
        try:
          year_rows = parser(path, year)
        except Exception as e:
          logger.exception(f'{year}년 {kind} 파싱 실패: {e}')
          failed_jobs.append(f'{year}/{kind}/parse')
          continue
        logger.info(f'{year}년 {kind}: 파싱 {len(year_rows)}행')
        all_rows.extend(year_rows)

    ctx.close()
    browser.close()

  return all_rows, downloaded, failed_jobs


def _maybe_cleanup_downloads(files: list[Path], keep: bool) -> None:
  if keep:
    logger.info(f'다운로드 파일 보존: {len(files)}개 in {DOWNLOAD_DIR}')
    return
  for f in files:
    try:
      f.unlink(missing_ok=True)
    except Exception as e:
      logger.debug(f'  파일 삭제 실패 {f.name}: {e}')


def main() -> int:
  args = parse_args()
  current_year = datetime.now(timezone.utc).year
  year_to = args.year_to or current_year
  year_range = list(range(args.year_from, year_to + 1))
  logger.info(
    f'현대차 판매 수집: 연도 {year_range[0]}~{year_range[-1]} kind={args.kind} '
    f'(dry_run={args.dry_run}, keep_downloads={args.keep_downloads})'
  )

  DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

  if args.use_cache_only:
    logger.info('cache-only mode: Playwright skip, data/_hyundai_downloads/ 파일만 parse')
    all_rows, downloaded, failed_jobs = _run_cache_only_pipeline(args, year_range)
  else:
    all_rows, downloaded, failed_jobs = _run_pipeline(args, year_range)

  sales_rows, export_rows = _split_export_rows(all_rows)
  sales_rows = dedupe_rows(sales_rows)
  export_rows = dedupe_export_rows(export_rows)
  print_summary(sales_rows)
  print_export_summary(export_rows)
  if failed_jobs:
    logger.warning(f'실패 작업({len(failed_jobs)}): {failed_jobs}')

  if args.dry_run:
    logger.success('dry-run 종료 (DB 쓰기 없음)')
    _maybe_cleanup_downloads(downloaded, args.keep_downloads)
    return 0 if not failed_jobs else 1

  if not sales_rows and not export_rows:
    logger.warning('적재할 행 없음 — DB 호출 생략')
    _maybe_cleanup_downloads(downloaded, args.keep_downloads)
    return 1 if failed_jobs else 0

  try:
    with WriteSession() as w:
      if sales_rows:
        w.table('hyundai_sales').upsert(
          sales_rows,
          on_conflict='period_type,year_period,region,factory,vehicle_model',
        ).execute()
        logger.success(f'hyundai_sales upsert 완료: {len(sales_rows)}행')
      if export_rows:
        w.table('hyundai_export_regions').upsert(
          export_rows,
          on_conflict='period_type,year_period,source,region_name',
        ).execute()
        logger.success(f'hyundai_export_regions upsert 완료: {len(export_rows)}행')
  except Exception as e:
    logger.error(f'upsert 실패: {e}')
    _maybe_cleanup_downloads(downloaded, args.keep_downloads)
    return 2

  _maybe_cleanup_downloads(downloaded, args.keep_downloads)
  return 0 if not failed_jobs else 1


if __name__ == '__main__':
  try:
    sys.exit(main())
  except KeyboardInterrupt:
    logger.warning('사용자 중단')
    sys.exit(130)
  except Exception as e:
    logger.exception(f'예기치 못한 오류: {e}')
    sys.exit(1)
