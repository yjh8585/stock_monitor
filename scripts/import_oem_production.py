#!/usr/bin/env python3
"""MarkLines OEM 글로벌 생산량 엑셀 → oem_production_model_country_month 적재.

판매 쪽 `import_oem_sales.py`의 쌍둥이다. 차이는 두 가지뿐:
  1. 메타 컬럼이 **6개**다 (판매는 PowerTrain 포함 7개). 월 컬럼이 index 6부터 시작한다.
  2. 사전 집계 4종을 만들지 않는다 — 생산은 현재 스텔란티스 탭 차트 1만 쓰고,
     조회가 (group, country) 필터라 모델 단위 테이블 직접 조회로 충분하다.
     (판매는 /oem 전체 탭이 group·PT·type/segment 집계를 쓰기 때문에 4종을 둔다.)

⚠️ country = **생산(공장) 국가**. 판매 테이블의 country(판매 시장)와 의미가 다르다.

처리 흐름:
  1. 참고/oem 생산량/MarkLines_product_data*.xlsx 전부 로드
     (연도별 2020~2023 + 최신 파일 2024~. GHA 러너에는 최신 파일만 존재한다 —
      참고/ 가 .gitignore라 연도별 과거 파일은 로컬 1회 적재분이 DB에 남아 있다.)
  2. 행 = (Country, Group, Maker, Type, Segment, Model) + 월별(YYYYMM) production
  3. (group, country, model, year_month) 단위로 합산 후 upsert

재실행 안전: PRIMARY KEY 충돌 시 UPDATE (멱등).

사용법:
  python scripts/import_oem_production.py
"""
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from loguru import logger

sys.path.insert(0, str(Path(__file__).parent))
from lib.bootstrap import init_script  # noqa: E402

init_script(__file__)

from lib.db import upsert_rows  # noqa: E402

EXCEL_DIR = Path(__file__).resolve().parents[1] / '참고' / 'oem 생산량'
EXCEL_GLOB = 'MarkLines_product_data*.xlsx'

#: 생산 엑셀의 메타 컬럼 수. 월 컬럼은 이 다음부터 시작한다.
#: 판매(7, PowerTrain 포함)와 다르므로 상수로 고정해 오프셋 실수를 막는다.
N_META_COLS = 6

#: 엑셀 header(2행)의 메타 컬럼 기대값. 다르면 다른 export를 받은 것.
EXPECTED_META = ('Country', 'Group', 'Maker/Brand', 'Type', 'Segment', 'Model')


def iter_excel_rows(excel_paths: list[Path]):
    """엑셀 파일들에서 (meta dict, year_month, production) yield.

    엑셀 구조: header row=2, 컬럼 0~5 = 메타, 컬럼 6~ = 월별(YYYYMM) 생산대수.
    production이 0/None/문자열이면 스킵한다.

    ⚠️ 월 컬럼은 데이터가 도착하지 않은 **미래 월까지 미리 만들어져 있다**
       (2026-07 실측: 데이터는 202606까지인데 컬럼은 202612까지). 빈 셀을 0으로 적재하면
       "생산 0대"라는 거짓 사실이 DB에 박히므로, 값이 있는 셀만 적재한다.
       그 결과 DB의 '없는 월'은 미도착과 진짜 0을 구분하지 못하지만, 스텔란티스 북미는
       2021년 이후 전 월이 nonzero라 실무상 문제되지 않는다(lib/stellantis-forecast가
       국가별 최신 월의 min()으로 잘라 쓴다).
    """
    for path in excel_paths:
        logger.info(f'엑셀 로딩: {path.name}')
        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            ws = wb['Sheet1']
            header = [c.value for c in ws[2]]
            meta = tuple(header[:N_META_COLS])
            if meta != EXPECTED_META:
                logger.error(f'{path.name}: 메타 컬럼 불일치 — 기대 {EXPECTED_META}, 실제 {meta}')
                raise ValueError(f'{path.name} 메타 컬럼 불일치')
            month_cols = [
                (i, int(v))
                for i, v in enumerate(header)
                if isinstance(v, (int, float)) and v and v > 200000
            ]
            if not month_cols:
                logger.warning(f'{path.name}: 월 컬럼 없음, 스킵')
                continue

            n_rows = 0
            n_cells = 0
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
                country, group, _maker, _vtype, _segment, model = (
                    row[i] or '' for i in range(N_META_COLS)
                )
                if not group:  # Group 없는 행(빈 줄·푸터) 스킵
                    continue
                meta_row = {
                    'country': country.strip() if isinstance(country, str) else '',
                    'oem_group': group.strip() if isinstance(group, str) else '',
                    'model': model.strip() if isinstance(model, str) else '',
                }
                n_rows += 1
                for col_idx, ym in month_cols:
                    v = row[col_idx]
                    if not isinstance(v, (int, float)):
                        continue
                    production = int(v)
                    if production <= 0:
                        continue
                    n_cells += 1
                    yield meta_row, ym, production
            logger.info(f'  → {n_rows}행, {n_cells}개 (모델×월) 셀')
        finally:
            wb.close()


def aggregate(rows_iter) -> dict:
    """(group, country, model, ym) → production 합산.

    같은 키가 여러 행으로 쪼개질 수 있다(Type/Segment가 달라도 모델명이 같은 경우) → 합산.
    """
    model_country_month: dict[tuple[str, str, str, int], int] = defaultdict(int)
    for meta, ym, production in rows_iter:
        if not meta['country'] or not meta['model']:
            continue
        key = (meta['oem_group'], meta['country'], meta['model'], ym)
        model_country_month[key] += production
    return model_country_month


def main() -> int:
    excel_paths = sorted(EXCEL_DIR.glob(EXCEL_GLOB))
    if not excel_paths:
        logger.error(f'엑셀 파일 없음: {EXCEL_DIR}/{EXCEL_GLOB}')
        return 1
    logger.info(f'엑셀 파일 {len(excel_paths)}개 발견')

    model_country_month = aggregate(iter_excel_rows(excel_paths))
    if not model_country_month:
        logger.error('집계 결과 0행 — 엑셀 구조 변경 의심')
        return 1
    months = {ym for (_, _, _, ym) in model_country_month}
    logger.info(
        f'집계 결과: group×country×model×month={len(model_country_month):,} '
        f'({min(months)}~{max(months)})'
    )

    upsert_rows(
        'oem_production_model_country_month',
        [
            {'oem_group': g, 'country': c, 'model': m, 'year_month': ym, 'production': p}
            for (g, c, m, ym), p in model_country_month.items()
        ],
        conflict_cols='oem_group,country,model,year_month',
    )
    logger.success('OEM 생산량 적재 완료')
    return 0


if __name__ == '__main__':
    sys.exit(main())
