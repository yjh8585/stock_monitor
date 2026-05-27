#!/usr/bin/env python3
"""4사(Stellantis NA / KG모빌리티 / 현대차 / 기아) IR 데이터 소스 audit.

PR1 1회용. PR2~5에서 회사별 차종 판매 수집 스크립트·DB 스키마를 finalize하기 위해
실제 데이터 구조(dimension, granularity, powertrain 가용성, 자동화 난이도)를 조사하고
data/_oem_audit_report.md 로 종합 리포트를 남긴다.

DB 쓰기 없음. fetch 실패는 graceful — 한 회사가 실패해도 나머지 진행.

사용:
  scripts/venv/Scripts/python.exe scripts/audit_oem_companies_sources.py

작업 완료 후 scripts/_archive/ 로 이동 예정 (PR1 머지 후).
"""
import argparse
import io
import re
import sys
import traceback
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup
from loguru import logger

from lib.bootstrap import init_script

init_script(__file__)

# -- 상수 ----------------------------------------------------------------------

USER_AGENT = (
  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
  '(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'
)
REQUEST_TIMEOUT = 30
DOWNLOAD_DIR = Path(__file__).resolve().parents[1] / 'data' / '_audit_downloads'
REPORT_PATH = Path(__file__).resolve().parents[1] / 'data' / '_oem_audit_report.md'
EXCEL_SAMPLE_ROWS = 6   # 샘플 행 수
EXCEL_SAMPLE_COLS = 16  # 샘플 컬럼 수

# 4사 시작 URL
STELLANTIS_SEARCH_URL = 'https://www.prnewswire.com/search/news/?keyword=fca+us&pagesize=25'
KG_MOBILITY_URL = 'https://www.kg-mobility.com/cm/ir-data/sales-performance'
HYUNDAI_URL = 'https://www.hyundai.com/worldwide/ko/company/ir/ir-resources/sales-results'
KIA_URL = 'https://worldwide.kia.com/ko/company/investor-relations/library/performance-and-plans'

# powertrain 키워드 (영문/한글) — Excel/HTML 어디서든 발견 시 "가능성 있음"으로 기록
POWERTRAIN_KEYWORDS = [
  'powertrain', 'power train', 'engine', 'fuel',
  'ev', 'hev', 'phev', 'fcev', 'bev', 'ice', 'hybrid', 'electric',
  '하이브리드', '전기', '내연', '연료전지', '플러그인',
]


def build_session() -> requests.Session:
  """공용 requests 세션."""
  s = requests.Session()
  s.headers.update({
    'User-Agent': USER_AGENT,
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.9,en;q=0.8',
  })
  return s


def safe_fetch(session: requests.Session, url: str, *, stream: bool = False) -> requests.Response | None:
  """fetch 실패 시 None 반환 + 로그. 호출자는 None 처리."""
  try:
    r = session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True, stream=stream)
    if r.status_code != 200:
      logger.warning(f'  HTTP {r.status_code} {url}')
      return None
    return r
  except Exception as e:
    logger.warning(f'  fetch 실패 {url}: {e}')
    return None


def download_file(session: requests.Session, url: str, dest: Path) -> Path | None:
  """파일 다운로드 → dest 경로 저장. 실패 시 None."""
  r = safe_fetch(session, url, stream=True)
  if r is None:
    return None
  ctype = (r.headers.get('Content-Type') or '').lower()
  if 'html' in ctype and 'xlsx' in dest.suffix.lower():
    logger.warning(f'  Content-Type=html, 엑셀 아님 (url={url})')
    return None
  dest.parent.mkdir(parents=True, exist_ok=True)
  with dest.open('wb') as f:
    for chunk in r.iter_content(64 * 1024):
      if chunk:
        f.write(chunk)
  size_kb = dest.stat().st_size / 1024
  logger.info(f'  다운로드 OK: {dest.name} ({size_kb:.1f} KB)')
  return dest


def find_powertrain_hits(text: str) -> list[str]:
  """텍스트에서 powertrain 관련 키워드 hit 추출(중복 제거, 소문자)."""
  if not text:
    return []
  low = text.lower()
  hits = sorted({kw for kw in POWERTRAIN_KEYWORDS if kw in low})
  return hits


def sample_excel(path: Path) -> dict:
  """엑셀 sheet별 첫 N행 × M열 sample + powertrain 단서 추출.

  반환: {'sheets': [{'name', 'dims', 'sample_rows', 'powertrain_hits'}], 'error'?}
  """
  out: dict = {'sheets': []}
  try:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
  except Exception as e:
    return {'sheets': [], 'error': f'엑셀 열기 실패: {e}'}

  try:
    for sname in wb.sheetnames:
      ws = wb[sname]
      sample_rows: list[list[str]] = []
      flat_text_parts: list[str] = []
      row_iter = ws.iter_rows(
        min_row=1, max_row=EXCEL_SAMPLE_ROWS,
        min_col=1, max_col=EXCEL_SAMPLE_COLS,
        values_only=True,
      )
      for row in row_iter:
        line = [('' if v is None else str(v)) for v in row]
        sample_rows.append(line)
        flat_text_parts.extend(line)
      hits = find_powertrain_hits(' '.join(flat_text_parts))
      out['sheets'].append({
        'name': sname,
        'dims': f'{ws.max_row}x{ws.max_column}',
        'sample_rows': sample_rows,
        'powertrain_hits': hits,
      })
  finally:
    wb.close()

  return out


def fmt_sheet_section(sheet: dict) -> str:
  """엑셀 sheet sample → 마크다운 섹션."""
  lines = [f"  - **시트 `{sheet['name']}`** (dims: {sheet['dims']})"]
  pt = sheet['powertrain_hits']
  if pt:
    lines.append(f"    - powertrain 단서: `{', '.join(pt)}`")
  else:
    lines.append('    - powertrain 단서: **없음**')
  lines.append('    - 샘플:')
  lines.append('      ```')
  for row in sheet['sample_rows']:
    # cell 값이 너무 길면 30자로 자름
    trimmed = [c[:30] for c in row]
    lines.append('      ' + ' | '.join(trimmed))
  lines.append('      ```')
  return '\n'.join(lines)


# -- 회사별 audit -------------------------------------------------------------

def audit_stellantis_na(session: requests.Session) -> dict:
  """Stellantis NA (구 FCA US): prnewswire 검색 → 분기 보도자료 본문 표 분석."""
  result: dict = {
    'company': 'Stellantis NA (FCA US)',
    'source_root': STELLANTIS_SEARCH_URL,
    'notes': [],
    'releases': [],
  }
  logger.info(f"[Stellantis] 검색 페이지 fetch: {STELLANTIS_SEARCH_URL}")
  r = safe_fetch(session, STELLANTIS_SEARCH_URL)
  if r is None:
    result['notes'].append('검색 페이지 fetch 실패')
    return result

  soup = BeautifulSoup(r.text, 'html.parser')
  candidates: list[tuple[str, str]] = []
  for a in soup.find_all('a', href=True):
    href = a['href']
    if '/news-releases/' not in href:
      continue
    text = a.get_text(' ', strip=True)
    low = text.lower()
    # FCA US 또는 Stellantis NA 분기/연간 sales 보도자료만
    if 'fca' in low and ('quarter' in low or 'sales' in low or 'results' in low):
      abs_url = urljoin('https://www.prnewswire.com/', href)
      candidates.append((abs_url, text))
  # 중복 제거(URL 기준)
  uniq: dict[str, str] = {}
  for url, text in candidates:
    if url not in uniq:
      uniq[url] = text
  releases = list(uniq.items())[:5]
  result['notes'].append(f'검색 결과 후보 {len(releases)}건')

  if not releases:
    result['notes'].append('FCA US 키워드 후보 없음 — 최근 prnewswire가 명칭 변경 가능')
    return result

  # archive 가용성 단서: 검색 결과 후보 URL 다 기록
  result['notes'].append('archive 단서(검색 결과 후보 URL):')
  for i, (u, t) in enumerate(releases):
    result['notes'].append(f'  {i+1}. {t[:90]} → {u}')

  # 첫 1건 본문 fetch (전체 12행까지 sample)
  url0, title0 = releases[0]
  logger.info(f"[Stellantis] 첫 보도자료 fetch: {title0[:80]}")
  r2 = safe_fetch(session, url0)
  if r2 is None:
    result['notes'].append(f'본문 fetch 실패: {url0}')
    result['releases'] = [{'url': url0, 'title': title0}]
    return result

  soup2 = BeautifulSoup(r2.text, 'html.parser')
  tables = soup2.find_all('table')
  table_summaries = []
  for ti, table in enumerate(tables[:5]):
    rows = table.find_all('tr')
    sample = []
    for tr in rows[:12]:  # 6 → 12로 확대 (모델 더 많이)
      cells = [td.get_text(' ', strip=True)[:30] for td in tr.find_all(['th', 'td'])]
      sample.append(cells)
    table_text = table.get_text(' ', strip=True)
    table_summaries.append({
      'index': ti,
      'rows_total': len(rows),
      'sample': sample,
      'powertrain_hits': find_powertrain_hits(table_text),
    })

  body_text = soup2.get_text(' ', strip=True)
  body_hits = find_powertrain_hits(body_text)

  # 분기 보고서 구조 단서 (Q1/Q2/Q3/Q4 키워드 출현)
  quarter_hits = sorted(set(re.findall(r'(?i)Q[1-4]\s+\d{4}', body_text)))[:8]

  result['releases'].append({
    'url': url0,
    'title': title0,
    'tables_total': len(tables),
    'tables_sampled': table_summaries,
    'body_powertrain_hits': body_hits,
    'quarter_keywords': quarter_hits,
  })
  result['notes'].append(
    f"본문 표 {len(tables)}개, body powertrain 키워드: "
    f"{body_hits or '없음'}, 분기 키워드: {quarter_hits or '없음'}"
  )
  return result


def collect_links_via_playwright(
  url: str, *, wait_selector: str | None = None, wait_timeout_ms: int = 15000,
) -> dict:
  """Playwright로 페이지 렌더링 후 모든 <a> 링크 + 다운로드 이벤트 감지.

  반환:
    {
      'rendered_html_len': int,
      'all_links': [(href, text)],
      'excel_links': [(href, text)],
      'download_urls': [url, ...],   # download 이벤트로 캡처된 실제 다운로드 URL
      'download_buttons': [(selector_hint, text)],  # 다운로드 후보 버튼/링크
      'select_options': [(name, [option_text...])],  # 연도 등 select 요소
      'error': str | None,
    }
  실패 시 'error'에 사유 기록.
  """
  out: dict = {
    'rendered_html_len': 0,
    'all_links': [],
    'excel_links': [],
    'download_urls': [],
    'download_buttons': [],
    'select_options': [],
    'error': None,
  }
  try:
    from playwright.sync_api import sync_playwright
  except Exception as e:
    out['error'] = f'playwright import 실패: {e}'
    return out

  try:
    with sync_playwright() as p:
      browser = p.chromium.launch(headless=True)
      context = browser.new_context(
        user_agent=USER_AGENT,
        locale='ko-KR',
        accept_downloads=True,
      )
      page = context.new_page()

      def on_download(d):
        out['download_urls'].append(d.url)

      page.on('download', on_download)
      try:
        page.goto(url, timeout=wait_timeout_ms, wait_until='domcontentloaded')
      except Exception as e:
        out['error'] = f'page.goto 실패: {e}'
        context.close()
        browser.close()
        return out

      # 페이지 안정화: SPA가 fetch한 데이터 그리는 동안 잠깐 대기
      try:
        page.wait_for_load_state('networkidle', timeout=wait_timeout_ms)
      except Exception:
        pass  # networkidle 도달 못 해도 진행
      if wait_selector:
        try:
          page.wait_for_selector(wait_selector, timeout=wait_timeout_ms)
        except Exception:
          pass

      html = page.content()
      out['rendered_html_len'] = len(html)
      anchors = page.eval_on_selector_all(
        'a',
        '''els => els.map(a => [a.getAttribute('href') || '', (a.innerText || '').trim().slice(0, 80)])''',
      )
      for href, text in anchors:
        if not href:
          continue
        out['all_links'].append((href, text))
        if re.search(r'\.(xlsx?|xls)(\?|$)', href, re.IGNORECASE):
          abs_url = urljoin(url, href)
          out['excel_links'].append((abs_url, text or '(no text)'))

      # 다운로드 버튼 후보 — onclick / class / 텍스트가 download/excel/엑셀/다운로드/내려받기
      try:
        btns = page.eval_on_selector_all(
          'a, button',
          r'''els => els.filter(e => {
            const t = ((e.innerText || '') + ' ' + (e.getAttribute('class') || '') + ' ' + (e.getAttribute('onclick') || '')).toLowerCase();
            return /(download|excel|엑셀|다운로드|내려받기|xls)/i.test(t);
          }).slice(0, 20).map(e => [(e.tagName + (e.getAttribute('class') ? '.' + e.getAttribute('class').split(/\s+/).join('.') : '')).toLowerCase().slice(0, 80), (e.innerText || '').trim().slice(0, 60)])''',
        )
        out['download_buttons'] = btns
      except Exception:
        pass
      # select 요소 (연도 등)
      try:
        selects = page.eval_on_selector_all(
          'select',
          '''els => els.slice(0, 10).map(s => [s.getAttribute('name') || s.getAttribute('id') || '(unnamed)', Array.from(s.options).slice(0, 10).map(o => (o.textContent || '').trim()).filter(Boolean)])''',
        )
        out['select_options'] = selects
      except Exception:
        pass

      context.close()
      browser.close()
  except Exception as e:
    out['error'] = f'playwright 실행 실패: {e}'
  return out


def audit_kg_mobility(session: requests.Session) -> dict:
  """KG모빌리티: sales-performance 페이지에서 엑셀 링크 추출 → 첫 1개 다운로드/파싱.

  requests로 SPA shell만 받으면 Playwright로 렌더링 후 재시도.
  """
  result: dict = {
    'company': 'KG모빌리티',
    'source_root': KG_MOBILITY_URL,
    'notes': [],
    'files': [],
  }
  logger.info(f"[KG] 페이지 fetch: {KG_MOBILITY_URL}")
  r = safe_fetch(session, KG_MOBILITY_URL)
  excel_links: list[tuple[str, str]] = []
  if r is not None:
    soup = BeautifulSoup(r.text, 'html.parser')
    for a in soup.find_all('a', href=True):
      href = a['href'].strip()
      text = a.get_text(' ', strip=True)
      if re.search(r'\.(xlsx?|xls)(\?|$)', href, re.IGNORECASE):
        excel_links.append((urljoin(KG_MOBILITY_URL, href), text or '(no text)'))
    result['notes'].append(
      f'requests 응답 {len(r.text)} chars, 정적 엑셀 링크 {len(excel_links)}건'
    )
  else:
    result['notes'].append('requests fetch 실패')

  # Playwright 보강
  if not excel_links:
    logger.info('[KG] Playwright 보강 시도')
    pw = collect_links_via_playwright(KG_MOBILITY_URL)
    if pw['error']:
      result['notes'].append(f'Playwright 보강 실패: {pw["error"]}')
    else:
      result['notes'].append(
        f"Playwright 렌더링 후 HTML {pw['rendered_html_len']} chars, "
        f"링크 {len(pw['all_links'])}개, 엑셀 {len(pw['excel_links'])}건, "
        f"download 이벤트 {len(pw['download_urls'])}건"
      )
      excel_links = pw['excel_links']
      # download_urls는 사용자 클릭 시뮬레이션 없이는 보통 안 잡힘. 참고용.
      if pw['download_buttons']:
        result['notes'].append(
          f"다운로드 버튼 후보 {len(pw['download_buttons'])}개 (예: "
          f"{pw['download_buttons'][0][0]} / 텍스트={pw['download_buttons'][0][1]!r})"
        )
      if pw['select_options']:
        opts_summary = ', '.join(
          f"{name}=[{', '.join(opts[:4])}...]" for name, opts in pw['select_options'][:3]
        )
        result['notes'].append(f'select 요소: {opts_summary}')

  if not excel_links:
    result['notes'].append(
      '정적/렌더링 모두 엑셀 링크 미발견 — JS click 필요. PR2에서 page.click(연도) → '
      'page.expect_download()로 캡처 필요. KG 사이트는 ASP.NET WebForms 추정(__doPostBack 가능)'
    )
    return result

  # 첫 1개 다운로드
  url0, label0 = excel_links[0]
  ext = '.xlsx' if 'xlsx' in url0.lower() else '.xls'
  fname = f'kg_mobility_audit{ext}'
  dest = DOWNLOAD_DIR / fname
  saved = download_file(session, url0, dest)
  if saved:
    sample = sample_excel(saved)
    result['files'].append({
      'url': url0,
      'label': label0,
      'local': str(saved.relative_to(REPORT_PATH.parent.parent)),
      'sample': sample,
    })
  else:
    result['files'].append({'url': url0, 'label': label0, 'error': '다운로드 실패'})

  return result


def audit_excel_site(
  session: requests.Session, *, company: str, url: str, fname_prefix: str,
) -> dict:
  """현대/기아 공통 패턴 — requests 우선, 실패 시 Playwright 보강."""
  result: dict = {
    'company': company,
    'source_root': url,
    'notes': [],
    'files': [],
  }
  logger.info(f"[{company}] 페이지 fetch: {url}")
  r = safe_fetch(session, url)
  excel_links: list[tuple[str, str]] = []
  if r is not None:
    soup = BeautifulSoup(r.text, 'html.parser')
    for a in soup.find_all('a', href=True):
      href = a['href'].strip()
      text = a.get_text(' ', strip=True)
      if re.search(r'\.(xlsx?|xls)(\?|$)', href, re.IGNORECASE):
        excel_links.append((urljoin(url, href), text or '(no text)'))
    result['notes'].append(
      f'requests 응답 {len(r.text)} chars, 정적 엑셀 링크 {len(excel_links)}건'
    )

    # SPA / 연도 단서
    raw = r.text
    spa_hints: list[str] = []
    if re.search(r'(__next|window\.__data|hydrate|__nuxt)', raw, re.IGNORECASE):
      spa_hints.append('SPA 마커 감지')
    if re.search(r'(performance|판매실적|수출실적|sales[_-]?result)', raw, re.IGNORECASE):
      spa_hints.append('실적 키워드 본문 노출')
    if spa_hints:
      result['notes'].append('단서: ' + ' / '.join(spa_hints))
  else:
    result['notes'].append('requests fetch 실패')

  if not excel_links:
    logger.info(f'[{company}] Playwright 보강 시도')
    pw = collect_links_via_playwright(url)
    if pw['error']:
      result['notes'].append(f'Playwright 보강 실패: {pw["error"]}')
    else:
      result['notes'].append(
        f"Playwright 렌더링 후 HTML {pw['rendered_html_len']} chars, "
        f"링크 {len(pw['all_links'])}개, 엑셀 {len(pw['excel_links'])}건"
      )
      excel_links = pw['excel_links']
      # 데이터 없는 SPA 흔적이라도 다른 패턴(.do?cmd=download 등) 단서 살펴봄
      asp_pattern_hits = [
        (h, t) for h, t in pw['all_links']
        if re.search(r'(download|file|excel|attach)', h, re.IGNORECASE)
      ][:5]
      if asp_pattern_hits:
        result['notes'].append(
          f'다운로드 패턴 후보 {len(asp_pattern_hits)}개 (예: {asp_pattern_hits[0][0][:120]})'
        )
      if pw['download_buttons']:
        # 상위 3개 버튼 텍스트만 미리보기
        sample = ' / '.join(
          f"{sel[:40]}={text!r}" for sel, text in pw['download_buttons'][:3]
        )
        result['notes'].append(
          f"다운로드 버튼 후보 {len(pw['download_buttons'])}개: {sample}"
        )
      if pw['select_options']:
        opts_summary = ' / '.join(
          f"{name}=[{', '.join(opts[:5])}{'...' if len(opts) > 5 else ''}]"
          for name, opts in pw['select_options'][:3]
        )
        result['notes'].append(f'select 요소: {opts_summary}')

  if not excel_links:
    result['notes'].append(
      'requests/Playwright 모두 직접 엑셀 링크 미발견 — 연도 select + 다운로드 버튼 클릭 필요. '
      'PR3/4에서 page.select_option() → page.click() → expect_download() 패턴 권장.'
    )
    return result

  for url0, label0 in excel_links[:2]:
    ext = '.xlsx' if 'xlsx' in url0.lower() else '.xls'
    safe_label = re.sub(r'[^a-zA-Z0-9_-]+', '_', label0)[:40] or 'unnamed'
    fname = f'{fname_prefix}_{safe_label}{ext}'
    dest = DOWNLOAD_DIR / fname
    saved = download_file(session, url0, dest)
    if saved:
      sample = sample_excel(saved)
      result['files'].append({
        'url': url0,
        'label': label0,
        'local': str(saved.relative_to(REPORT_PATH.parent.parent)),
        'sample': sample,
      })
    else:
      result['files'].append({'url': url0, 'label': label0, 'error': '다운로드 실패'})

  return result


def audit_hyundai(session: requests.Session) -> dict:
  """현대차: ir-resources/sales-results 페이지 분석."""
  return audit_excel_site(
    session, company='현대차', url=HYUNDAI_URL, fname_prefix='hyundai',
  )


def audit_kia(session: requests.Session) -> dict:
  """기아: library/performance-and-plans 페이지 분석."""
  return audit_excel_site(
    session, company='기아', url=KIA_URL, fname_prefix='kia',
  )


# -- 리포트 생성 --------------------------------------------------------------

def render_stellantis_section(d: dict) -> str:
  lines = [f"## {d['company']}", f"- **소스 루트**: {d['source_root']}", '']
  for n in d['notes']:
    lines.append(f"- {n}")
  if not d.get('releases'):
    lines.append('')
    return '\n'.join(lines)
  lines.append('')
  lines.append('### 보도자료 분석')
  for rel in d['releases']:
    lines.append(f"- **{rel.get('title', '(no title)')}**")
    lines.append(f"  - URL: {rel['url']}")
    if 'tables_total' in rel:
      lines.append(f"  - 본문 표 수: {rel['tables_total']}")
      lines.append(
        f"  - 본문 powertrain 키워드: "
        f"{', '.join(rel['body_powertrain_hits']) if rel['body_powertrain_hits'] else '없음'}"
      )
      for t in rel['tables_sampled']:
        lines.append(
          f"  - 표 #{t['index']} (rows={t['rows_total']}, "
          f"powertrain hits: {', '.join(t['powertrain_hits']) or '없음'})"
        )
        if t['sample']:
          lines.append('    ```')
          for row in t['sample']:
            lines.append('    ' + ' | '.join(row))
          lines.append('    ```')
  return '\n'.join(lines)


def render_excel_company_section(d: dict) -> str:
  lines = [f"## {d['company']}", f"- **소스 루트**: {d['source_root']}", '']
  for n in d['notes']:
    lines.append(f"- {n}")
  if not d.get('files'):
    lines.append('')
    return '\n'.join(lines)
  lines.append('')
  lines.append('### 다운로드 파일 분석')
  for f in d['files']:
    lines.append(f"- **{f.get('label', '(no label)')}**")
    lines.append(f"  - URL: {f['url']}")
    if 'error' in f:
      lines.append(f"  - 에러: {f['error']}")
      continue
    if 'local' in f:
      lines.append(f"  - 로컬: `{f['local']}`")
    sample = f.get('sample', {})
    if sample.get('error'):
      lines.append(f"  - 샘플 에러: {sample['error']}")
      continue
    for sheet in sample.get('sheets', []):
      lines.append(fmt_sheet_section(sheet))
  return '\n'.join(lines)


def render_overall_assessment(results: list[dict]) -> str:
  """4사 종합 평가 — 자동화 난이도 + powertrain 가용성 + PR2~5 권고."""
  lines = ['## 종합 평가', '']

  # 자동화 난이도 표
  lines.append('### 자동화 난이도 (PR2~5 수집 스크립트 작성용)')
  lines.append('')
  lines.append('| 회사 | 정적 링크 가용 | 자동화 난이도 | 권장 도구 |')
  lines.append('|------|----------------|---------------|-----------|')
  for r in results:
    company = r['company']
    files_n = len(r.get('files', []))
    releases_n = len(r.get('releases', []))
    has_static = files_n > 0 or releases_n > 0
    if 'fetch 실패' in ' '.join(r['notes']):
      level = '미상 (fetch 실패)'
      tool = 'Playwright 또는 수동 검토'
    elif has_static:
      level = 'LOW (requests + bs4)'
      tool = 'requests + openpyxl/bs4'
    else:
      level = 'MEDIUM~HIGH'
      tool = 'Playwright (브라우저 렌더링)'
    static_label = 'O' if has_static else 'X'
    lines.append(f'| {company} | {static_label} | {level} | {tool} |')

  # powertrain 가용성
  lines.append('')
  lines.append('### Powertrain 정보 가용성')
  lines.append('')
  lines.append('| 회사 | 엑셀/표 단서 | 본문 단서 | 평가 |')
  lines.append('|------|--------------|-----------|------|')
  for r in results:
    sheet_hits: set[str] = set()
    body_hits: set[str] = set()
    for f in r.get('files', []):
      for s in f.get('sample', {}).get('sheets', []):
        sheet_hits.update(s.get('powertrain_hits', []))
    for rel in r.get('releases', []):
      body_hits.update(rel.get('body_powertrain_hits', []))
      for t in rel.get('tables_sampled', []):
        sheet_hits.update(t.get('powertrain_hits', []))
    has_strong = bool({'powertrain', 'power train'} & sheet_hits)
    has_indirect = bool(sheet_hits | body_hits)
    if has_strong:
      verdict = '**가능** (powertrain 컬럼 추정)'
    elif has_indirect:
      verdict = '간접 (모델명 기반 매핑 필요)'
    else:
      verdict = '**없음** (모델→PT 매핑 테이블 별도)'
    lines.append(
      f"| {r['company']} | {', '.join(sorted(sheet_hits)) or '-'} | "
      f"{', '.join(sorted(body_hits)) or '-'} | {verdict} |"
    )

  lines.append('')
  lines.append('### PR2~5 finalize 권고')
  lines.append('')
  lines.append('**1. Powertrain 전략**')
  lines.append('- 4사 모두 엑셀/표에 powertrain 컬럼 **없음**. Stellantis 본문에 정성적 언급만.')
  lines.append('- 차종→powertrain 매핑 테이블(`vehicle_powertrain_map`, plan에 정의)을 회사별로 운영해야 함.')
  lines.append('- Phase 1(PR2~5): best-effort — 매핑 안 된 차종은 `powertrain=NULL` 허용.')
  lines.append('  매핑 컴포넌트(`CompanyPowertrainMixChart`)는 NULL 비중을 별도 카테고리로 표시.')
  lines.append('- Phase 2: 회사 IR 보도자료/연간보고서에서 모델별 powertrain 정보를 사람이 1회 정리해 시드.')
  lines.append('')
  lines.append('**2. Granularity 확정**')
  lines.append('- Stellantis NA: **분기** 단위만 (보도자료). 월 단위 미제공.')
  lines.append('- 현대/기아: **월 단위 + 연간** 양쪽. 다운로드 버튼이 "YYYY년 ..." 형태로 연도 단위 파일.')
  lines.append('- KG모빌리티: 페이지 라벨 미확보지만 통상 **월 단위** (매월 갱신 IR 표준).')
  lines.append('- 공통 컴포넌트 `CompanyTimeSeriesChart`는 month/quarter 양쪽 처리 필수.')
  lines.append('')
  lines.append('**3. 자동화 전략 (Playwright)**')
  lines.append('- 정적 엑셀 링크는 4사 중 0사. 모두 `expect_download()` 패턴 필요.')
  lines.append('- 발견된 selector 패턴 (audit 본문 참고):')
  lines.append('  - KG: `button.btn.file-down` (텍스트 "다운로드")')
  lines.append('  - 현대: `button.btn-download` (텍스트 "YYYY년 차종별 매출실적파일 다운로드" 등 3종)')
  lines.append('  - 기아: audit에서 직접 selector 미감지 — PR4 작업자가 페이지 수동 검사 1회 필요')
  lines.append('- 권장 PR 패턴:')
  lines.append('  ```python')
  lines.append('  with page.expect_download(timeout=30000) as dl_info:')
  lines.append('      page.click("button.btn-download:has-text(\'2025년 차종별\')")')
  lines.append('  download = dl_info.value')
  lines.append('  download.save_as(dest_path)')
  lines.append('  ```')
  lines.append('- GHA workflow에서 Playwright 브라우저 다운로드: `microsoft/playwright-github-action`')
  lines.append('  사용 (`scripts/sync_oem_excel.py`가 이미 `MARKLINES_COOKIE` 패턴 — 참고).')
  lines.append('')
  lines.append('**4. Archive (2021년 이전)**')
  lines.append('- Stellantis: prnewswire는 영구 URL — Q1/Q2/Q3/Q4 검색으로 2018+ 백필 가능.')
  lines.append('  검색 후보 URL 패턴 활용. 사이트 robots/rate limit 준수.')
  lines.append('- 현대/기아: 페이지에 보통 최근 2~3년만 노출. 더 옛 데이터는')
  lines.append('  `web.archive.org` 스냅샷 또는 회사 IR 부서 직접 문의.')
  lines.append('- KG모빌리티: 동일 — 페이지 옛 연도는 selector로 노출 안 됨.')
  lines.append('')
  lines.append('**5. DB 스키마 (plan 대비 영향)**')
  lines.append('- 현대 `factory text` 컬럼: 다운로드 파일명에 "해외 공장별 판매" 확인 — 컬럼 그대로 OK.')
  lines.append('- 기아 `export_region text` 컬럼: audit 미확정. 페이지에 "지역별 수출" 키워드는 본문 노출되나')
  lines.append('  버튼 텍스트는 미감지 — PR4에서 실제 파일 확인 후 컬럼명 finalize.')
  lines.append('- Stellantis `brand text` 컬럼: 표 #0에 모델만 있고 Jeep/Ram 그룹 헤더는 본문 텍스트 추정.')
  lines.append('  PR5에서 표 파싱 시 모델 → 브랜드 매핑(예: Compass/Wrangler/Gladiator → Jeep) 필요.')
  lines.append('- `vehicle_powertrain_map` 테이블: 4사 모두 필요. PR2(KG)에서 함께 마이그레이션 도입 권장.')
  return '\n'.join(lines)


def build_report(results: list[dict]) -> str:
  now = datetime.now(timezone.utc).astimezone().strftime('%Y-%m-%d %H:%M %Z')
  header = [
    '# OEM 4사 데이터 소스 Audit 리포트',
    '',
    f'- 생성: {now}',
    f'- 스크립트: `scripts/audit_oem_companies_sources.py` (PR1 1회용)',
    f'- 다운로드 디렉토리: `data/_audit_downloads/`',
    '',
    '> 본 리포트는 PR2~5 회사별 차종 판매 수집 스크립트·DB 스키마 finalize를 위한 1회 audit 결과다.',
    '> 자동 fetch만 시도했으므로 일부 회사는 Playwright/로그인 필요 — 본문 "자동화 난이도" 표 참고.',
    '',
  ]
  sections: list[str] = []
  for r in results:
    if r['company'].startswith('Stellantis'):
      sections.append(render_stellantis_section(r))
    else:
      sections.append(render_excel_company_section(r))
  overall = render_overall_assessment(results)
  return '\n\n'.join(header + sections + [overall, ''])


# -- 메인 --------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
  p = argparse.ArgumentParser(description='OEM 4사 데이터 소스 audit (PR1 1회용)')
  p.add_argument('--skip', nargs='*', default=[], help='건너뛸 회사 키 (stellantis/kg/hyundai/kia)')
  return p.parse_args()


def main() -> int:
  args = parse_args()
  DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

  session = build_session()
  results: list[dict] = []

  audits = [
    ('stellantis', audit_stellantis_na),
    ('kg', audit_kg_mobility),
    ('hyundai', audit_hyundai),
    ('kia', audit_kia),
  ]

  for key, fn in audits:
    if key in args.skip:
      logger.info(f'[{key}] skip')
      continue
    logger.info(f'=== {key} audit 시작 ===')
    try:
      r = fn(session)
    except Exception as e:
      logger.error(f'[{key}] audit 실패: {e}\n{traceback.format_exc()}')
      r = {
        'company': key,
        'source_root': '(unknown)',
        'notes': [f'audit 실패: {e}'],
      }
    results.append(r)

  report = build_report(results)
  REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
  REPORT_PATH.write_text(report, encoding='utf-8', newline='\n')
  logger.success(f'리포트 작성: {REPORT_PATH}')
  return 0


if __name__ == '__main__':
  try:
    sys.exit(main())
  except Exception as e:
    logger.error(f'audit 실패: {e}')
    sys.exit(1)
