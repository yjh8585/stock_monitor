#!/usr/bin/env python3
"""model.xlsx 구조 확인."""
import sys
from pathlib import Path
import openpyxl
from lib.bootstrap import init_script

init_script(__file__)

DOWNLOAD_DIR = Path(__file__).resolve().parent.parent / 'data' / '_hyundai_downloads'


def inspect(year: int) -> None:
  path = DOWNLOAD_DIR / f'{year}_model.xlsx'
  wb = openpyxl.load_workbook(path, data_only=True)
  ws = wb[wb.sheetnames[0]]
  print(f'\n=== {year}_model.xlsx (max_row={ws.max_row}) ===')
  for r in range(1, ws.max_row + 1):
    b = ws.cell(r, 2).value
    c = ws.cell(r, 3).value
    total_p = ws.cell(r, 16).value
    sb = str(b or '').strip()
    sc = str(c or '').strip()
    if sb or sc:
      print(f'  r{r}: B="{sb}" C="{sc}" P(Total)={total_p}')


if __name__ == '__main__':
  inspect(2025)
