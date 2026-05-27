"""'연간', '연결_월', '월' 3개 시트의 GM 직수출 행을 sil별로 집계."""

from __future__ import annotations

from collections import defaultdict
from openpyxl import load_workbook

XLSX = r"참고\손익\자료정리_월별손익.xlsx"


def inspect(ws, sheet_name: str, customer_col_idx: int, sil_col_idx: int,
            year_col_idx: int, revenue_col_idx: int, month_col_idx: int | None = None) -> None:
    """customer='GM 직수출' 행을 sil × year로 집계."""
    print(f"\n=== {sheet_name} ===")
    agg: dict[tuple[int, str], dict] = defaultdict(lambda: {"rows": 0, "revenue": 0.0, "months": set()})
    for row in ws.iter_rows(min_row=4, values_only=True):
        customer = row[customer_col_idx - 1] if customer_col_idx - 1 < len(row) else None
        if customer != "GM 직수출":
            continue
        year = row[year_col_idx - 1]
        sil = row[sil_col_idx - 1]
        revenue = row[revenue_col_idx - 1]
        try:
            y = int(year)
            rev = float(revenue or 0.0)
        except (TypeError, ValueError):
            continue
        key = (y, str(sil) if sil else "(empty)")
        agg[key]["rows"] += 1
        agg[key]["revenue"] += rev
        if month_col_idx is not None:
            m = row[month_col_idx - 1]
            try:
                agg[key]["months"].add(int(m))
            except (TypeError, ValueError):
                pass

    for k in sorted(agg.keys()):
        y, sil = k
        d = agg[k]
        months_str = f" months={sorted(d['months'])}" if d["months"] else ""
        print(f"  {y} / {sil:>5}: {d['rows']:>3}행, 매출 {d['revenue']:>12,.2f}{months_str}")


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)

    # 연간 시트: 연도(B=2), 기준(C=3), 실(D=4), ..., 거래처(H=8), 매출(I=9). monthly 없음.
    inspect(wb["연간"], "연간 시트",
            customer_col_idx=8, sil_col_idx=4, year_col_idx=2, revenue_col_idx=9)

    # 연결_월 시트: 기간(B=2), 연도(C=3), 월(D=4), 기준(E=5), 실(F=6), ..., 거래처(J=10), 매출(K=11)
    inspect(wb["연결_월"], "연결_월 시트",
            customer_col_idx=10, sil_col_idx=6, year_col_idx=3, revenue_col_idx=11, month_col_idx=4)

    # 월 시트: 동일 구조
    inspect(wb["월"], "월 시트",
            customer_col_idx=10, sil_col_idx=6, year_col_idx=3, revenue_col_idx=11, month_col_idx=4)


if __name__ == "__main__":
    main()
