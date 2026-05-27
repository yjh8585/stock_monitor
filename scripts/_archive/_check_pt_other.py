"""PowerTrain 원본 고유 값 전수 조사 — Other 분류 원인 파악."""
import openpyxl
from pathlib import Path
from collections import defaultdict

EXCEL_DIR = Path(__file__).resolve().parents[1] / '참고' / 'oem 판매량'
PT_TOKEN_MAP = {'EV': 'EV', 'FCV': 'FCV', 'PHV': 'PHEV', 'HV': 'HV', 'MHV': 'HV', 'ICE': 'ICE'}
PT_PRIORITY = ['EV', 'FCV', 'PHEV', 'HV', 'ICE']


def normalize(raw: str) -> str:
    if not raw or raw == 'N/A':
        return 'Other'
    tokens = [t.strip().upper() for t in str(raw).split('/') if t.strip()]
    mapped = {PT_TOKEN_MAP[t] for t in tokens if t in PT_TOKEN_MAP}
    if not mapped:
        return 'Other'
    for pt in PT_PRIORITY:
        if pt in mapped:
            return pt
    return 'Other'


def main() -> None:
    # raw 값 → (판매량합계, 행수, 정규화결과)
    raw_stats: dict[str, list] = defaultdict(lambda: [0, 0, ''])

    for path in sorted(p for p in EXCEL_DIR.glob('*.xlsx') if not p.name.startswith('~$')):
        print(f'  로딩: {path.name}')
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['Sheet1']
        header = [c.value for c in ws[2]]
        month_cols = [
            i for i, v in enumerate(header)
            if isinstance(v, (int, float)) and v and v > 200000
        ]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if not (len(row) > 1 and row[1]):
                continue
            raw_pt = str(row[6] or '') if len(row) > 6 else ''
            total = sum((row[i] or 0) for i in month_cols if len(row) > i and isinstance(row[i], (int, float)))
            entry = raw_stats[raw_pt]
            entry[0] += total
            entry[1] += 1
            entry[2] = normalize(raw_pt)
        wb.close()

    # Other로 분류된 것만 필터
    other_rows = {k: v for k, v in raw_stats.items() if v[2] == 'Other'}
    all_rows = raw_stats

    print()
    print('=== Other로 분류된 원본 값 ===')
    print(f"{'원본 값':<40} {'판매량':>14} {'행수':>7}")
    print('-' * 65)
    for raw, (sales, cnt, _) in sorted(other_rows.items(), key=lambda x: -x[1][0]):
        display = repr(raw) if raw != raw.strip() else raw
        print(f'{display:<40} {sales:>14,.0f} {cnt:>7,}')

    print()
    print('=== 전체 고유 원본 값 목록 ===')
    print(f"{'원본 값':<40} {'→ 정규화':<10} {'판매량':>14} {'행수':>7}")
    print('-' * 75)
    for raw, (sales, cnt, norm) in sorted(all_rows.items(), key=lambda x: -x[1][0]):
        display = repr(raw) if not raw else raw
        print(f'{display:<40} {norm:<10} {sales:>14,.0f} {cnt:>7,}')


if __name__ == '__main__':
    main()
