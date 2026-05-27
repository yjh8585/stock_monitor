"""모든 시트의 첫 몇 행 + 시트 이름을 UTF-8 파일로 저장.

PowerShell stdout cp949 인코딩 문제 회피용.
"""

from __future__ import annotations

import json
from openpyxl import load_workbook

XLSX = r"참고\손익\자료정리_월별손익.xlsx"
OUT = r"scripts\_inspect_pnl_all_sheets.json"


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    result: dict[str, dict] = {"sheets": []}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list] = []
        for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=True):
            cleaned = [
                (str(c)[:50] if c is not None else None)
                for c in row[:18]
            ]
            rows.append(cleaned)
        result["sheets"].append(
            {
                "name": name,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "head": rows,
            }
        )

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)
    print(f"saved to {OUT}")


if __name__ == "__main__":
    main()
