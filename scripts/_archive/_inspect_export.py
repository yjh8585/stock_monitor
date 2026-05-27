#!/usr/bin/env python3
"""export.xlsx 구조 확인."""
import sys
from pathlib import Path
import openpyxl
from lib.bootstrap import init_script

init_script(__file__)

DOWNLOAD_DIR = Path(__file__).resolve().parent.parent / 'data' / '_hyundai_downloads'


def inspect(year: int) -> None:
  path = DOWNLOAD_DIR / f'{year}_export.xlsx'
  wb = openpyxl.load_workbook(path, data_only=True)
  ws = wb[wb.sheetnames[0]]
  print(f'\n=== {year}_export.xlsx (max_row={ws.max_row}, max_col={ws.max_column}) ===')
  for r in range(1, min(ws.max_row, 30) + 1):
    cells = []
    for col in range(1, min(ws.max_column, 16) + 1):
      v = ws.cell(r, col).value
      cells.append(f'{chr(64+col)}={v}')
    print(f'  r{r}: {" | ".join(cells)}')


if __name__ == '__main__':
  inspect(2025)
