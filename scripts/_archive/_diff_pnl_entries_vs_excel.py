"""엑셀 '연간' 시트 raw 합 vs DB pnl_entries 합 비교.

목표: pnl_entries에 합계 행(예: '전 사', '제품계')이 raw로 잘못 적재됐는지 확인.

- '연간' 시트의 컬럼: B(2)=연도, C(3)=기준(연결/별도), D(4)=실, E(5)=부문, F(6)=공장,
  G(7)=제품, H(8)=거래처, I(9)=매출, ...
- 거래처가 비어있거나 '제품계', '전 사' 같은 패턴 있는 행은 합계로 간주
"""

from __future__ import annotations

import json
from collections import defaultdict
from openpyxl import load_workbook

XLSX = r"참고\손익\자료정리_월별손익.xlsx"


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["연간"]
    print(f"연간 시트: {ws.max_row}행 × {ws.max_column}열")

    # 헤더 확인 (row 3)
    header = [c for c in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
    print(f"\n헤더 row 3: {header[:11]}")

    # 거래처 NULL/합계 패턴 식별
    AGGREGATE_PATTERNS = {"제품계", "전 사", "전사", None, "", "제품 매출 기준"}

    by_year_basis_clean: dict[tuple[int, str], dict] = defaultdict(
        lambda: {"revenue": 0.0, "rows": 0}
    )
    by_year_basis_dirty: dict[tuple[int, str], dict] = defaultdict(
        lambda: {"revenue": 0.0, "rows": 0}
    )
    customer_distrib: dict[tuple[int, str], dict[str, int]] = defaultdict(
        lambda: defaultdict(int)
    )

    for row in ws.iter_rows(min_row=4, values_only=True):
        year = row[1]  # B (0-indexed=1)
        basis = row[2]  # C
        revenue = row[8]  # I
        customer = row[7]  # H
        if year is None:
            continue
        try:
            y = int(year)
            rev = float(revenue or 0.0)
        except (TypeError, ValueError):
            continue
        b = str(basis) if basis else "(empty)"
        key = (y, b)
        cust_label = customer if customer is not None else "(NULL)"
        customer_distrib[key][cust_label] += 1

        if customer in AGGREGATE_PATTERNS:
            by_year_basis_dirty[key]["revenue"] += rev
            by_year_basis_dirty[key]["rows"] += 1
        else:
            by_year_basis_clean[key]["revenue"] += rev
            by_year_basis_clean[key]["rows"] += 1

    print("\n=== 엑셀 '연간' 시트: 연도×기준별 합계 ===")
    print(f"{'연도':<6}{'기준':<6}{'합계행 매출':>16}{'합계행 수':>10}{'개별행 매출':>16}{'개별행 수':>10}")
    keys = sorted(set(by_year_basis_clean.keys()) | set(by_year_basis_dirty.keys()))
    for k in keys:
        y, b = k
        d_clean = by_year_basis_clean.get(k, {"revenue": 0, "rows": 0})
        d_dirty = by_year_basis_dirty.get(k, {"revenue": 0, "rows": 0})
        print(
            f"{y:<6}{b:<6}"
            f"{d_dirty['revenue']:>16,.2f}{d_dirty['rows']:>10}"
            f"{d_clean['revenue']:>16,.2f}{d_clean['rows']:>10}"
        )

    # 합계행에 어떤 거래처 라벨이 있는지 출력
    print("\n=== 합계행 후보(거래처가 NULL/특수값)의 분포 ===")
    for k in sorted(customer_distrib.keys()):
        d = customer_distrib[k]
        agg_rows = {label: cnt for label, cnt in d.items() if label in AGGREGATE_PATTERNS}
        if agg_rows:
            print(f"  {k}: {agg_rows}")

    # JSON 저장
    out = "scripts/_diff_pnl_entries_vs_excel.json"
    payload = {
        "clean": {
            f"{k[0]}_{k[1]}": v for k, v in by_year_basis_clean.items()
        },
        "dirty": {
            f"{k[0]}_{k[1]}": v for k, v in by_year_basis_dirty.items()
        },
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"\n→ saved: {out}")


if __name__ == "__main__":
    main()
