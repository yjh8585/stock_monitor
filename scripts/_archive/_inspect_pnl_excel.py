"""엑셀 원본 손익 데이터 검사 — 정리_연결 시트와 연결_월 시트 구조 파악.

사용: python scripts/_inspect_pnl_excel.py
"""

from __future__ import annotations

import os
from openpyxl import load_workbook

XLSX = r"참고\손익\자료정리_월별손익.xlsx"


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    print(f"=== sheets ===")
    for s in wb.sheetnames:
        print(f"  - {s}")

    for sheet_name in ["정리_연결", "연결_월"]:
        if sheet_name not in wb.sheetnames:
            print(f"  ! sheet missing: {sheet_name}")
            continue
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name}: {ws.max_row} rows × {ws.max_column} cols ===")
        # 상위 12행 출력
        rows = list(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True))
        for i, row in enumerate(rows, start=1):
            # 빈 컬럼 제거해서 보기 좋게
            non_empty = [(j, c) for j, c in enumerate(row, start=1) if c is not None]
            print(f"  row{i:2d}: {non_empty[:15]}")


if __name__ == "__main__":
    main()
