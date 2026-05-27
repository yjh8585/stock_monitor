"""엑셀 원본 vs DB 비교 — 매출 불일치 추적.

- 정리_연결 시트의 1번 섹션 (전사 실적): 연도별 매출/영업이익 추출
- 연결_월 시트: 연도+월별로 매출 SUM → annual로 derive 후 정리_연결과 비교
- DB pnl_entries / pnl_cost_structure 와도 비교

출력은 표 형태.
"""

from __future__ import annotations

from collections import defaultdict
from openpyxl import load_workbook

XLSX = r"참고\손익\자료정리_월별손익.xlsx"


def fmt(n: float | int | None) -> str:
    if n is None:
        return "    -"
    return f"{float(n):>12,.2f}"


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)

    # ── 1. 정리_연결 시트의 "1. 전사 실적" 섹션
    ws1 = wb["정리_연결"]
    print("=== 정리_연결: 1. 전사 실적 (연도별) ===")
    print(f"{'연도':<6}{'매출':>14}{'영업이익':>14}")
    # row 6~11 (2024~2029)
    summary1: dict[int, dict[str, float]] = {}
    for row in ws1.iter_rows(min_row=6, max_row=11, values_only=True):
        year = row[4]  # 컬럼 5 (0-indexed = 4)
        revenue = row[6]  # 컬럼 7
        op_income = row[19]  # 컬럼 20
        if year is None:
            continue
        summary1[int(year)] = {"revenue": revenue or 0.0, "op": op_income or 0.0}
        print(f"{year:<6}{fmt(revenue)}{fmt(op_income)}")

    # ── 2. 연결_월 시트: 연도×월 매출 합 → 연간 합산
    ws2 = wb["연결_월"]
    print(f"\n=== 연결_월 시트 총 행 수: {ws2.max_row} ===")

    # 컬럼: B(2)=기간, C(3)=연도, D(4)=월, E(5)=회사, F(6)=실, G(7)=부문, H(8)=공장,
    #       I(9)=제품, J(10)=거래처, K(11)=매출, L(12)=재료, M(13)=%, N(14)=노무, O(15)=%,
    #       P(16)=경비, ...
    annual_by_year: dict[int, dict[str, float]] = defaultdict(
        lambda: {"revenue": 0.0, "op_income": 0.0, "rows": 0}
    )
    monthly_by_ym: dict[tuple[int, int], float] = defaultdict(float)

    for row in ws2.iter_rows(min_row=4, values_only=True):
        year = row[2]
        month = row[3]
        revenue = row[10]
        if year is None or month is None:
            continue
        # 영업이익 컬럼은 보통 마지막 쪽. 컬럼 25 부근에 있을 가능성. 일단 매출만.
        try:
            y = int(year)
            m = int(month)
            rev = float(revenue or 0.0)
        except (TypeError, ValueError):
            continue
        annual_by_year[y]["revenue"] += rev
        annual_by_year[y]["rows"] += 1
        monthly_by_ym[(y, m)] += rev

    print("\n=== 연결_월 시트 연도별 매출 합 (월별 → 연간 derive) ===")
    print(f"{'연도':<6}{'매출 합':>16}{'행 수':>8}")
    for y in sorted(annual_by_year.keys()):
        d = annual_by_year[y]
        print(f"{y:<6}{fmt(d['revenue'])}{int(d['rows']):>8}")

    print("\n=== 연결_월 2026 1~3월 매출 ===")
    for m in (1, 2, 3):
        v = monthly_by_ym.get((2026, m), 0.0)
        print(f"  2026-{m:02d}: {fmt(v)}")
    print(f"  합계      : {fmt(sum(monthly_by_ym.get((2026, m), 0.0) for m in (1, 2, 3)))}")


if __name__ == "__main__":
    main()
