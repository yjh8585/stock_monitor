#!/usr/bin/env python3
"""Step 8c: PDF 전체 페이지 텍스트 + 모델 엑셀 전체 row dump."""
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = PROJECT_ROOT / 'data' / '_kia_audit_excel'
PDF_DIR = PROJECT_ROOT / 'data' / '_kia_audit_pdf'


def dump_full_xlsx(path: Path):
  print(f'\n=== FULL DUMP {path.name} ===')
  wb = openpyxl.load_workbook(path, data_only=True)
  ws = wb[wb.sheetnames[0]]
  print(f'  rows={ws.max_row} cols={ws.max_column}')
  for r in range(1, ws.max_row + 1):
    cells = []
    for c in range(1, min(ws.max_column + 1, 19)):
      v = ws.cell(r, c).value
      if v is None: cells.append('')
      elif isinstance(v, str): cells.append(v[:18])
      else: cells.append(str(v)[:10])
    if any(c.strip() for c in cells):
      print(f'    r{r:03d}: {" | ".join(cells)}')
  wb.close()


def dump_pdf_all(path: Path):
  print(f'\n========= FULL PDF {path.name} =========')
  from pypdf import PdfReader
  r = PdfReader(str(path))
  for i in range(len(r.pages)):
    txt = (r.pages[i].extract_text() or '').strip()
    if len(txt) < 50:
      continue
    print(f'  === p{i+1} (chars={len(txt)}) ===')
    print(txt[:2500])
    print()


# Model 파일 풀 덤프 (Domestic + Export 두 section)
dump_full_xlsx(EXCEL_DIR / 'kia_2025_model.xlsx')

# Export 파일 (region별)
dump_full_xlsx(EXCEL_DIR / 'kia_2025_export.xlsx')

# PDF — 4분기 (재무 데이터 있을 가능성 더 큼)
dump_pdf_all(PDF_DIR / 'kia_2025_4q_business.pdf')
