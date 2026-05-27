"""KG 2025년 엑셀 1건 다운로드해서 헤더/구조 정찰. 작업 후 삭제."""
import sys
import io
import os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
from playwright.sync_api import sync_playwright
import openpyxl

DEST = Path('data/_kg_downloads')
DEST.mkdir(parents=True, exist_ok=True)

with sync_playwright() as p:
  browser = p.chromium.launch()
  ctx = browser.new_context(accept_downloads=True)
  page = ctx.new_page()
  page.goto('https://www.kg-mobility.com/cm/ir-data/sales-performance', wait_until='networkidle', timeout=60000)

  for year in (2025, 2023):
    print(f'=== year={year} ===')
    # div.sale-item이 "YYYY년"을 포함하는 것 안의 다운로드 버튼
    selector = f'div.sale-item:has-text("{year}년") button.btn.file-down'
    btn = page.locator(selector).first
    try:
      with page.expect_download(timeout=30000) as dl_info:
        btn.click()
      dl = dl_info.value
      filename = dl.suggested_filename
      print(f'  suggested_filename: {filename}')
      dest_path = DEST / f'_probe_{year}_{filename}'
      dl.save_as(str(dest_path))
      size = dest_path.stat().st_size
      print(f'  saved: {dest_path} ({size} bytes)')

      # openpyxl로 시트 구조 살핌
      wb = openpyxl.load_workbook(dest_path, data_only=True)
      for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'  sheet="{sheet_name}" rows={ws.max_row} cols={ws.max_column}')
        # 처음 20행 출력 (각 행의 처음 15개 셀)
        for r in range(1, min(ws.max_row, 25) + 1):
          row_vals = []
          for c in range(1, min(ws.max_column, 18) + 1):
            v = ws.cell(r, c).value
            if v is None:
              s = ''
            else:
              s = str(v)
              if len(s) > 15:
                s = s[:12] + '...'
            row_vals.append(s)
          print(f'    r{r}: ' + ' | '.join(row_vals))
        print()
      wb.close()
    except Exception as e:
      print(f'  ERROR: {e}')
  ctx.close()
  browser.close()
