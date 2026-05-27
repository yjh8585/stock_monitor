#!/usr/bin/env python3
"""미국/유럽 현지 판매 엑셀 audit — 구조·dimension 확인 (전체)."""
from pathlib import Path

import openpyxl

from lib.bootstrap import init_script

init_script(__file__)

DEST_DIR = Path(__file__).resolve().parent.parent / 'data' / '_hyundai_audit_retail'


def dump(p: Path) -> None:
  print(f'\n========== {p.name} ==========')
  wb = openpyxl.load_workbook(p, data_only=True, read_only=False)
  for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'[Sheet: {sn}] dim={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}')
    for r in range(1, ws.max_row + 1):
      row = []
      for c in range(1, min(ws.max_column + 1, 20)):
        v = ws.cell(r, c).value
        if v is None:
          row.append('.')
        else:
          s = str(v)
          if len(s) > 22:
            s = s[:20] + '..'
          row.append(s)
      print(f'  r{r:2d}: {" | ".join(row)}')


if __name__ == '__main__':
  # 유럽만 전체 dump (시장 점유율/Total 위치 확인)
  for p in sorted(DEST_DIR.glob('2024_유럽*.xlsx')):
    dump(p)
    break
