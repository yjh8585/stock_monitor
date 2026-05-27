#!/usr/bin/env python3
"""모든 연도 factory.xlsx Grand Total vs parser 차이 추적.

각 plant Total을 Excel에서 읽어 parser sum과 비교 → 누락 plant 식별.
"""
import sys
from pathlib import Path
import openpyxl

from lib.bootstrap import init_script

init_script(__file__)

from collect_hyundai_sales import (  # noqa: E402
  DOWNLOAD_DIR,
  parse_factory_excel,
)


def excel_grand_total(year: int) -> int:
  path = DOWNLOAD_DIR / f'{year}_factory.xlsx'
  wb = openpyxl.load_workbook(path, data_only=True)
  ws = wb[wb.sheetnames[0]]
  for r in range(1, ws.max_row + 1):
    sb = str(ws.cell(r, 2).value or '').strip()
    if sb.lower().startswith('grand total'):
      v = ws.cell(r, 16).value
      try:
        return int(v) if v else 0
      except (TypeError, ValueError):
        return 0
  return 0


def excel_plant_totals(year: int) -> list[tuple[int, str, int]]:
  """각 plant Total 추출 (r, plant_name, total).

  Walk: B에 plant 이름(non-region/non-Total/non-Sub-total)이 등장한 row,
        뒤이어 'Total' row의 P 컬럼이 plant total.
  """
  path = DOWNLOAD_DIR / f'{year}_factory.xlsx'
  wb = openpyxl.load_workbook(path, data_only=True)
  ws = wb[wb.sheetnames[0]]
  plants = []
  current_plant = None
  for r in range(1, ws.max_row + 1):
    sb = str(ws.cell(r, 2).value or '').strip()
    sc = str(ws.cell(r, 3).value or '').strip()
    p_total = ws.cell(r, 16).value
    if sb.lower().startswith('grand total'):
      break
    if sb and not sc and sb not in ('Domestic', 'Export', 'Sub-total', 'Total'):
      current_plant = sb
    elif sb == 'Total' and not sc and current_plant:
      try:
        plants.append((r, current_plant, int(p_total or 0)))
      except (TypeError, ValueError):
        plants.append((r, current_plant, 0))
      current_plant = None
  return plants


def main() -> None:
  for year in [2021, 2022, 2023, 2024, 2025, 2026]:
    grand = excel_grand_total(year)
    plants = excel_plant_totals(year)
    plant_sum = sum(t for _, _, t in plants)

    rows = parse_factory_excel(DOWNLOAD_DIR / f'{year}_factory.xlsx', year)
    parser_sum = sum(r['sales_units'] for r in rows)

    print(f'\n=== {year} ===')
    print(f'  Excel Grand Total:  {grand:,}')
    print(f'  Excel plant Totals: {plant_sum:,} (n={len(plants)})')
    print(f'  Parser sum:         {parser_sum:,}')
    print(f'  Grand - Parser:     {grand - parser_sum:+,}')

    # plant별 parser sum
    by_plant = {}
    for r in rows:
      by_plant[r['factory']] = by_plant.get(r['factory'], 0) + r['sales_units']
    for r_num, plant, plant_total in plants:
      ps = by_plant.get(plant, 0)
      diff = plant_total - ps
      mark = '' if diff == 0 else '  <-- DIFF'
      print(f'    {plant:<10s}  excel_total={plant_total:>10,}  parser={ps:>10,}  diff={diff:+,}{mark}')


if __name__ == '__main__':
  main()
