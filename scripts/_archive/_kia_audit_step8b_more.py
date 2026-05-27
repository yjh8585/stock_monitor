#!/usr/bin/env python3
"""Step 8b: 더 깊은 엑셀 구조 분석 (전체 row 스캔) + PDF text 추출."""
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = PROJECT_ROOT / 'data' / '_kia_audit_excel'
PDF_DIR = PROJECT_ROOT / 'data' / '_kia_audit_pdf'


def scan_section_headers(path: Path, max_rows: int = 100):
  """엑셀의 section header(B열 텍스트 with C열 빈) 패턴 추출."""
  print(f'\n=== {path.name}: section header scan ===')
  wb = openpyxl.load_workbook(path, data_only=True)
  for sn in wb.sheetnames[:1]:
    ws = wb[sn]
    print(f'  sheet={sn} rows={ws.max_row} cols={ws.max_column}')
    # 모든 row 의 B+C 만 확인 — 빈 row 또는 section/모델 row 구분
    sections = []
    for r in range(1, ws.max_row + 1):
      a = str(ws.cell(r, 1).value or '').strip()
      b = str(ws.cell(r, 2).value or '').strip()
      c = str(ws.cell(r, 3).value or '').strip()
      d = str(ws.cell(r, 4).value or '').strip()
      total_col = None
      # 토탈 컬럼 (대략 6열 부근)
      tv = ws.cell(r, 6).value
      if tv is not None:
        try: total_col = int(tv)
        except: pass
      if a or b or c or d:
        # 모델 행이 아닐 가능성 (예: 'Domestic', 'U.S. Plant', section header)
        if (a and not c and not d) or (b and not c and not d) or (c and 'Sub' in c) or (c and 'Total' in c):
          sections.append((r, a[:35], b[:35], c[:35], d[:35]))
    print(f'  potential sections (B/A header):')
    for s in sections[:50]:
      print(f'    r{s[0]:03d}: A="{s[1]}" B="{s[2]}" C="{s[3]}" D="{s[4]}"')
  wb.close()


# Model 파일의 전체 section 구조 (Domestic + Export)
scan_section_headers(EXCEL_DIR / 'kia_2025_model.xlsx')

# Factory 파일의 plant section 구조
scan_section_headers(EXCEL_DIR / 'kia_2025_factory.xlsx')

# Export 파일의 region 구조
scan_section_headers(EXCEL_DIR / 'kia_2025_export.xlsx')

# 'other' = retail (총 13 sheet, 월별 + Total)
scan_section_headers(EXCEL_DIR / 'kia_2025_other.xlsx')


# PDF 내용 첫 페이지 텍스트 (목차 / 손익 위치 확인)
print('\n\n========= PDF samples =========')
try:
  from pypdf import PdfReader
  for fp in [PDF_DIR / 'kia_2025_3q_business.pdf', PDF_DIR / 'kia_2025_4q_business.pdf']:
    if not fp.exists():
      continue
    print(f'\n=== {fp.name} ({fp.stat().st_size/1024:.0f} KB) ===')
    r = PdfReader(str(fp))
    n = len(r.pages)
    print(f'  pages: {n}')
    for i in range(min(n, 8)):
      txt = (r.pages[i].extract_text() or '').strip()
      print(f'  --- p{i+1} (chars={len(txt)}) ---')
      print(txt[:1200])
      print()
except ImportError:
  print('pypdf not installed — run pip install pypdf')
except Exception as e:
  print(f'  ERR: {e}')
