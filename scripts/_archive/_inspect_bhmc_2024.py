#!/usr/bin/env python3
"""2024 factory.xlsx BHMC section 구조 확인."""
import sys
from pathlib import Path
import openpyxl
from lib.bootstrap import init_script

init_script(__file__)

DOWNLOAD_DIR = Path(__file__).resolve().parent.parent / 'data' / '_hyundai_downloads'


def inspect() -> None:
  path = DOWNLOAD_DIR / '2024_factory.xlsx'
  wb = openpyxl.load_workbook(path, data_only=True)
  ws = wb[wb.sheetnames[0]]
  print(f'=== 2024_factory.xlsx (max_row={ws.max_row}) - BHMC section ===')
  in_bhmc = False
  bhmc_start = None
  for r in range(1, ws.max_row + 1):
    sb = str(ws.cell(r, 2).value or '').strip()
    sc = str(ws.cell(r, 3).value or '').strip()
    if sb == 'BHMC' and not sc:
      in_bhmc = True
      bhmc_start = r
    elif in_bhmc and sb in ('HMI','HAOS','HMMA','HMGMA','HMMC','HMMR','HMB','HMMI','HTBC','KMX','Others','Russia','Vietnam','Singapore','CKD','HMTR','Grand Total'):
      break
    if in_bhmc:
      p = ws.cell(r, 16).value
      print(f'  r{r}: B="{sb}" C="{sc}" P(Total)={p}')


if __name__ == '__main__':
  inspect()
