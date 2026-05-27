"""엑셀 '연간' 시트의 'GM 직수출' 행이 어느 sil에 속하는지 확인."""

from __future__ import annotations

from collections import defaultdict
from openpyxl import load_workbook

XLSX = r"참고\손익\자료정리_월별손익.xlsx"


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["연간"]

    # 컬럼: B(2)=연도, C(3)=기준, D(4)=실, E(5)=부문, F(6)=공장, G(7)=제품, H(8)=거래처, I(9)=매출
    # 0-indexed: 1, 2, 3, 4, 5, 6, 7, 8

    by_year_sil: dict[tuple[int, str], dict] = defaultdict(
        lambda: {"revenue": 0.0, "rows": 0, "products": []}
    )

    for row in ws.iter_rows(min_row=4, values_only=True):
        year = row[1]
        sil = row[3]
        customer = row[7]
        revenue = row[8]
        product = row[6]
        if customer != "GM 직수출":
            continue
        try:
            y = int(year)
            rev = float(revenue or 0.0)
        except (TypeError, ValueError):
            continue
        key = (y, str(sil) if sil else "(empty)")
        by_year_sil[key]["revenue"] += rev
        by_year_sil[key]["rows"] += 1
        by_year_sil[key]["products"].append(f"{product}:{rev:.2f}")

    print("=== 엑셀 '연간' 시트의 'GM 직수출' 행: 연도 × sil ===")
    for key in sorted(by_year_sil.keys()):
        y, sil = key
        d = by_year_sil[key]
        print(f"  {y} / {sil:>5}: {d['rows']:>3}행, 매출 {d['revenue']:>12,.2f}")
        for p in d["products"][:5]:
            print(f"    - {p}")


if __name__ == "__main__":
    main()
