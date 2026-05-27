#!/usr/bin/env python3
"""2021 factory.xlsx 'Others' section 구조."""
import sys
from pathlib import Path
import openpyxl
from lib.bootstrap import init_script

init_script(__file__)

DOWNLOAD_DIR = Path(__file__).resolve().parent.parent / 'data' / '_hyundai_downloads'


def inspect(year: int) -> None:
  path = DOWNLOAD_DIR / f'{year}_factory.xlsx'
  wb = openpyxl.load_workbook(path, data_only=True)
  ws = wb[wb.sheetnames[0]]
  print(f'\n=== {year}_factory.xlsx - Others section ===')
  in_others = False
  for r in range(1, ws.max_row + 1):
    sb = str(ws.cell(r, 2).value or '').strip()
    sc = str(ws.cell(r, 3).value or '').strip()
    if sb == 'Others' and not sc:
      in_others = True
    elif in_others and sb in ('HMI','HAOS','BHMC','HMMA','HMGMA','HMMC','HMMR','HMB','HMMI','HTBC','KMX','Russia','Vietnam','Singapore','CKD','HMTR','Grand Total'):
      break
    if in_others:
      p = ws.cell(r, 16).value
      print(f'  r{r}: B="{sb}" C="{sc}" P(Total)={p}')


if __name__ == '__main__':
  for y in [2021, 2022]:
    inspect(y)
