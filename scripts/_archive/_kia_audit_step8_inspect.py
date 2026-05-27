#!/usr/bin/env python3
"""Step 8: 샘플 엑셀/PDF 구조 분석."""
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = PROJECT_ROOT / 'data' / '_kia_audit_excel'
PDF_DIR = PROJECT_ROOT / 'data' / '_kia_audit_pdf'


def inspect_xlsx(path: Path):
  print(f'\n=== {path.name} ({path.stat().st_size/1024:.0f} KB) ===')
  try:
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f'  sheets: {wb.sheetnames}')
    for sn in wb.sheetnames:
      ws = wb[sn]
      print(f'  --- sheet "{sn}" (rows={ws.max_row}, cols={ws.max_column}) ---')
      # 첫 15행, 첫 18열
      for r in range(1, min(ws.max_row + 1, 16)):
        row = []
        for c in range(1, min(ws.max_column + 1, 19)):
          v = ws.cell(r, c).value
          if v is None:
            row.append('')
          elif isinstance(v, str):
            row.append(v[:25])
          else:
            row.append(str(v)[:15])
        # 빈 행 skip
        if not any(row):
          continue
        # tab으로 구분
        print(f'    r{r:02d}: {" | ".join(row)}')
    wb.close()
  except Exception as e:
    print(f'  ERR: {e}')


def inspect_pdf(path: Path, pages: int = 10):
  print(f'\n=== {path.name} ({path.stat().st_size/1024:.0f} KB) ===')
  try:
    from pypdf import PdfReader
  except ImportError:
    print('  pypdf not installed')
    return
  try:
    r = PdfReader(str(path))
    n = len(r.pages)
    print(f'  pages: {n}')
    for i in range(min(n, pages)):
      txt = r.pages[i].extract_text() or ''
      txt = txt.strip()
      print(f'  --- page {i+1} (chars={len(txt)}) ---')
      # 앞 1500자만
      print(txt[:1500])
      print()
  except Exception as e:
    print(f'  ERR: {e}')


# 엑셀 1개씩 (한국어)
for label in ['model', 'factory', 'export']:
  inspect_xlsx(EXCEL_DIR / f'kia_2025_{label}.xlsx')

# 영문 비교 (model만)
inspect_xlsx(EXCEL_DIR / 'kia_2025_en_model.xlsx')

# 'other' 226KB — 별도 검토
inspect_xlsx(EXCEL_DIR / 'kia_2025_other.xlsx')
