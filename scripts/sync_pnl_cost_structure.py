#!/usr/bin/env python3
"""손익 엑셀의 '비용비율' 시트 → Supabase pnl_cost_structure 적재.

비용비율 시트 헤더(1행):
  연도 | 연간/월 | 계획/실적 | 계정분류 | 계정명 | 밸류(백만원)

데이터 2행부터 시작.
"""
import argparse
import sys
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')

sys.path.insert(0, str(Path(__file__).parent))
from lib.db import upsert_rows  # noqa: E402
from lib.management_excel import resolve_excel_path  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402

def _latest_excel() -> Path:
  return resolve_excel_path()


EXCEL_PATH = _latest_excel()
TABLE = 'pnl_cost_structure'
SHEET = '비용비율'
CONFLICT_COLS = 'period_year,period_kind,period_month,kind,account'
DATA_START_ROW = 2

KIND_MAP = {'실적': 'actual', '계획': 'plan'}


def parse_period(raw_period: Any) -> tuple[str, int] | None:
  """'연간' → ('annual', 0). 1..12 → ('monthly', N)."""
  if raw_period is None:
    return None
  if isinstance(raw_period, (int, float)) and not isinstance(raw_period, bool):
    m = int(raw_period)
    if 1 <= m <= 12:
      return ('monthly', m)
    return None
  s = str(raw_period).strip()
  if s == '연간':
    return ('annual', 0)
  return None


def main() -> int:
  parser = argparse.ArgumentParser(description='비용비율 시트 → Supabase pnl_cost_structure 적재')
  parser.add_argument('--dry-run', action='store_true', help='실제 upsert 없이 파싱·요약만')
  parser.add_argument('--revalidate-prod', action='store_true',
                      help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                           '로컬 수동 실행 시 프로덕션 stale 방지용')
  args = parser.parse_args()

  if not EXCEL_PATH.exists():
    logger.error(f'엑셀 파일 없음: {EXCEL_PATH}')
    return 1

  logger.info(f'엑셀 로드: {EXCEL_PATH}')
  wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
  try:
    ws = wb[SHEET]
    # 헤더 검증
    expected = ['연도', '연간/월', '계획/실적', '계정분류', '계정명', '밸류(백만원)']
    actual = [str(ws.cell(1, c).value or '').strip() for c in range(1, 7)]
    if actual != expected:
      logger.error(f'[{SHEET}] 헤더 불일치: 기대 {expected}, 실제 {actual}')
      return 2

    rows: list[dict[str, Any]] = []
    skipped = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
      raw_year, raw_period, raw_kind, category, account, value = row[:6]
      if raw_year is None or value is None:
        skipped += 1
        continue
      period = parse_period(raw_period)
      if period is None:
        skipped += 1
        continue
      kind = KIND_MAP.get(str(raw_kind).strip())
      if kind is None:
        skipped += 1
        continue
      rows.append({
        'period_year': int(raw_year),
        'period_kind': period[0],
        'period_month': period[1],
        'kind': kind,
        'category': str(category).strip(),
        'account': str(account).strip(),
        'value_mwon': float(value),
      })
  finally:
    wb.close()

  logger.info(f'적재 대상 {len(rows)}행 (SKIP {skipped})')
  if args.dry_run:
    logger.success('dry-run 완료')
    return 0
  try:
    n = upsert_rows(TABLE, rows, conflict_cols=CONFLICT_COLS)
    logger.success(f'pnl_cost_structure upsert 완료: {n}행')
    if args.revalidate_prod:
      revalidate_prod_for_tables([TABLE])
    return 0
  except Exception as e:
    logger.error(f'upsert 실패: {e}')
    return 3


if __name__ == '__main__':
  sys.exit(main())
