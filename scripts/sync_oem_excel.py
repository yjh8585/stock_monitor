"""MarkLines OEM 판매량 Excel 자동 다운로드 + 변경 감지 + DB 동기화.

처리 흐름
--------
1. MARKLINES_COOKIE 환경변수 + requests.Session 으로 인증
2. https://www.marklines.com/en/vehicle_sales/search?rf=left_menu GET
3. HTML 파싱 → "Latest month" 링크의 href = Excel 다운로드 URL
4. Excel 임시 파일로 다운로드 (스트리밍)
5. 구조 검증 — header 7번째 컬럼이 'PowerTrain' 인지 확인 (다른 export 오인 방지)
6. MD5 비교 → 기존 참고/oem 판매량/MarkLines_sales_data_en.xlsx 와 동일 시 EXIT 0
7. 다르면 파일 교체 → import_oem_sales.main() 호출 → DB upsert (멱등)

종료 코드
--------
0: 정상 (변경 없음 또는 업데이트 성공)
1: MARKLINES_COOKIE 미설정 / 쿠키 만료
2: Excel 다운로드 링크 파싱 실패
3: Excel 다운로드 HTTP 오류
4: 다운로드 파일 구조 불일치 (잘못된 export — 로컬 파일 보존)
5: import_oem_sales.main() 실패

사용법
-----
  python scripts/sync_oem_excel.py
"""
import hashlib
import os
import shutil
import sys
import tempfile
from pathlib import Path
from urllib.parse import urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')

sys.path.insert(0, str(Path(__file__).parent))
import import_oem_sales  # noqa: E402

EXCEL_DIR = Path(__file__).resolve().parents[1] / '참고' / 'oem 판매량'
TARGET_FILE = EXCEL_DIR / 'MarkLines_sales_data_en.xlsx'
SEARCH_URL = 'https://www.marklines.com/en/vehicle_sales/search?rf=left_menu'
USER_AGENT = (
  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
  '(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'
)
REQUEST_TIMEOUT = 60
EXPECTED_HEADER_PREFIX = ('Country', 'Group', 'Maker/Brand', 'Type', 'Segment', 'Model', 'PowerTrain')


def build_session() -> requests.Session:
  """MARKLINES_COOKIE 기반 인증 세션 생성."""
  cookie = os.environ.get('MARKLINES_COOKIE', '').strip()
  if not cookie:
    logger.error('MARKLINES_COOKIE 미설정 — .env.local 에 등록 필요')
    sys.exit(1)
  s = requests.Session()
  s.headers.update({
    'User-Agent': USER_AGENT,
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9,ko;q=0.8',
    'Cookie': cookie,
    'Referer': 'https://www.marklines.com/en/',
  })
  return s


def fetch_search_page(session: requests.Session) -> str:
  """검색 페이지 HTML 반환. 쿠키 만료 시 EXIT 1."""
  r = session.get(SEARCH_URL, timeout=REQUEST_TIMEOUT, allow_redirects=True)
  if r.status_code != 200:
    logger.error(f'검색 페이지 HTTP {r.status_code} — 쿠키 만료 가능')
    sys.exit(1)
  if '/login' in r.url or 'login_form' in r.text.lower():
    logger.error('로그인 페이지로 리다이렉트 — MARKLINES_COOKIE 만료')
    sys.exit(1)
  return r.text


def find_latest_excel_url(html: str) -> str:
  """HTML에서 "Latest month" (case-insensitive) 링크의 absolute URL 반환."""
  soup = BeautifulSoup(html, 'html.parser')
  candidates: list[tuple[str, str]] = []
  for a in soup.find_all('a', href=True):
    href = a['href']
    text = a.get_text(strip=True)
    if 'latest month' in text.lower():
      candidates.append((href, text))
  if not candidates:
    logger.error('"Latest month" 링크 미발견 — MarkLines 페이지 구조 변경 가능')
    sys.exit(2)
  href, label = candidates[0]
  full_url = urljoin(SEARCH_URL, href)
  logger.info(f'Excel 링크 발견: "{label}" → {full_url}')
  return full_url


def download_excel(session: requests.Session, url: str) -> Path:
  """Excel을 임시 파일로 다운로드. 경로 반환."""
  r = session.get(url, timeout=REQUEST_TIMEOUT, stream=True, allow_redirects=True)
  if r.status_code != 200:
    logger.error(f'Excel 다운로드 HTTP {r.status_code} — URL: {url}')
    sys.exit(3)
  ctype = r.headers.get('Content-Type', '')
  if 'html' in ctype.lower():
    logger.error(f'다운로드 응답이 HTML — 인증 실패 가능 (Content-Type={ctype})')
    sys.exit(3)
  tmp = tempfile.NamedTemporaryFile(prefix='marklines_', suffix='.xlsx', delete=False)
  size = 0
  with tmp as f:
    for chunk in r.iter_content(chunk_size=64 * 1024):
      if chunk:
        f.write(chunk)
        size += len(chunk)
  logger.info(f'다운로드 완료: {size/1024/1024:.2f} MB → {tmp.name}')
  return Path(tmp.name)


def validate_excel_structure(path: Path) -> bool:
  """다운로드한 Excel이 기존 형식과 호환되는지 검증.

  - Sheet1 존재
  - row 2 의 1~7번 컬럼이 ('Country','Group','Maker/Brand','Type','Segment','Model','PowerTrain')
  - 8번째 컬럼이 YYYYMM 정수
  실패 시 잘못된 export (예: production_data) 다운받은 것 — 로컬 파일 보존.
  """
  try:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
  except Exception as e:
    logger.error(f'Excel 열기 실패: {e}')
    return False
  try:
    if 'Sheet1' not in wb.sheetnames:
      logger.error(f'Sheet1 누락. 시트 목록: {wb.sheetnames}')
      return False
    ws = wb['Sheet1']
    header = tuple(ws.cell(2, c).value for c in range(1, 8))
    if header != EXPECTED_HEADER_PREFIX:
      logger.error(
        f'헤더 불일치 — 기대: {EXPECTED_HEADER_PREFIX}, 실제: {header}'
      )
      return False
    first_month = ws.cell(2, 8).value
    if not (isinstance(first_month, (int, float)) and first_month > 200000):
      logger.error(f'8번째 컬럼이 YYYYMM이 아님: {first_month!r}')
      return False
    logger.info(f'구조 검증 통과 (시작 월={int(first_month)})')
    return True
  finally:
    wb.close()


def md5_of(path: Path) -> str:
  h = hashlib.md5()
  with path.open('rb') as f:
    for chunk in iter(lambda: f.read(64 * 1024), b''):
      h.update(chunk)
  return h.hexdigest()


def main() -> int:
  EXCEL_DIR.mkdir(parents=True, exist_ok=True)
  session = build_session()
  html = fetch_search_page(session)
  excel_url = find_latest_excel_url(html)
  tmp_path = download_excel(session, excel_url)

  try:
    if not validate_excel_structure(tmp_path):
      logger.error('다운로드 파일 구조 불일치 — 로컬 파일 보존, 종료')
      return 4

    new_md5 = md5_of(tmp_path)
    if TARGET_FILE.exists():
      old_md5 = md5_of(TARGET_FILE)
      if new_md5 == old_md5:
        logger.success(f'변경 없음 (md5={new_md5[:12]}…) — DB upsert 스킵')
        return 0
      logger.info(f'변경 감지 (old={old_md5[:12]}… new={new_md5[:12]}…)')
    else:
      logger.info(f'기존 파일 없음 — 신규 저장 (md5={new_md5[:12]}…)')

    shutil.move(str(tmp_path), str(TARGET_FILE))
    logger.info(f'파일 교체 완료: {TARGET_FILE}')

    rc = import_oem_sales.main()
    if rc != 0:
      logger.error(f'import_oem_sales.main() 실패 (rc={rc})')
      return 5
    logger.success('OEM Excel 동기화 완료')
    return 0
  finally:
    if tmp_path.exists():
      tmp_path.unlink(missing_ok=True)


if __name__ == '__main__':
  sys.exit(main())
