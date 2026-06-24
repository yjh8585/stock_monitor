#!/usr/bin/env python3
"""인원 시트(자료정리_월별손익*.xlsx '인원') → Supabase personnel_entries 적재.

인원수 비노출: 요약은 (지역·구분)별 행수·시점 커버리지·null 카운트만 출력.
사용자가 직접 실행한다. WriteSession으로 자동 revalidate('personnel_entries').

사용법
-----
  python scripts/sync_personnel.py --dry-run
  python scripts/sync_personnel.py

종료 코드
--------
0 정상
2 헤더 검증 실패
"""
import argparse
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / '.env')
load_dotenv(Path(__file__).parent.parent / '.env.local')
sys.path.insert(0, str(Path(__file__).parent))
from lib.db import WriteSession  # noqa: E402
from lib.management_excel import resolve_excel_path  # noqa: E402
from lib.revalidate import revalidate_prod_for_tables  # noqa: E402

SHEET = '인원'
TABLE = 'personnel_entries'
CONFLICT = 'region,detail,kind,period_date'
HEADER_ROW = 1
DATA_START = 2

COL = {'year': 1, 'month': 2, 'day': 3, 'region': 4, 'detail': 5, 'kind': 6, 'headcount': 7}
EXPECTED_HEADERS = {
  1: '연도', 2: '월', 3: '일',
  4: '지역', 5: '상세', 6: '구분', 7: '인원 수',
}
VALID_REGIONS = {'국내', '외주', '미국', '중국', '우즈벡', '이인텔리전스'}
VALID_KINDS = {'임원', '사무', '생산'}
BATCH_SIZE = 500


def _int(v: Any) -> int | None:
  if v is None or v == '' or isinstance(v, bool):
    return None
  if isinstance(v, (int, float)):
    return int(v)
  return None


def _txt(v: Any) -> str:
  return '' if v is None else str(v).strip()


def validate_headers(ws) -> list[str]:
  errs = []
  for c, exp in EXPECTED_HEADERS.items():
    actual = _txt(ws.cell(HEADER_ROW, c).value)
    if actual != exp:
      errs.append(f'  컬럼 {c}: 기대 "{exp}" 실제 "{actual}"')
  return errs


def row_to_entry(ws, r: int) -> dict[str, Any] | None:
  y = ws.cell(r, COL['year']).value
  m = ws.cell(r, COL['month']).value
  d = ws.cell(r, COL['day']).value
  if not all(isinstance(x, (int, float)) for x in (y, m, d)):
    return None
  region = _txt(ws.cell(r, COL['region']).value)
  if region not in VALID_REGIONS:
    return None
  detail = _txt(ws.cell(r, COL['detail']).value)
  kind = _txt(ws.cell(r, COL['kind']).value)
  if kind not in VALID_KINDS:
    return None
  try:
    period_date = date(int(y), int(m), int(d)).isoformat()
  except ValueError:
    return None
  return {
    'region': region,
    'detail': detail,
    'kind': kind,
    'period_date': period_date,
    'headcount': _int(ws.cell(r, COL['headcount']).value),
  }


def summarize(entries: list[dict[str, Any]]) -> None:
  """(지역·구분) 행수·시점 커버리지·null 카운트. 인원수 비노출."""
  agg = defaultdict(lambda: {'rows': 0, 'dates': set(), 'nulls': 0})
  for e in entries:
    k = (e['region'], e['kind'])
    agg[k]['rows'] += 1
    agg[k]['dates'].add(e['period_date'])
    if e['headcount'] is None:
      agg[k]['nulls'] += 1
  logger.info('--- 인원 요약 (지역·구분) — 인원수 비노출 ---')
  for k in sorted(agg.keys()):
    v = agg[k]
    logger.info(f'  {k} | rows={v["rows"]} | dates={sorted(v["dates"])} | nulls={v["nulls"]}')


def _latest_excel() -> Path:
  return resolve_excel_path()


def main() -> int:
  ap = argparse.ArgumentParser(description='인원 시트 → Supabase personnel_entries 적재')
  ap.add_argument('--dry-run', action='store_true', help='실제 upsert 없이 파싱·요약만')
  ap.add_argument('--revalidate-prod', action='store_true',
                  help='적재 후 프로덕션 캐시도 추가 무효화 (NEXT_REVALIDATE_PROD_URL). '
                       '로컬 수동 실행 시 프로덕션 stale 방지용')
  args = ap.parse_args()

  path = _latest_excel()
  logger.info(f'엑셀 로드: {path}')
  wb = openpyxl.load_workbook(path, data_only=True)
  try:
    ws = wb[SHEET]
    errs = validate_headers(ws)
    if errs:
      logger.error(f'[{SHEET}] 헤더 불일치:\n' + '\n'.join(errs))
      return 2
    entries: list[dict[str, Any]] = []
    for r in range(DATA_START, ws.max_row + 1):
      e = row_to_entry(ws, r)
      if e is not None:
        entries.append(e)
    logger.info(f'[{SHEET}] {len(entries)}행 파싱 완료')
  finally:
    wb.close()

  summarize(entries)

  if args.dry_run:
    logger.success('dry-run 완료')
    return 0
  if not entries:
    logger.warning('적재할 행 없음')
    return 0

  with WriteSession() as w:
    for i in range(0, len(entries), BATCH_SIZE):
      chunk = entries[i:i + BATCH_SIZE]
      w.table(TABLE).upsert(chunk, on_conflict=CONFLICT).execute()
  logger.success(f'personnel_entries upsert 완료: {len(entries)}행')
  if args.revalidate_prod:
    revalidate_prod_for_tables([TABLE])
  return 0


if __name__ == '__main__':
  sys.exit(main())
